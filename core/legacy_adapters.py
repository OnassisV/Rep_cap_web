"""Adaptadores de datos legacy para submenus de Gestion de Capacitacion."""

# Utilidad para detectar el anio actual en filtros por vigencia.
from datetime import datetime
# Tipado explicito para listas/diccionarios del adaptador.
from typing import Any
# Utilidad para normalizar textos (acentos, mayusculas, espacios).
import unicodedata
import re
import os
import csv
import io
import json
import ast
import logging
from pathlib import Path
import pymysql

# Reutiliza la conexion MySQL ya configurada en el backend de autenticacion.
from accounts.db import get_connection
# Importa settings para resolver rutas absolutas del proyecto Django.
from django.conf import settings

logger = logging.getLogger("core")

# Lista heredada de DNIs que no deben entrar a reportes operativos.
DNIS_EXCLUIDOS_PLANTILLA = {
    "40231243",
    "09616505",
    "25790623",
    "40773688",
    "46935574",
    "40608330",
    "10818490",
    "10525745",
    "10090934",
    "40222323",
    "42116227",
    "18113751",
    "44803184",
    "43989121",
    "41410367",
    "08977083",
    "87654321",
}


def _project_base_dir() -> Path:
    """Retorna la carpeta raiz del proyecto actual de Django."""
    return Path(getattr(settings, "BASE_DIR", Path.cwd()))


def _core_config_dir(crear: bool = False) -> Path:
    """Retorna carpeta de configuraciones internas de core (core/config)."""
    ruta = _project_base_dir() / "core" / "config"
    if crear:
        ruta.mkdir(parents=True, exist_ok=True)
    return ruta


def _actividades_fuera_dir(crear: bool = False) -> Path:
    """Retorna carpeta compartida de archivos auxiliares (Actividades_fuera)."""
    ruta = _project_base_dir() / "Actividades_fuera"
    if crear:
        ruta.mkdir(parents=True, exist_ok=True)
    return ruta


def _normalizar_texto(valor: Any) -> str:
    """Convierte cualquier valor a texto normalizado para comparaciones robustas."""
    # Convierte a string y recorta espacios en extremos.
    texto = str(valor or "").strip()
    # Elimina tildes para comparar "implementacion" y "implementacion".
    texto = unicodedata.normalize("NFKD", texto)
    texto = texto.encode("ascii", "ignore").decode("ascii")
    # Unifica en minusculas para comparaciones case-insensitive.
    return texto.lower()


def obtener_catalogo_iged_por_region() -> dict[str, list[dict[str, Any]]]:
    """Retorna regiones e IGED desde iged_s3 para selects encadenados en el registro."""
    sql = """
        SELECT
            TRIM(region) AS region,
            TRIM(`NOMBRE IGED`) AS nombre_iged,
            TRIM(`tipo IGED`) AS tipo_iged,
            `CODIGO_R` AS codigo_r
        FROM iged_s3
        WHERE region IS NOT NULL
          AND TRIM(region) <> ''
          AND `NOMBRE IGED` IS NOT NULL
          AND TRIM(`NOMBRE IGED`) <> ''
        ORDER BY region, `NOMBRE IGED`
    """

    catalogo: dict[str, list[dict[str, Any]]] = {}
    vistos: set[tuple[str, str]] = set()

    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(sql)
                rows = cursor.fetchall()
    except Exception:
        # Si la tabla no esta disponible, el formulario sigue cargando con listas vacias.
        return {}

    for row in rows:
        region = str(row.get("region") or "").strip()
        nombre_iged = str(row.get("nombre_iged") or "").strip()
        tipo_iged = str(row.get("tipo_iged") or "").strip()
        codigo_r = row.get("codigo_r")

        if not region or not nombre_iged:
            continue

        llave = (region, nombre_iged)
        if llave in vistos:
            continue
        vistos.add(llave)

        catalogo.setdefault(region, []).append(
            {
                "nombre": nombre_iged,
                "tipo": tipo_iged,
                "codigo_r": codigo_r,
            }
        )

    return catalogo


def crear_registro_capacitacion(
    payload: dict[str, Any],
    creado_por: str,
    creado_nombre: str,
) -> dict[str, Any]:
    """Crea un registro de capacitacion usando Django ORM (modelo Capacitacion)."""
    from core.models import Capacitacion

    # Extrae y normaliza campos escalares del payload.
    def _str(key: str, max_len: int = 0) -> str:
        val = str(payload.get(key) or "").strip()
        return val[:max_len] if max_len else val

    def _date(key: str):
        val = _str(key)
        if not val:
            return None
        try:
            from datetime import date as _date_cls
            # Acepta YYYY-MM-DD
            return _date_cls.fromisoformat(val)
        except (ValueError, TypeError):
            return None

    def _int(key: str):
        val = _str(key)
        if not val:
            return None
        try:
            return int(val)
        except (ValueError, TypeError):
            return None

    try:
        cap = Capacitacion.objects.create(
            # Solicitud (Paso 1)
            sol_origen_institucional=_str("sol_origen_institucional", 60),
            sol_numero_oficio=_str("sol_numero_oficio", 120),
            sol_fecha_oficio=_date("sol_fecha_oficio"),
            sol_archivo_oficio=_str("sol_archivo_oficio", 500),
            sol_region_iged=_str("sol_region_iged", 120),
            sol_iged_nombre=_str("sol_iged_nombre", 250),
            sol_es_replica=_str("sol_es_replica", 10),
            sol_tiene_matriz=_str("sol_tiene_matriz", 10),
            sol_tiene_diagnostico=_str("sol_tiene_diagnostico", 10),
            sol_responde_desempeno=_str("sol_responde_desempeno", 10),
            # Identificacion
            cap_nombre=_str("cap_nombre", 255),
            cap_codigo=_str("cap_codigo", 120),
            cap_tipo=_str("cap_tipo", 100),
            cap_estrategia=_str("cap_estrategia", 255),
            cap_prioridad=_str("cap_prioridad", 30),
            cap_anio=_int("cap_anio") or datetime.now().year,
            cap_direccion=_str("cap_direccion", 120),
            pob_tipo=_str("pob_tipo", 120),
            pob_ambito=_str("pob_ambito", 120),
            # Estado
            cap_estado=Capacitacion.Estado.FORMULADA,
            paso_actual=1,
            # Metadata
            creado_por=str(creado_por or "").strip() or "sistema",
            creado_nombre=str(creado_nombre or "").strip(),
        )
        return {"ok": True, "id": cap.pk}
    except Exception as error:
        return {"ok": False, "error": str(error)}


def _parsear_anio(valor: Any) -> int | None:
    """Intenta convertir un valor de anio a entero; retorna None si falla."""
    try:
        return int(str(valor).strip())
    except Exception:
        return None


def obtener_filas_oferta_formativa() -> list[dict[str, Any]]:
    """Obtiene filas de capacitaciones desde cap_capacitaciones con aliases compatibles."""
    query = (
        "SELECT cap_codigo AS codigo, cap_anio AS anio, cap_estado AS condicion,"
        " creado_nombre AS especialista_cargo, cap_tipo AS tipo_proceso_formativo,"
        " cap_nombre AS denominacion_proceso_formativo"
        " FROM cap_capacitaciones"
    )
    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(query)
                return list(cursor.fetchall())
    except Exception:
        return []


def filtrar_capacitaciones_para_usuario(
    filas: list[dict[str, Any]],
    role_effective: str,
    display_name: str,
    username: str,
    excluir_sincronicas: bool = False,
) -> list[dict[str, Any]]:
    """Aplica filtros de rol y condicion permitida, similar a plantillas.py."""
    # Normaliza rol para comparar sin ambiguedad.
    rol_normalizado = _normalizar_texto(role_effective)

    # Define identificadores del usuario para comparar contra especialista_cargo.
    propietarios_validos = {
        _normalizar_texto(display_name),
        _normalizar_texto(username),
    }

    # Condiciones habilitadas en el flujo legacy de gestion.
    condiciones_permitidas = {
        "cerrado",
        "implementacion",
        "en implementacion",
    }

    filas_filtradas: list[dict[str, Any]] = []
    for fila in filas:
        # Permite excluir procesos sincronicos del modulo de gestion regular.
        if excluir_sincronicas:
            tipo_proceso = _normalizar_texto(fila.get("tipo_proceso_formativo", ""))
            if "sincron" in tipo_proceso:
                continue

        # Evalua condicion de estado de la capacitacion.
        condicion = _normalizar_texto(fila.get("condicion", ""))
        if condicion not in condiciones_permitidas:
            continue

        # Si el rol es usuario estandar, solo ve sus propias capacitaciones.
        if rol_normalizado == "usuario estandar":
            especialista = _normalizar_texto(fila.get("especialista_cargo", ""))
            if especialista not in propietarios_validos:
                continue

        # Conserva anio parseado para filtros de interfaz.
        fila_copia = dict(fila)
        fila_copia["anio_num"] = _parsear_anio(fila.get("anio"))
        filas_filtradas.append(fila_copia)

    # Ordena por codigo para mantener consistencia visual.
    filas_filtradas.sort(key=lambda x: str(x.get("codigo", "")).strip())
    return filas_filtradas


def aplicar_filtro_anio(
    filas: list[dict[str, Any]],
    anio_param: str,
) -> tuple[list[int], int | None, list[dict[str, Any]]]:
    """Resuelve anios disponibles y aplica filtro seleccionado."""
    # Extrae y ordena anios validos en orden descendente.
    anios_disponibles = sorted(
        {fila["anio_num"] for fila in filas if fila.get("anio_num") is not None},
        reverse=True,
    )

    # Si no hay anios validos, retorna datos originales.
    if not anios_disponibles:
        return [], None, filas

    # Intenta usar anio recibido por querystring.
    anio_solicitado = _parsear_anio(anio_param)
    if anio_solicitado not in anios_disponibles:
        # Fallback: prioriza anio actual si existe, sino el mas reciente.
        anio_actual = datetime.now().year
        anio_solicitado = anio_actual if anio_actual in anios_disponibles else anios_disponibles[0]

    # Filtra filas por anio resuelto.
    filas_anio = [fila for fila in filas if fila.get("anio_num") == anio_solicitado]
    return anios_disponibles, anio_solicitado, filas_anio


def obtener_resumen_estandares(codigos: list[str]) -> dict[str, dict[str, Any]]:
    """Obtiene resumen por codigo desde tabla estandares_calidad."""
    # Si no hay codigos, evita consulta innecesaria.
    if not codigos:
        return {}

    # Limpia codigos vacios y conserva orden de entrada.
    codigos_limpios = [str(c or "").strip() for c in codigos if str(c or "").strip()]
    if not codigos_limpios:
        return {}

    try:
        # Construye placeholders dinamicos para consulta IN parametrizada.
        placeholders = ", ".join(["%s"] * len(codigos_limpios))
        query = f"""
            SELECT
                codigo,
                COUNT(*) AS respuestas_totales,
                COUNT(DISTINCT capitulo) AS capitulos_completados,
                MAX(fecha_guardado) AS ultima_actualizacion
            FROM estandares_calidad
            WHERE codigo IN ({placeholders})
            GROUP BY codigo
        """

        # Ejecuta consulta y transforma filas en diccionario por codigo.
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(query, codigos_limpios)
                filas = list(cursor.fetchall())

        resumen: dict[str, dict[str, Any]] = {}
        for fila in filas:
            codigo = str(fila.get("codigo", "")).strip()
            if not codigo:
                continue
            resumen[codigo] = {
                "respuestas_totales": int(fila.get("respuestas_totales") or 0),
                "capitulos_completados": int(fila.get("capitulos_completados") or 0),
                "ultima_actualizacion": fila.get("ultima_actualizacion"),
            }
        return resumen
    except Exception:
        # Si falla la lectura (tabla inexistente o error de conexion), retorna vacio.
        return {}


def construir_resumen_estandares_por_capacitacion(
    capacitaciones: list[dict[str, Any]],
    resumen_estandares: dict[str, dict[str, Any]],
    total_capitulos: int = 4,
) -> list[dict[str, Any]]:
    """Combina oferta filtrada con completitud de estandares por codigo."""
    filas_resultado: list[dict[str, Any]] = []

    for cap in capacitaciones:
        # Lee identificador unico del proceso formativo.
        codigo = str(cap.get("codigo", "")).strip()
        resumen = resumen_estandares.get(codigo, {})

        # Extrae indicadores con fallback a cero.
        capitulos = int(resumen.get("capitulos_completados", 0))
        respuestas = int(resumen.get("respuestas_totales", 0))

        # Calcula porcentaje de completitud sobre 4 capitulos (modelo legacy).
        if total_capitulos > 0:
            porcentaje = round((capitulos / total_capitulos) * 100, 1)
        else:
            porcentaje = 0.0

        # Acumula fila final para render de tabla en Django.
        filas_resultado.append(
            {
                "codigo": codigo,
                "proceso": f"{cap.get('tipo_proceso_formativo', '')} {cap.get('denominacion_proceso_formativo', '')}".strip(),
                "especialista": cap.get("especialista_cargo", ""),
                "anio": cap.get("anio_num"),
                "capitulos_completados": capitulos,
                "total_capitulos": total_capitulos,
                "respuestas_totales": respuestas,
                "porcentaje_completitud": porcentaje,
                "ultima_actualizacion": resumen.get("ultima_actualizacion"),
            }
        )

    # Ordena para mostrar primero mayor completitud y luego codigo.
    filas_resultado.sort(
        key=lambda x: (-float(x.get("porcentaje_completitud", 0)), str(x.get("codigo", "")))
    )
    return filas_resultado


def extraer_id_capacitacion(codigo: str) -> str:
    """Extrae id numerico simple desde codigo compuesto (ej: 25001I-288 -> 288)."""
    # Limpia texto de entrada.
    codigo_limpio = str(codigo or "").strip()
    if not codigo_limpio:
        return ""

    # Busca segmento numerico final despues de guion.
    match = re.search(r"-\s*(\d+)\s*$", codigo_limpio)
    if match:
        return match.group(1)

    # Para codigos sin guion, replica comportamiento legacy de plantillas.py:
    # usa los ultimos 3 caracteres (ej: 26001X -> 01X).
    if len(codigo_limpio) >= 3:
        return codigo_limpio[-3:]
    return codigo_limpio


def _a_float(valor: Any) -> float:
    """Convierte un valor a float tolerando None/textos vacios."""
    try:
        return float(valor)
    except Exception:
        return 0.0


def _a_int(valor: Any) -> int:
    """Convierte un valor a int de forma segura."""
    try:
        return int(float(valor))
    except Exception:
        return 0


def _normalizar_dni(valor: Any) -> str:
    """Normaliza un DNI a formato de digitos (8 digitos cuando aplica)."""
    # Extrae solo caracteres numericos para tolerar formatos variados.
    digitos = re.sub(r"\D+", "", str(valor or ""))
    if not digitos:
        return ""
    # Para DNIs peruanos estandar, rellena con ceros a la izquierda.
    if len(digitos) <= 8:
        return digitos.zfill(8)
    return digitos


def leer_retiros_manual(codigo: str) -> list[str]:
    """Lee lista de DNIs de retiros manuales desde cap_retiros_manual en MySQL."""
    codigo = str(codigo or "").strip()
    if not codigo:
        return []
    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT dni FROM cap_retiros_manual WHERE codigo = %s ORDER BY dni",
                    [codigo],
                )
                return [row["dni"] for row in cursor.fetchall()]
    except Exception:
        logger.exception("Error leyendo retiros manuales para %s", codigo)
        return []


def agregar_retiros_manual(codigo: str, dnis_raw: list[str]) -> tuple[bool, int]:
    """Agrega DNIs a cap_retiros_manual y retorna cantidad nueva insertada."""
    codigo = str(codigo or "").strip()
    if not codigo:
        return False, 0

    nuevos = {_normalizar_dni(dni) for dni in dnis_raw}
    nuevos = {dni for dni in nuevos if dni}
    if not nuevos:
        return True, 0

    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                insertados = 0
                for dni in sorted(nuevos):
                    cursor.execute(
                        "INSERT IGNORE INTO cap_retiros_manual (codigo, dni) VALUES (%s, %s)",
                        [codigo, dni],
                    )
                    insertados += cursor.rowcount
            connection.commit()
        return True, insertados
    except Exception:
        logger.exception("Error agregando retiros manuales para %s", codigo)
        return False, 0


def eliminar_retiro_manual(codigo: str, dni: str) -> bool:
    """Elimina un DNI puntual de cap_retiros_manual."""
    codigo = str(codigo or "").strip()
    dni = _normalizar_dni(dni)
    if not codigo or not dni:
        return False

    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(
                    "DELETE FROM cap_retiros_manual WHERE codigo = %s AND dni = %s",
                    [codigo, dni],
                )
            connection.commit()
        return True
    except Exception:
        logger.exception("Error eliminando retiro manual %s/%s", codigo, dni)
        return False


def limpiar_retiros_manual(codigo: str) -> bool:
    """Elimina todos los retiros manuales de un codigo."""
    codigo = str(codigo or "").strip()
    if not codigo:
        return False
    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(
                    "DELETE FROM cap_retiros_manual WHERE codigo = %s",
                    [codigo],
                )
            connection.commit()
        return True
    except Exception:
        logger.exception("Error limpiando retiros manuales para %s", codigo)
        return False


def obtener_metricas_seguimiento(codigo: str) -> dict[str, Any]:
    """Calcula metricas base del seguimiento usando bbdd_difoca."""
    # Codigo requerido para consultas.
    codigo = str(codigo or "").strip()
    if not codigo:
        return {
            "postulaciones": 0,
            "matriculaciones": 0,
            "participaciones": 0,
            "finalizaciones": 0,
            "certificaciones": 0,
            "retiros": 0,
            "avance_promedio": 0.0,
            "nota_promedio": 0.0,
        }

    query = """
        SELECT
            COUNT(DISTINCT dni) AS postulaciones,
            SUM(CASE WHEN estado = 2 THEN 1 ELSE 0 END) AS matriculaciones,
            SUM(CASE WHEN estado = 2 AND compromiso = 20 THEN 1 ELSE 0 END) AS participaciones,
            SUM(CASE WHEN (aprobados_certificados = 1 OR desaprobado_permanente = 1) THEN 1 ELSE 0 END) AS finalizaciones,
            SUM(CASE WHEN aprobados_certificados = 1 THEN 1 ELSE 0 END) AS certificaciones,
            SUM(CASE WHEN (retiros = 1 OR desaprobado_abandono = 1) THEN 1 ELSE 0 END) AS retiros,
            AVG(CASE WHEN avance_curso_certificacion IS NOT NULL THEN avance_curso_certificacion END) AS avance_promedio,
            AVG(CASE WHEN promedio_final_general IS NOT NULL THEN promedio_final_general END) AS nota_promedio
        FROM bbdd_difoca
        WHERE codigo = %s
    """

    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(query, (codigo,))
                row = cursor.fetchone() or {}
    except Exception:
        row = {}

    return {
        "postulaciones": _a_int(row.get("postulaciones")),
        "matriculaciones": _a_int(row.get("matriculaciones")),
        "participaciones": _a_int(row.get("participaciones")),
        "finalizaciones": _a_int(row.get("finalizaciones")),
        "certificaciones": _a_int(row.get("certificaciones")),
        "retiros": _a_int(row.get("retiros")),
        "avance_promedio": round(_a_float(row.get("avance_promedio")), 2),
        "nota_promedio": round(_a_float(row.get("nota_promedio")), 2),
    }


def construir_alertas_seguimiento(metricas: dict[str, Any]) -> list[dict[str, str]]:
    """Genera alertas operativas a partir de metricas de seguimiento."""
    # Lee metricas base.
    postulaciones = _a_int(metricas.get("postulaciones"))
    matriculaciones = _a_int(metricas.get("matriculaciones"))
    participaciones = _a_int(metricas.get("participaciones"))
    finalizaciones = _a_int(metricas.get("finalizaciones"))
    certificaciones = _a_int(metricas.get("certificaciones"))
    retiros = _a_int(metricas.get("retiros"))

    alertas: list[dict[str, str]] = []

    # Evita divisiones por cero en tasas.
    tasa_matricula = (matriculaciones / postulaciones * 100) if postulaciones else 0
    tasa_participacion = (participaciones / matriculaciones * 100) if matriculaciones else 0
    tasa_finalizacion = (finalizaciones / participaciones * 100) if participaciones else 0
    tasa_certificacion = (certificaciones / finalizaciones * 100) if finalizaciones else 0

    # Reglas operativas simples para semaforo de seguimiento.
    if postulaciones == 0:
        alertas.append({"nivel": "critico", "mensaje": "La capacitacion no tiene postulaciones registradas."})
    if matriculaciones > 0 and tasa_participacion < 70:
        alertas.append(
            {
                "nivel": "alto",
                "mensaje": f"Participacion baja ({tasa_participacion:.1f}%). Revisar seguimiento de ingreso al curso.",
            }
        )
    if participaciones > 0 and tasa_finalizacion < 70:
        alertas.append(
            {
                "nivel": "alto",
                "mensaje": f"Finalizacion baja ({tasa_finalizacion:.1f}%). Revisar actividades obligatorias.",
            }
        )
    if finalizaciones > 0 and tasa_certificacion < 70:
        alertas.append(
            {
                "nivel": "medio",
                "mensaje": f"Certificacion baja ({tasa_certificacion:.1f}%). Revisar formula de promedio.",
            }
        )
    if retiros > 0:
        alertas.append(
            {
                "nivel": "medio",
                "mensaje": f"Se detectaron {retiros} retiro(s) o abandono(s) en el proceso.",
            }
        )
    if not alertas:
        alertas.append({"nivel": "ok", "mensaje": "Sin alertas criticas para el proceso seleccionado."})
    return alertas


def _get_aula_connection() -> pymysql.connections.Connection:
    """Abre conexion a BD de Aula Virtual usando settings.AULA_DB."""
    # Lee configuracion declarada en settings del proyecto.
    db_config = settings.AULA_DB
    return pymysql.connect(
        host=db_config["host"],
        user=db_config["user"],
        password=db_config["password"],
        database=db_config["database"],
        port=int(db_config["port"]),
        autocommit=True,
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
    )


def _obtener_actividades_aula_por_curso(curso_id: int) -> dict[str, list[dict[str, Any]]]:
    """Obtiene actividades de aula por tipo para un curso numerico."""
    # Inicializa salida con listas vacias para mantener forma estable.
    result: dict[str, list[dict[str, Any]]] = {
        "Ejercicio": [],
        "Tarea": [],
        "Encuesta": [],
    }

    try:
        with _get_aula_connection() as connection:
            with connection.cursor() as cursor:
                # Ejercicios activos del curso.
                cursor.execute(
                    """
                    SELECT id, title
                    FROM c_quiz
                    WHERE c_id = %s AND active = 1
                    ORDER BY id ASC
                    """,
                    (int(curso_id),),
                )
                for row in cursor.fetchall():
                    result["Ejercicio"].append(
                        {
                            "id": str(row.get("id", "")).strip(),
                            "nombre": str(row.get("title", "")).strip(),
                        }
                    )

                # Tareas de primer nivel (sin parent) del curso.
                cursor.execute(
                    """
                    SELECT id, title
                    FROM c_student_publication
                    WHERE c_id = %s AND parent_id = 0
                    ORDER BY id ASC
                    """,
                    (int(curso_id),),
                )
                for row in cursor.fetchall():
                    result["Tarea"].append(
                        {
                            "id": str(row.get("id", "")).strip(),
                            "nombre": str(row.get("title", "")).strip(),
                        }
                    )

                # Encuestas del curso.
                cursor.execute(
                    """
                    SELECT survey_id AS id, title
                    FROM c_survey
                    WHERE c_id = %s
                    ORDER BY survey_id ASC
                    """,
                    (int(curso_id),),
                )
                for row in cursor.fetchall():
                    result["Encuesta"].append(
                        {
                            "id": str(row.get("id", "")).strip(),
                            "nombre": str(row.get("title", "")).strip(),
                        }
                    )
    except Exception:
        # Si no hay conectividad a aula, retorna listas vacias controladas.
        return result

    return result


def obtener_alertas_actividades_plataforma(codigo: str) -> dict[str, Any]:
    """Compara actividades de Aula vs estructura y lista pendientes por tipo."""
    codigo = str(codigo or "").strip()
    if not codigo:
        return {"curso_id": "", "pendientes": {}, "total": 0, "error": "Codigo vacio."}

    # Solo los codigos con sufijo numerico (ej: 25001I-288) pueden mapearse a curso aula.
    curso_id_texto = extraer_id_capacitacion(codigo)
    if not curso_id_texto.isdigit():
        return {
            "curso_id": curso_id_texto,
            "pendientes": {"Ejercicio": [], "Tarea": [], "Encuesta": []},
            "total": 0,
            "error": "Codigo sin identificador numerico de curso.",
        }

    curso_id = int(curso_id_texto)
    actividades_aula = _obtener_actividades_aula_por_curso(curso_id)

    # Obtiene estructura actual y crea mapa de codigos registrados por tipo.
    estructura = obtener_estructura_por_codigo(codigo)
    codigos_estructura_por_tipo: dict[str, set[str]] = {
        "Ejercicio": set(),
        "Tarea": set(),
        "Encuesta": set(),
    }
    for row in estructura:
        tipo = str(row.get("tipo", "")).strip().title()
        codigo_actividad = str(row.get("codigo_actividad", "")).strip()
        if tipo in codigos_estructura_por_tipo and codigo_actividad:
            codigos_estructura_por_tipo[tipo].add(codigo_actividad)

    # Filtra pendientes: estan en plataforma pero no en estructura.
    pendientes: dict[str, list[dict[str, Any]]] = {
        "Ejercicio": [],
        "Tarea": [],
        "Encuesta": [],
    }
    for tipo, items in actividades_aula.items():
        codigos_registrados = codigos_estructura_por_tipo.get(tipo, set())
        for item in items:
            item_id = str(item.get("id", "")).strip()
            if not item_id:
                continue
            if item_id not in codigos_registrados:
                pendientes[tipo].append(item)

    total = sum(len(items) for items in pendientes.values())
    return {
        "curso_id": curso_id_texto,
        "pendientes": pendientes,
        "total": total,
        "error": "",
    }


def obtener_estructura_por_codigo(codigo: str) -> list[dict[str, Any]]:
    """Lista actividades de la tabla estructura para una capacitacion."""
    # Convierte codigo compuesto a id simple de estructura/formula.
    id_cap = extraer_id_capacitacion(codigo)
    if not id_cap:
        return []

    query = """
        SELECT
            id_estructura,
            id_capacitacion,
            actividad,
            grupo,
            codigo_actividad,
            tipo,
            cumplimiento_nota,
            origen,
            inicio,
            fin,
            escala,
            obligatoria,
            enlace,
            aplica_a,
            observaciones
        FROM estructura
        WHERE id_capacitacion = %s
        ORDER BY id_estructura ASC
    """
    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(query, (id_cap,))
                return list(cursor.fetchall())
    except Exception:
        return []


# Columnas base del reporte nominal heredadas del flujo legacy.
COLUMNAS_NOMINAL_BASE: list[str] = [
    "tipo_documento",
    "dni",
    "apellidos",
    "nombres",
    "genero",
    "fecha_nacimiento",
    "email",
    "telefono_celular",
    "telefono_fijo",
    "actualizo_datos",
    "region",
    "tipo_iged",
    "codigo_iged",
    "nombre_iged",
    "nivel_puesto",
    "nombre_puesto",
    "regimen_laboral",
    "ultima conexion",
    "estado",
    "compromiso",
    "%_avance_certificacion",
    "promedio_final_general",
    "promedio_final_condicion",
    "situacion_del_participante",
    "aprobados/certificados",
    "desaprobado/permanente",
    "desaprobado/abandono",
    "ev_progreso_aprendizaje",
    "mantuvo_o_progreso",
    "progreso",
    "nivel_c._entrada",
    "nivel_c._salida",
    "obs",
]


def _normalizar_fragmento_archivo(valor: Any) -> str:
    """Limpia texto para usarlo como parte de nombre de archivo local."""
    texto = str(valor or "").strip()
    if not texto:
        return "actividad"

    # Evita caracteres no validos en nombre de archivo de Windows.
    texto = re.sub(r"[\\/:*?\"<>|]+", " ", texto)
    # Compacta espacios para no generar nombres excesivos.
    texto = re.sub(r"\s+", " ", texto).strip()
    # Limita longitud para mantener rutas legibles.
    return texto[:120] if texto else "actividad"


def _ruta_excel_actividad_fuera_desde_fila(codigo: str, fila: dict[str, Any]) -> Path:
    """Construye la ruta del Excel asociado a una actividad de origen Fuera."""
    return _actividades_fuera_dir(crear=True) / _nombre_archivo_actividad_fuera(codigo, fila)


def _nombre_archivo_actividad_fuera(codigo: str, fila: dict[str, Any]) -> str:
    """Genera nombre de archivo unico para una actividad fuera."""
    id_simple = extraer_id_capacitacion(codigo)
    id_estructura = str(fila.get("id_estructura", "") or "").strip()
    codigo_actividad = str(fila.get("codigo_actividad", "") or "").strip()
    actividad = str(fila.get("actividad", "") or "").strip()
    base = codigo_actividad if codigo_actividad else actividad
    fragmento = _normalizar_fragmento_archivo(base)
    return f"excel_{id_simple}_{id_estructura}_{fragmento}.xlsx"


def _obtener_actividad_estructura_por_id(id_estructura: int) -> dict[str, Any]:
    """Retorna una fila de estructura por id para operaciones de archivos por actividad."""
    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT
                        id_estructura,
                        id_capacitacion,
                        actividad,
                        grupo,
                        codigo_actividad,
                        cumplimiento_nota,
                        origen,
                        inicio,
                        fin,
                        escala,
                        tipo,
                        obligatoria,
                        enlace,
                        observaciones,
                        aplica_a
                    FROM estructura
                    WHERE id_estructura = %s
                    LIMIT 1
                    """,
                    (int(id_estructura),),
                )
                return cursor.fetchone() or {}
    except Exception:
        return {}


def obtener_actividades_plantilla(codigo: str, estructura: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    """Separa actividades por origen y adjunta metadata de archivo para origen Fuera."""
    codigo = str(codigo or "").strip()
    if not codigo:
        return {"plataforma": [], "fuera": []}

    # Carga ids de archivos existentes en BD de una sola vez.
    ids_fuera = [
        int(row.get("id_estructura") or 0)
        for row in estructura
        if _normalizar_texto(row.get("origen", "")) == "fuera"
    ]
    archivos_bd: dict[int, int] = {}
    if ids_fuera:
        try:
            with get_connection() as connection:
                with connection.cursor() as cursor:
                    placeholders = ",".join(["%s"] * len(ids_fuera))
                    cursor.execute(
                        f"SELECT id_estructura, tamanio FROM cap_archivos_actividad_fuera WHERE id_estructura IN ({placeholders})",
                        tuple(ids_fuera),
                    )
                    for r in cursor.fetchall():
                        archivos_bd[int(r["id_estructura"])] = int(r.get("tamanio") or 0)
        except Exception:
            pass

    plataforma: list[dict[str, Any]] = []
    fuera: list[dict[str, Any]] = []

    for row in estructura:
        origen = _normalizar_texto(row.get("origen", ""))
        if origen == "plataforma":
            plataforma.append(
                {
                    "id_estructura": int(row.get("id_estructura") or 0),
                    "actividad": str(row.get("actividad", "") or "").strip(),
                    "codigo_actividad": str(row.get("codigo_actividad", "") or "").strip(),
                    "cumplimiento_nota": str(row.get("cumplimiento_nota", "") or "").strip(),
                    "tipo": str(row.get("tipo", "") or "").strip(),
                    "grupo": str(row.get("grupo", "") or "").strip(),
                }
            )
            continue

        if origen != "fuera":
            continue

        id_est = int(row.get("id_estructura") or 0)
        existe = id_est in archivos_bd
        nombre = _nombre_archivo_actividad_fuera(codigo, row)
        fuera.append(
            {
                "id_estructura": id_est,
                "actividad": str(row.get("actividad", "") or "").strip(),
                "codigo_actividad": str(row.get("codigo_actividad", "") or "").strip(),
                "cumplimiento_nota": str(row.get("cumplimiento_nota", "") or "").strip(),
                "tipo": str(row.get("tipo", "") or "").strip(),
                "grupo": str(row.get("grupo", "") or "").strip(),
                "archivo_path": "",
                "archivo_nombre": nombre,
                "archivo_existe": existe,
                "archivo_size": archivos_bd.get(id_est, 0),
            }
        )

    return {"plataforma": plataforma, "fuera": fuera}


def guardar_excel_actividad_fuera(codigo: str, id_estructura: int, contenido_bytes: bytes) -> bool:
    """Guarda un Excel cargado para una actividad de origen Fuera en BD."""
    codigo = str(codigo or "").strip()
    if not codigo or not contenido_bytes:
        return False

    fila = _obtener_actividad_estructura_por_id(id_estructura)
    if not fila:
        return False

    # Verifica que la actividad pertenezca al codigo actualmente seleccionado.
    id_simple = extraer_id_capacitacion(codigo)
    id_fila = str(fila.get("id_capacitacion", "") or "").strip()
    if not id_simple or id_fila != id_simple:
        return False

    # Solo aplica a actividades de origen Fuera.
    if _normalizar_texto(fila.get("origen", "")) != "fuera":
        return False

    nombre = _nombre_archivo_actividad_fuera(codigo, fila)
    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO cap_archivos_actividad_fuera
                        (id_estructura, nombre_archivo, contenido, tamanio, fecha_carga)
                    VALUES (%s, %s, %s, %s, NOW())
                    ON DUPLICATE KEY UPDATE
                        nombre_archivo = VALUES(nombre_archivo),
                        contenido = VALUES(contenido),
                        tamanio = VALUES(tamanio),
                        fecha_carga = NOW()
                    """,
                    (int(id_estructura), nombre, contenido_bytes, len(contenido_bytes)),
                )
            connection.commit()
        return True
    except Exception:
        return False


def eliminar_excel_actividad_fuera(codigo: str, id_estructura: int) -> bool:
    """Elimina el Excel asociado a una actividad de origen Fuera desde BD."""
    codigo = str(codigo or "").strip()
    if not codigo:
        return False

    fila = _obtener_actividad_estructura_por_id(id_estructura)
    if not fila:
        return False

    id_simple = extraer_id_capacitacion(codigo)
    id_fila = str(fila.get("id_capacitacion", "") or "").strip()
    if not id_simple or id_fila != id_simple:
        return False

    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(
                    "DELETE FROM cap_archivos_actividad_fuera WHERE id_estructura = %s",
                    (int(id_estructura),),
                )
            connection.commit()
        return True
    except Exception:
        return False


def obtener_excel_actividad_fuera_info(codigo: str, id_estructura: int) -> dict[str, Any]:
    """Retorna metadata de descarga del Excel de una actividad de origen Fuera desde BD."""
    _empty = {"exists": False, "file_name": "", "actividad": "", "contenido": b""}
    codigo = str(codigo or "").strip()
    if not codigo:
        return _empty

    fila = _obtener_actividad_estructura_por_id(id_estructura)
    if not fila:
        return _empty

    id_simple = extraer_id_capacitacion(codigo)
    id_fila = str(fila.get("id_capacitacion", "") or "").strip()
    if not id_simple or id_fila != id_simple:
        return _empty

    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT nombre_archivo, contenido FROM cap_archivos_actividad_fuera WHERE id_estructura = %s LIMIT 1",
                    (int(id_estructura),),
                )
                row = cursor.fetchone()
                if not row:
                    return _empty
                return {
                    "exists": True,
                    "file_name": row.get("nombre_archivo", "actividad_fuera.xlsx"),
                    "actividad": str(fila.get("actividad", "") or "").strip(),
                    "contenido": row.get("contenido", b""),
                }
    except Exception:
        return _empty


def _ruta_columnas_nominal_config() -> Path:
    """Retorna ruta del JSON de configuracion nominal compatible con legacy."""
    return _core_config_dir(crear=True) / "columnas_nominal_config.json"


def _leer_config_nominal() -> dict[str, Any]:
    """Lee configuracion nominal desde JSON local con tolerancia a errores."""
    ruta = _ruta_columnas_nominal_config()
    if not ruta.exists():
        return {}
    try:
        data = json.loads(ruta.read_text(encoding="utf-8-sig"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _guardar_config_nominal(config: dict[str, Any]) -> bool:
    """Guarda configuracion nominal en disco."""
    ruta = _ruta_columnas_nominal_config()
    try:
        ruta.write_text(
            json.dumps(config, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return True
    except Exception:
        return False


def _ordenar_grupos(grupos: list[str]) -> list[str]:
    """Ordena grupos numericos primero y luego alfanumericos."""
    def key_fn(valor: str) -> tuple[int, Any]:
        texto = str(valor or "").strip()
        if texto.isdigit():
            return (0, int(texto))
        return (1, texto.lower())

    return sorted(grupos, key=key_fn)


def obtener_config_nominal_reporte(codigo: str, estructura: list[dict[str, Any]]) -> dict[str, Any]:
    """Construye configuracion de reporte nominal (columnas, titulo y grupos)."""
    codigo = str(codigo or "").strip()
    id_simple = extraer_id_capacitacion(codigo)
    if not codigo or not id_simple:
        return {
            "titulo_nominal": "Reporte Nominal",
            "columnas_disponibles": [],
            "columnas_seleccionadas": [],
            "grupos": [],
            "grupos_detalle": [],
            "nombres_grupo": {},
        }

    config = _leer_config_nominal()

    # Lista ordenada de actividades para columnas dinamicas del nominal.
    actividades: list[str] = []
    seen_actividades: set[str] = set()
    for row in estructura:
        actividad = str(row.get("actividad", "") or "").strip()
        if actividad and actividad not in seen_actividades:
            seen_actividades.add(actividad)
            actividades.append(actividad)

    # Replica ajuste legacy: no duplicar compromiso si ya existe como actividad.
    columnas_base = list(COLUMNAS_NOMINAL_BASE)
    if "compromiso" in [item.lower() for item in actividades]:
        columnas_base = [col for col in columnas_base if _normalizar_texto(col) != "compromiso"]

    columnas_disponibles = columnas_base + [
        actividad for actividad in actividades if actividad not in columnas_base
    ]

    columnas_guardadas_raw = config.get(id_simple, columnas_disponibles)
    if not isinstance(columnas_guardadas_raw, list):
        columnas_guardadas_raw = columnas_disponibles

    columnas_seleccionadas: list[str] = []
    seen_cols: set[str] = set()
    for col in columnas_guardadas_raw:
        col_txt = str(col or "").strip()
        if (
            col_txt
            and col_txt in columnas_disponibles
            and col_txt not in seen_cols
        ):
            seen_cols.add(col_txt)
            columnas_seleccionadas.append(col_txt)
    if not columnas_seleccionadas:
        columnas_seleccionadas = list(columnas_disponibles)

    nominal_titles = config.get("nominal_titles", {})
    if not isinstance(nominal_titles, dict):
        nominal_titles = {}
    titulo_nominal = str(
        nominal_titles.get(id_simple, f"Reporte Nominal {id_simple}")
    ).strip() or f"Reporte Nominal {id_simple}"

    # Detecta grupos definidos en estructura para configurar sus denominaciones.
    grupos_set: set[str] = set()
    for row in estructura:
        grupo = str(row.get("grupo", "") or "").strip()
        if grupo:
            grupos_set.add(grupo)
    grupos = _ordenar_grupos(list(grupos_set)) if grupos_set else ["General"]

    group_names_root = config.get("group_names", {})
    if not isinstance(group_names_root, dict):
        group_names_root = {}
    group_names_code = group_names_root.get(id_simple, {})
    if not isinstance(group_names_code, dict):
        group_names_code = {}

    nombres_grupo: dict[str, str] = {}
    for grupo in grupos:
        default_name = f"Grupo {grupo}" if str(grupo).strip().isdigit() else str(grupo).strip()
        nombre = str(group_names_code.get(grupo, default_name)).strip() or default_name
        nombres_grupo[grupo] = nombre

    grupos_detalle = [
        {"grupo": grupo, "nombre": nombres_grupo.get(grupo, "")}
        for grupo in grupos
    ]

    return {
        "titulo_nominal": titulo_nominal,
        "columnas_disponibles": columnas_disponibles,
        "columnas_seleccionadas": columnas_seleccionadas,
        "grupos": grupos,
        "grupos_detalle": grupos_detalle,
        "nombres_grupo": nombres_grupo,
    }


def guardar_config_nominal_reporte(
    codigo: str,
    columnas_disponibles: list[str],
    columnas_seleccionadas: list[str],
    titulo_nominal: str,
    nombres_grupo: dict[str, str],
) -> bool:
    """Guarda configuracion nominal del codigo seleccionado."""
    codigo = str(codigo or "").strip()
    id_simple = extraer_id_capacitacion(codigo)
    if not codigo or not id_simple:
        return False

    config = _leer_config_nominal()
    columnas_validas = {str(col or "").strip() for col in columnas_disponibles if str(col or "").strip()}

    seleccion_final: list[str] = []
    seen: set[str] = set()
    for col in columnas_seleccionadas:
        col_txt = str(col or "").strip()
        if col_txt and col_txt in columnas_validas and col_txt not in seen:
            seen.add(col_txt)
            seleccion_final.append(col_txt)
    if not seleccion_final:
        # Si no llega ninguna seleccion valida, conserva todas para no dejar vacio.
        seleccion_final = [col for col in columnas_disponibles if str(col or "").strip()]

    config[id_simple] = seleccion_final

    if "nominal_titles" not in config or not isinstance(config.get("nominal_titles"), dict):
        config["nominal_titles"] = {}
    config["nominal_titles"][id_simple] = str(titulo_nominal or f"Reporte Nominal {id_simple}").strip()

    if "group_names" not in config or not isinstance(config.get("group_names"), dict):
        config["group_names"] = {}
    if id_simple not in config["group_names"] or not isinstance(config["group_names"].get(id_simple), dict):
        config["group_names"][id_simple] = {}

    for grupo, nombre in (nombres_grupo or {}).items():
        grupo_txt = str(grupo or "").strip()
        if not grupo_txt:
            continue
        default_name = f"Grupo {grupo_txt}" if grupo_txt.isdigit() else grupo_txt
        config["group_names"][id_simple][grupo_txt] = str(nombre or "").strip() or default_name

    return _guardar_config_nominal(config)


def agregar_actividad_estructura(codigo: str, data: dict[str, Any]) -> bool:
    """Inserta una actividad nueva en estructura para la capacitacion indicada."""
    id_cap = extraer_id_capacitacion(codigo)
    if not id_cap:
        return False

    query = """
        INSERT INTO estructura (
            id_capacitacion,
            actividad,
            codigo_actividad,
            cumplimiento_nota,
            origen,
            inicio,
            fin,
            escala,
            tipo,
            obligatoria,
            enlace,
            observaciones,
            aplica_a,
            grupo,
            items
        ) VALUES (
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1, ''
        )
    """
    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(
                    query,
                    (
                        id_cap,
                        str(data.get("actividad", "")).strip(),
                        str(data.get("codigo_actividad", "")).strip(),
                        str(data.get("cumplimiento_nota", "Cumplimiento")).strip(),
                        str(data.get("origen", "Plataforma")).strip(),
                        data.get("inicio"),
                        data.get("fin"),
                        str(data.get("escala", "Vigesimal")).strip(),
                        str(data.get("tipo", "Ejercicio")).strip(),
                        1 if bool(data.get("obligatoria")) else 0,
                        1 if bool(data.get("enlace")) else 0,
                        str(data.get("observaciones", "")).strip(),
                        str(data.get("aplica_a", "Ambos")).strip(),
                    ),
                )
        return True
    except Exception:
        return False


def actualizar_actividad_estructura(id_estructura: int, data: dict[str, Any]) -> bool:
    """Actualiza una actividad existente en estructura."""
    query = """
        UPDATE estructura
        SET
            actividad = %s,
            codigo_actividad = %s,
            cumplimiento_nota = %s,
            origen = %s,
            inicio = %s,
            fin = %s,
            escala = %s,
            tipo = %s,
            obligatoria = %s,
            enlace = %s,
            observaciones = %s,
            aplica_a = %s
        WHERE id_estructura = %s
    """
    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(
                    query,
                    (
                        str(data.get("actividad", "")).strip(),
                        str(data.get("codigo_actividad", "")).strip(),
                        str(data.get("cumplimiento_nota", "Cumplimiento")).strip(),
                        str(data.get("origen", "Plataforma")).strip(),
                        data.get("inicio"),
                        data.get("fin"),
                        str(data.get("escala", "Vigesimal")).strip(),
                        str(data.get("tipo", "Ejercicio")).strip(),
                        1 if bool(data.get("obligatoria")) else 0,
                        1 if bool(data.get("enlace")) else 0,
                        str(data.get("observaciones", "")).strip(),
                        str(data.get("aplica_a", "Ambos")).strip(),
                        int(id_estructura),
                    ),
                )
        return True
    except Exception:
        return False


def eliminar_actividad_estructura(id_estructura: int) -> bool:
    """Elimina una actividad puntual de estructura por id."""
    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute("DELETE FROM estructura WHERE id_estructura = %s", (int(id_estructura),))
        return True
    except Exception:
        return False


def obtener_formulas_promedio(codigo: str) -> list[dict[str, Any]]:
    """Retorna formulas registradas para una capacitacion."""
    id_cap = extraer_id_capacitacion(codigo)
    if not id_cap:
        return []

    query = """
        SELECT id_formula, id_capacitacion, aplica_a, formula, fecha_registro
        FROM formula_promedio
        WHERE id_capacitacion = %s
        ORDER BY fecha_registro DESC, id_formula DESC
    """
    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(query, (id_cap,))
                return list(cursor.fetchall())
    except Exception:
        return []


def guardar_formula_promedio(codigo: str, aplica_a: str, formula: str) -> bool:
    """Inserta o actualiza formula de promedio para una capacitacion."""
    id_cap = extraer_id_capacitacion(codigo)
    if not id_cap:
        return False

    aplica_a = str(aplica_a or "Ambos").strip()
    formula = str(formula or "").strip()
    if not formula:
        return False

    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                # Verifica existencia previa por id_capacitacion + aplica_a.
                cursor.execute(
                    """
                    SELECT id_formula
                    FROM formula_promedio
                    WHERE id_capacitacion = %s AND aplica_a = %s
                    LIMIT 1
                    """,
                    (id_cap, aplica_a),
                )
                row = cursor.fetchone()
                if row and row.get("id_formula"):
                    cursor.execute(
                        """
                        UPDATE formula_promedio
                        SET formula = %s, fecha_registro = NOW()
                        WHERE id_formula = %s
                        """,
                        (formula, int(row["id_formula"])),
                    )
                else:
                    cursor.execute(
                        """
                        INSERT INTO formula_promedio (id_capacitacion, aplica_a, formula, fecha_registro)
                        VALUES (%s, %s, %s, NOW())
                        """,
                        (id_cap, aplica_a, formula),
                    )
        return True
    except Exception:
        return False


def eliminar_formula_promedio(id_formula: int) -> bool:
    """Elimina una formula de promedio por id."""
    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute("DELETE FROM formula_promedio WHERE id_formula = %s", (int(id_formula),))
        return True
    except Exception:
        return False


def _obtener_participantes_por_dni(codigo: str, dnis: list[str]) -> dict[str, dict[str, Any]]:
    """Retorna mapa dni->fila participante para el codigo indicado."""
    codigo = str(codigo or "").strip()
    dnis_limpios = [_normalizar_dni(dni) for dni in dnis if _normalizar_dni(dni)]
    if not codigo or not dnis_limpios:
        return {}

    placeholders = ", ".join(["%s"] * len(dnis_limpios))
    query = f"""
        SELECT
            dni,
            apellidos,
            nombres,
            email,
            region,
            nombre_iged,
            estado,
            retiros,
            desaprobado_abandono,
            desaprobado_permanente,
            ultimo_acceso_curso
        FROM bbdd_difoca
        WHERE codigo = %s
          AND dni IN ({placeholders})
    """
    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(query, (codigo, *dnis_limpios))
                rows = list(cursor.fetchall())
    except Exception:
        rows = []

    # Normaliza clave dni para accesos consistentes.
    return {_normalizar_dni(row.get("dni")): row for row in rows}


def obtener_retiros_por_codigo(codigo: str, limit: int = 200) -> list[dict[str, Any]]:
    """Lista retiros/abandono combinando BD y registro manual en JSON."""
    codigo = str(codigo or "").strip()
    if not codigo:
        return []

    query = """
        SELECT
            dni,
            apellidos,
            nombres,
            email,
            region,
            nombre_iged,
            estado,
            retiros,
            desaprobado_abandono,
            desaprobado_permanente,
            ultimo_acceso_curso
        FROM bbdd_difoca
        WHERE codigo = %s
          AND (
              retiros = 1
              OR desaprobado_abandono = 1
              OR estado <> 2
          )
        ORDER BY ultimo_acceso_curso DESC
        LIMIT %s
    """
    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(query, (codigo, int(limit)))
                rows_db = list(cursor.fetchall())
    except Exception:
        rows_db = []

    # Marca filas provenientes de BD para distinguir origen en la interfaz.
    for row in rows_db:
        row["origen_retiro"] = "BD"
        row["retiro_manual"] = 0

    # Integra retiros manuales guardados en core/config/retiros_<codigo>.json.
    dnis_manual = leer_retiros_manual(codigo)
    if not dnis_manual:
        return rows_db

    # Mapa de filas existentes en BD por DNI para merge sin duplicados.
    rows_por_dni = {_normalizar_dni(row.get("dni")): row for row in rows_db}
    detalle_manual = _obtener_participantes_por_dni(codigo, dnis_manual)

    for dni in dnis_manual:
        row_existente = rows_por_dni.get(dni)
        if row_existente:
            # Si ya existe por BD, solo etiqueta que tambien esta en listado manual.
            row_existente["retiro_manual"] = 1
            row_existente["origen_retiro"] = "BD + Manual"
            continue

        # Si no existe en BD por condicion de retiro, crea fila de apoyo.
        info = detalle_manual.get(dni, {})
        rows_db.append(
            {
                "dni": dni,
                "apellidos": info.get("apellidos", ""),
                "nombres": info.get("nombres", ""),
                "email": info.get("email", ""),
                "region": info.get("region", ""),
                "nombre_iged": info.get("nombre_iged", ""),
                "estado": info.get("estado"),
                "retiros": 1,
                "desaprobado_abandono": info.get("desaprobado_abandono"),
                "desaprobado_permanente": info.get("desaprobado_permanente"),
                "ultimo_acceso_curso": info.get("ultimo_acceso_curso"),
                "retiro_manual": 1,
                "origen_retiro": "Manual",
            }
        )

    # Reordena dejando primero retiros manuales y luego por ultimo acceso descendente.
    rows_db.sort(key=lambda row: str(row.get("ultimo_acceso_curso") or ""), reverse=True)
    rows_db.sort(key=lambda row: 0 if _a_int(row.get("retiro_manual")) == 1 else 1)
    return rows_db


def obtener_participantes_retiro_manual_por_codigo(codigo: str) -> list[dict[str, Any]]:
    """Retorna solo participantes del listado manual de retiros para un codigo."""
    # Normaliza codigo para evitar consultas con texto vacio.
    codigo = str(codigo or "").strip()
    if not codigo:
        return []

    # Lee DNIs manuales guardados en el JSON local de configuracion.
    dnis_manual = leer_retiros_manual(codigo)
    if not dnis_manual:
        return []

    # Busca detalle nominal de cada DNI en la base para completar nombres/region.
    detalle_por_dni = _obtener_participantes_por_dni(codigo, dnis_manual)
    rows: list[dict[str, Any]] = []

    # Conserva el orden del listado manual y rellena datos faltantes cuando aplique.
    for dni in dnis_manual:
        info = detalle_por_dni.get(dni, {})
        rows.append(
            {
                "dni": dni,
                "apellidos": info.get("apellidos", ""),
                "nombres": info.get("nombres", ""),
                "email": info.get("email", ""),
                "region": info.get("region", ""),
                "nombre_iged": info.get("nombre_iged", ""),
                "estado": info.get("estado"),
                "retiros": 1,
                "desaprobado_abandono": info.get("desaprobado_abandono"),
                "desaprobado_permanente": info.get("desaprobado_permanente"),
                "ultimo_acceso_curso": info.get("ultimo_acceso_curso"),
                "retiro_manual": 1,
                "origen_retiro": "Manual",
            }
        )

    # Ordena por ultimo acceso descendente cuando exista fecha en BD.
    rows.sort(key=lambda row: str(row.get("ultimo_acceso_curso") or ""), reverse=True)
    return rows


def obtener_resumen_satisfaccion(codigo: str) -> dict[str, Any]:
    """Construye resumen de satisfaccion global por codigo."""
    codigo = str(codigo or "").strip()
    if not codigo:
        return {"total_respuestas": 0, "por_categoria": [], "por_aspecto": []}

    query_categoria = """
        SELECT satisfaccion, COUNT(*) AS total
        FROM satisfaccion
        WHERE codigo = %s
        GROUP BY satisfaccion
        ORDER BY total DESC
    """
    query_aspecto = """
        SELECT aspectos, COUNT(*) AS total
        FROM satisfaccion
        WHERE codigo = %s
        GROUP BY aspectos
        ORDER BY total DESC
        LIMIT 10
    """
    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(query_categoria, (codigo,))
                por_categoria = list(cursor.fetchall())
                cursor.execute(query_aspecto, (codigo,))
                por_aspecto = list(cursor.fetchall())
    except Exception:
        por_categoria = []
        por_aspecto = []

    total_respuestas = sum(_a_int(item.get("total")) for item in por_categoria)
    return {
        "total_respuestas": total_respuestas,
        "por_categoria": por_categoria,
        "por_aspecto": por_aspecto,
    }


def obtener_rutas_plantilla(codigo: str) -> dict[str, Any]:
    """Obtiene rutas de archivos asociadas desde tabla plantillas."""
    codigo = str(codigo or "").strip()
    if not codigo:
        return {
            "excel": "",
            "py": "",
            "excel_existe": False,
            "py_existe": False,
            "excel_nombre": "",
            "py_nombre": "",
            "postulantes_path": "",
            "postulantes_existe": False,
            "postulantes_size": 0,
        }

    query = "SELECT EXCEL, PY FROM plantillas WHERE codigo = %s LIMIT 1"
    row: dict[str, Any] = {}
    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(query, (codigo,))
                row = cursor.fetchone() or {}
    except Exception:
        row = {}

    # Lee rutas absolutas heredadas del registro legacy.
    excel_path = str(row.get("EXCEL", "") or "").strip()
    py_path = str(row.get("PY", "") or "").strip()
    excel_nombre = Path(excel_path).name if excel_path else ""
    py_nombre = Path(py_path).name if py_path else ""
    post_info = obtener_postulantes_excel_info(codigo)
    return {
        "excel": excel_path,
        "py": py_path,
        "excel_existe": bool(excel_path and os.path.exists(excel_path)),
        "py_existe": bool(py_path and os.path.exists(py_path)),
        "excel_nombre": excel_nombre,
        "py_nombre": py_nombre,
        "postulantes_path": post_info["path"],
        "postulantes_existe": post_info["exists"],
        "postulantes_size": post_info["size_bytes"],
    }


def guardar_rutas_plantilla(codigo: str, excel_path: str, py_path: str) -> bool:
    """Inserta o actualiza rutas de plantilla en tabla plantillas."""
    codigo = str(codigo or "").strip()
    if not codigo:
        return False

    excel_path = str(excel_path or "").strip()
    py_path = str(py_path or "").strip()

    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                # Verifica si el codigo ya existe para decidir update/insert.
                cursor.execute("SELECT codigo FROM plantillas WHERE codigo = %s LIMIT 1", (codigo,))
                existe = cursor.fetchone() is not None
                if existe:
                    cursor.execute(
                        "UPDATE plantillas SET EXCEL = %s, PY = %s WHERE codigo = %s",
                        (excel_path, py_path, codigo),
                    )
                else:
                    cursor.execute(
                        "INSERT INTO plantillas (codigo, EXCEL, PY) VALUES (%s, %s, %s)",
                        (codigo, excel_path, py_path),
                    )
        return True
    except Exception:
        return False


def _ruta_postulantes_excel(codigo: str) -> Path:
    """Retorna ruta local del excel de postulantes por capacitacion."""
    id_simple = extraer_id_capacitacion(codigo)
    return _actividades_fuera_dir(crear=True) / f"postulantes_{id_simple}.xlsx"


def obtener_postulantes_excel_info(codigo: str) -> dict[str, Any]:
    """Devuelve metadata de archivo de postulantes en Actividades_fuera."""
    ruta = _ruta_postulantes_excel(codigo)
    existe = ruta.exists()
    return {
        "path": str(ruta),
        "file_name": ruta.name,
        "exists": existe,
        "size_bytes": int(ruta.stat().st_size) if existe else 0,
    }


def guardar_postulantes_excel(codigo: str, contenido_bytes: bytes) -> bool:
    """Guarda el excel de postulantes en la carpeta local compartida."""
    codigo = str(codigo or "").strip()
    if not codigo or not contenido_bytes:
        return False

    ruta = _ruta_postulantes_excel(codigo)
    try:
        ruta.write_bytes(contenido_bytes)
        return True
    except Exception:
        return False


def eliminar_postulantes_excel(codigo: str) -> bool:
    """Elimina archivo de postulantes local, si existe."""
    codigo = str(codigo or "").strip()
    if not codigo:
        return False

    ruta = _ruta_postulantes_excel(codigo)
    try:
        if ruta.exists():
            ruta.unlink()
        return True
    except Exception:
        return False


def _mean(values: list[float]) -> float:
    """Calcula promedio de una lista numerica."""
    if not values:
        return 0.0
    return sum(values) / len(values)


def _var_sample(values: list[float]) -> float:
    """Calcula varianza muestral (ddof=1) compatible con pandas.var()."""
    n = len(values)
    if n <= 1:
        return 0.0
    avg = _mean(values)
    return sum((x - avg) ** 2 for x in values) / (n - 1)


def analizar_confiabilidad_por_codigo(codigo: str) -> dict[str, Any]:
    """Calcula KR-20 sobre una matriz binaria derivada de bbdd_difoca."""
    codigo = str(codigo or "").strip()
    if not codigo:
        return {
            "ok": False,
            "error": "Codigo vacio.",
            "n_estudiantes": 0,
            "n_items": 0,
            "kr20": 0.0,
            "var_total": 0.0,
            "items": [],
            "problematicos": [],
        }

    # Consulta campos disponibles para construir matriz de items dicotomicos.
    query = """
        SELECT
            dni,
            compromiso,
            encuesta,
            ev_progreso_aprendizaje,
            mantuvo_o_progreso,
            progreso,
            aprobados_certificados,
            desaprobado_permanente,
            desaprobado_abandono,
            cuestionario_entrada,
            cuestionario_salida
        FROM bbdd_difoca
        WHERE codigo = %s
    """

    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(query, (codigo,))
                rows = list(cursor.fetchall())
    except Exception as error:
        return {
            "ok": False,
            "error": f"No se pudo leer bbdd_difoca: {error}",
            "n_estudiantes": 0,
            "n_items": 0,
            "kr20": 0.0,
            "var_total": 0.0,
            "items": [],
            "problematicos": [],
        }

    if not rows:
        return {
            "ok": False,
            "error": "No hay participantes para el codigo seleccionado.",
            "n_estudiantes": 0,
            "n_items": 0,
            "kr20": 0.0,
            "var_total": 0.0,
            "items": [],
            "problematicos": [],
        }

    # Definicion de items binarios (portado a partir de indicadores existentes).
    item_defs = [
        ("compromiso_20", "Compromiso completado (=20)", lambda r: 1 if _a_float(r.get("compromiso")) == 20 else 0),
        ("encuesta", "Encuesta respondida", lambda r: 1 if _a_float(r.get("encuesta")) >= 1 else 0),
        ("ev_progreso", "Evaluacion de progreso rendida", lambda r: 1 if _a_float(r.get("ev_progreso_aprendizaje")) >= 1 else 0),
        ("mantuvo_progreso", "Mantuvo o progreso", lambda r: 1 if _a_float(r.get("mantuvo_o_progreso")) >= 1 else 0),
        ("progreso", "Progreso logrado", lambda r: 1 if _a_float(r.get("progreso")) >= 1 else 0),
        ("certificado", "Aprobado/certificado", lambda r: 1 if _a_float(r.get("aprobados_certificados")) >= 1 else 0),
        ("sin_abandono", "Sin abandono", lambda r: 1 if _a_float(r.get("desaprobado_abandono")) == 0 else 0),
        ("sin_permanente", "Sin desaprobado permanente", lambda r: 1 if _a_float(r.get("desaprobado_permanente")) == 0 else 0),
        ("ce_aprueba", "Cuestionario entrada >= 11", lambda r: 1 if _a_float(r.get("cuestionario_entrada")) >= 11 else 0),
        ("cs_aprueba", "Cuestionario salida >= 11", lambda r: 1 if _a_float(r.get("cuestionario_salida")) >= 11 else 0),
        ("mejora_ce_cs", "Cuestionario salida > entrada", lambda r: 1 if _a_float(r.get("cuestionario_salida")) > _a_float(r.get("cuestionario_entrada")) else 0),
    ]

    # Construye matriz binaria por participante.
    matrix: list[list[float]] = []
    for row in rows:
        matrix.append([float(fn(row)) for _, _, fn in item_defs])

    n_estudiantes = len(matrix)
    n_items = len(item_defs)
    if n_estudiantes < 3 or n_items < 3:
        return {
            "ok": False,
            "error": "Se requieren al menos 3 estudiantes y 3 items para KR-20.",
            "n_estudiantes": n_estudiantes,
            "n_items": n_items,
            "kr20": 0.0,
            "var_total": 0.0,
            "items": [],
            "problematicos": [],
        }

    # Totales por participante y varianza muestral.
    totals = [sum(row) for row in matrix]
    var_total = _var_sample(totals)

    # p y q por item.
    p_values: list[float] = []
    q_values: list[float] = []
    for item_idx in range(n_items):
        col = [row[item_idx] for row in matrix]
        p = _mean(col)
        q = 1 - p
        p_values.append(p)
        q_values.append(q)

    # KR-20 principal.
    suma_pq = sum(p * q for p, q in zip(p_values, q_values))
    if var_total > 0 and n_items > 1:
        kr20 = (n_items / (n_items - 1)) * (1 - (suma_pq / var_total))
    else:
        kr20 = 0.0

    # Analisis "if item deleted".
    items_result: list[dict[str, Any]] = []
    for idx, (item_id, item_label, _) in enumerate(item_defs):
        # Matriz sin item actual.
        matrix_without = []
        for row in matrix:
            row_new = row[:idx] + row[idx + 1 :]
            matrix_without.append(row_new)

        n_items_new = len(item_defs) - 1
        if n_items_new <= 1:
            kr20_without = 0.0
        else:
            totals_new = [sum(row) for row in matrix_without]
            var_total_new = _var_sample(totals_new)
            if var_total_new > 0:
                p_new = []
                q_new = []
                for j in range(n_items_new):
                    col_new = [row[j] for row in matrix_without]
                    p_j = _mean(col_new)
                    p_new.append(p_j)
                    q_new.append(1 - p_j)
                suma_pq_new = sum(pj * qj for pj, qj in zip(p_new, q_new))
                kr20_without = (n_items_new / (n_items_new - 1)) * (1 - (suma_pq_new / var_total_new))
            else:
                kr20_without = 0.0

        impacto = kr20_without - kr20
        items_result.append(
            {
                "item_id": item_id,
                "item_label": item_label,
                "proporcion_aciertos": round(p_values[idx], 4),
                "dificultad": round(1 - p_values[idx], 4),
                "discriminacion": round(p_values[idx] * q_values[idx], 4),
                "kr20_sin_item": round(kr20_without, 4),
                "impacto": round(impacto, 4),
            }
        )

    # Ordena por mayor impacto positivo (preguntas que mas reducen confiabilidad).
    items_result.sort(key=lambda x: x.get("impacto", 0), reverse=True)
    problematicos = [item for item in items_result if float(item.get("impacto", 0)) > 0]

    return {
        "ok": True,
        "error": "",
        "n_estudiantes": n_estudiantes,
        "n_items": n_items,
        "kr20": round(kr20, 4),
        "var_total": round(var_total, 4),
        "items": items_result,
        "problematicos": problematicos,
    }


def interpretar_kr20(kr20: float) -> str:
    """Retorna interpretacion textual del KR-20."""
    kr20 = float(kr20 or 0)
    if kr20 >= 0.9:
        return "Excelente"
    if kr20 >= 0.8:
        return "Buena"
    if kr20 >= 0.7:
        return "Aceptable"
    if kr20 >= 0.6:
        return "Cuestionable"
    return "Deficiente"


def exportar_confiabilidad_csv(analisis: dict[str, Any]) -> str:
    """Serializa analisis de confiabilidad a CSV."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "item_id",
            "item_label",
            "proporcion_aciertos",
            "dificultad",
            "discriminacion",
            "kr20_sin_item",
            "impacto",
        ]
    )
    for row in analisis.get("items", []):
        writer.writerow(
            [
                row.get("item_id"),
                row.get("item_label"),
                row.get("proporcion_aciertos"),
                row.get("dificultad"),
                row.get("discriminacion"),
                row.get("kr20_sin_item"),
                row.get("impacto"),
            ]
        )
    return output.getvalue()


def obtener_certificados_detalle(codigo: str, limit: int = 5000) -> list[dict[str, Any]]:
    """Obtiene listado nominal de participantes certificados por codigo."""
    codigo = str(codigo or "").strip()
    if not codigo:
        return []

    query = """
        SELECT
            codigo,
            dni,
            apellidos,
            nombres,
            email,
            region,
            tipo_iged,
            nombre_iged,
            promedio_final_general,
            cuestionario_salida,
            ultimo_acceso_curso
        FROM bbdd_difoca
        WHERE codigo = %s
          AND aprobados_certificados = 1
        ORDER BY apellidos ASC, nombres ASC
        LIMIT %s
    """
    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(query, (codigo, int(limit)))
                return list(cursor.fetchall())
    except Exception:
        return []


def obtener_participantes_certificacion_para_emision(codigo: str, limit: int = 10000) -> list[dict[str, Any]]:
    """Retorna participantes aptos para emisión de certificados por código de curso.

    Filtros aplicados:
    - estado = 2
    - compromiso = 20
    - aprobados_certificados = 1
    """
    codigo = str(codigo or "").strip()
    if not codigo:
        return []

    query = """
        SELECT
            dni AS DNI,
            apellidos AS APELLIDOS,
            nombres AS NOMBRES,
            promedio_final_general AS NOTAS,
            nombre_puesto AS `NOMBRE DE PUESTO`,
            nombre_iged AS `NOMBRE IGED`
        FROM bbdd_difoca
        WHERE codigo = %s
          AND estado = 2
          AND compromiso = 20
          AND aprobados_certificados = 1
        ORDER BY apellidos ASC, nombres ASC
        LIMIT %s
    """
    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(query, (codigo, int(limit)))
                return list(cursor.fetchall())
    except Exception:
        return []


def obtener_participantes_certificados_lista_excel(codigo: str, limit: int = 10000) -> list[dict[str, Any]]:
    """Lista nominal de participantes certificados por codigo de curso, para descargar a Excel.

    Devuelve los campos solicitados: DNI, APELLIDOS, NOMBRES, TELEFONO, NOMBRE IGED,
    NIVEL DE PUESTO, PROMEDIO FINAL.

    Mismo filtro que `obtener_participantes_certificacion_para_emision`:
    - estado = 2
    - compromiso = 20
    - aprobados_certificados = 1
    """
    codigo = str(codigo or "").strip()
    if not codigo:
        return []

    query = """
        SELECT
            dni AS DNI,
            apellidos AS APELLIDOS,
            nombres AS NOMBRES,
            telefono_celular AS TELEFONO,
            nombre_iged AS `NOMBRE IGED`,
            nivel_puesto AS `NIVEL DE PUESTO`,
            promedio_final_general AS `PROMEDIO FINAL`
        FROM bbdd_difoca
        WHERE codigo = %s
          AND estado = 2
          AND compromiso = 20
          AND aprobados_certificados = 1
        ORDER BY apellidos ASC, nombres ASC
        LIMIT %s
    """
    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(query, (codigo, int(limit)))
                return list(cursor.fetchall())
    except Exception:
        return []


def resumir_certificados_por_region(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Agrupa certificados por region con promedio de nota final."""
    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        region = str(row.get("region", "")).strip() or "Sin region"
        nota = _a_float(row.get("promedio_final_general"))
        current = grouped.setdefault(region, {"region": region, "total": 0, "sum_nota": 0.0})
        current["total"] += 1
        current["sum_nota"] += nota

    result: list[dict[str, Any]] = []
    for _, item in grouped.items():
        total = int(item["total"])
        prom = round((item["sum_nota"] / total), 2) if total > 0 else 0.0
        result.append({"region": item["region"], "total": total, "nota_promedio": prom})

    result.sort(key=lambda x: (-int(x.get("total", 0)), str(x.get("region", ""))))
    return result


def exportar_certificados_csv(rows: list[dict[str, Any]]) -> str:
    """Serializa detalle nominal de certificados a CSV."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "codigo",
            "dni",
            "apellidos",
            "nombres",
            "email",
            "region",
            "tipo_iged",
            "nombre_iged",
            "promedio_final_general",
            "cuestionario_salida",
            "ultimo_acceso_curso",
        ]
    )
    for row in rows:
        writer.writerow(
            [
                row.get("codigo", ""),
                row.get("dni", ""),
                row.get("apellidos", ""),
                row.get("nombres", ""),
                row.get("email", ""),
                row.get("region", ""),
                row.get("tipo_iged", ""),
                row.get("nombre_iged", ""),
                row.get("promedio_final_general", ""),
                row.get("cuestionario_salida", ""),
                row.get("ultimo_acceso_curso", ""),
            ]
        )
    return output.getvalue()


def exportar_resumen_certificados_csv(rows: list[dict[str, Any]]) -> str:
    """Serializa resumen regional de certificados a CSV."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["region", "total", "nota_promedio"])
    for row in rows:
        writer.writerow([row.get("region", ""), row.get("total", 0), row.get("nota_promedio", 0)])
    return output.getvalue()


def _ruta_plantillas_generadas_dir(anio: int, crear: bool = False) -> Path:
    """Retorna carpeta local de plantillas generadas para un anio especifico."""
    ruta = _project_base_dir() / "Plantillas_generadas" / f"PF{int(anio)}"
    if crear:
        ruta.mkdir(parents=True, exist_ok=True)
    return ruta


def _ruta_metadata_plantilla_generada(codigo: str) -> Path:
    """Retorna ruta del JSON metadata de la ultima plantilla generada por codigo."""
    id_simple = extraer_id_capacitacion(codigo)
    return _core_config_dir(crear=True) / f"plantilla_generada_{id_simple}.json"


def _guardar_metadata_plantilla_generada(codigo: str, payload: dict[str, Any]) -> bool:
    """Guarda metadata de salida de la ultima plantilla generada."""
    ruta = _ruta_metadata_plantilla_generada(codigo)
    try:
        ruta.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return True
    except Exception:
        return False


def obtener_plantilla_generada_info(codigo: str) -> dict[str, Any]:
    """Retorna metadata y estado del archivo generado para la capacitacion."""
    codigo = str(codigo or "").strip()
    if not codigo:
        return {
            "exists": False,
            "path": "",
            "file_name": "",
            "size_bytes": 0,
            "generated_at": "",
            "files": [],
        }

    ruta_meta = _ruta_metadata_plantilla_generada(codigo)
    if not ruta_meta.exists():
        return {
            "exists": False,
            "path": "",
            "file_name": "",
            "size_bytes": 0,
            "generated_at": "",
            "files": [],
        }

    try:
        payload = json.loads(ruta_meta.read_text(encoding="utf-8-sig"))
    except Exception:
        payload = {}

    # Soporta metadata nueva (lista de archivos) y metadata legacy (un solo archivo).
    files_meta = payload.get("files")
    files: list[dict[str, Any]] = []
    etiquetas = {
        "main": "Plantilla generada",
        "nominal": "Reporte nominal",
        "iged": "Cumplimiento por IGED",
    }
    descargas = {
        "main": "plantilla_generada",
        "nominal": "plantilla_generada_nominal",
        "iged": "plantilla_generada_iged",
    }

    if isinstance(files_meta, list):
        for raw in files_meta:
            if not isinstance(raw, dict):
                continue
            kind = str(raw.get("kind", "main") or "main").strip().lower()
            path_text = str(raw.get("path", "") or "").strip()
            file_path = Path(path_text) if path_text else None
            exists_file = bool(file_path and file_path.exists())
            files.append(
                {
                    "kind": kind,
                    "label": str(raw.get("label", "") or etiquetas.get(kind, "Archivo generado")),
                    "path": str(file_path) if file_path else "",
                    "file_name": str(raw.get("file_name", "") or (file_path.name if file_path else "")),
                    "size_bytes": int(raw.get("size_bytes") or (file_path.stat().st_size if exists_file else 0)),
                    "exists": exists_file,
                    "download_kind": descargas.get(kind, "plantilla_generada"),
                }
            )

    if not files:
        # Compatibilidad con metadata antigua de un solo archivo.
        path_text = str(payload.get("path", "") or "").strip()
        file_path = Path(path_text) if path_text else None
        exists_file = bool(file_path and file_path.exists())
        files.append(
            {
                "kind": "main",
                "label": "Plantilla generada",
                "path": str(file_path) if file_path else "",
                "file_name": str(payload.get("file_name", "") or (file_path.name if file_path else "")),
                "size_bytes": int(payload.get("size_bytes") or (file_path.stat().st_size if exists_file else 0)),
                "exists": exists_file,
                "download_kind": "plantilla_generada",
            }
        )

        # Si existen salidas auxiliares junto al archivo principal, las incorpora
        # aunque la metadata antigua no traiga el campo "files".
        if file_path:
            stem = file_path.stem
            suffix = file_path.suffix or ".xlsx"
            for kind, label, sibling_name in [
                ("nominal", "Reporte nominal", f"{stem}_NOMINAL{suffix}"),
                ("iged", "Cumplimiento por IGED", f"{stem}_IGED{suffix}"),
            ]:
                sibling = file_path.with_name(sibling_name)
                if not sibling.exists():
                    continue
                files.append(
                    {
                        "kind": kind,
                        "label": label,
                        "path": str(sibling),
                        "file_name": sibling.name,
                        "size_bytes": int(sibling.stat().st_size),
                        "exists": True,
                        "download_kind": descargas.get(kind, "plantilla_generada"),
                    }
                )

    # Define archivo principal para compatibilidad de campos legacy.
    principal = next((f for f in files if f.get("kind") == "main"), None)
    if principal is None:
        principal = next((f for f in files if f.get("exists")), files[0] if files else {})
    existe_alguno = any(bool(f.get("exists")) for f in files)

    return {
        "exists": existe_alguno,
        "path": str(principal.get("path", "") or ""),
        "file_name": str(principal.get("file_name", "") or ""),
        "size_bytes": int(principal.get("size_bytes") or 0),
        "generated_at": str(payload.get("generated_at", "") or ""),
        "files": files,
    }


def _normalizar_header_excel(valor: Any) -> str:
    """Normaliza texto de cabecera de Excel para comparaciones robustas."""
    texto = str(valor or "").replace("_", " ").replace(".", " ").replace("/", " ")
    texto = re.sub(r"\s+", " ", texto).strip()
    return _normalizar_texto(texto)


def _leer_excel_postulantes_dni(path_excel: str) -> list[str]:
    """Lee DNI de un archivo de postulantes y retorna lista normalizada sin duplicados."""
    path_excel = str(path_excel or "").strip()
    if not path_excel or not os.path.exists(path_excel):
        return []

    try:
        # Import local para no romper si el paquete no esta instalado.
        from openpyxl import load_workbook
    except Exception:
        return []

    try:
        wb = load_workbook(path_excel, read_only=True, data_only=True)
        ws = wb.active
        filas = ws.iter_rows(min_row=1, max_row=1, values_only=True)
        headers = next(filas, ())
        if not headers:
            return []

        idx_dni = None
        for i, header in enumerate(headers):
            if _normalizar_header_excel(header) == "dni":
                idx_dni = i
                break
        if idx_dni is None:
            return []

        dnis: list[str] = []
        vistos: set[str] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or idx_dni >= len(row):
                continue
            dni = _normalizar_dni(row[idx_dni])
            if not dni or dni in vistos:
                continue
            vistos.add(dni)
            dnis.append(dni)
        return dnis
    except Exception:
        return []


def _leer_excel_fuera_dni_nota(path_excel: str) -> dict[str, Any]:
    """Lee mapa DNI->NOTA (o valor equivalente) desde Excel de actividad fuera."""
    path_excel = str(path_excel or "").strip()
    if not path_excel or not os.path.exists(path_excel):
        return {}

    try:
        from openpyxl import load_workbook
    except Exception:
        return {}

    try:
        wb = load_workbook(path_excel, read_only=True, data_only=True)
        ws = wb.active
        headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        if not headers:
            return {}

        idx_dni = None
        idx_nota = None
        for i, header in enumerate(headers):
            norm = _normalizar_header_excel(header)
            if norm == "dni" and idx_dni is None:
                idx_dni = i
            if norm in {"nota", "puntaje", "calificacion"} and idx_nota is None:
                idx_nota = i

        if idx_dni is None:
            return {}

        # Si no existe columna explicita de nota, usa la segunda columna.
        if idx_nota is None and len(headers) >= 2:
            idx_nota = 1 if idx_dni != 1 else 0

        resultado: dict[str, Any] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or idx_dni >= len(row):
                continue
            dni = _normalizar_dni(row[idx_dni])
            if not dni:
                continue

            nota = None
            if idx_nota is not None and idx_nota < len(row):
                nota = row[idx_nota]
                if isinstance(nota, str):
                    nota = nota.strip()
                    if nota == "":
                        nota = None
            resultado[dni] = nota
        return resultado
    except Exception:
        return {}


def _leer_excel_fuera_desde_bytes(contenido: bytes) -> dict[str, Any]:
    """Lee mapa DNI->NOTA desde bytes de un Excel en memoria."""
    if not contenido:
        return {}
    try:
        from io import BytesIO
        from openpyxl import load_workbook
    except Exception:
        return {}
    try:
        wb = load_workbook(BytesIO(contenido), read_only=True, data_only=True)
        ws = wb.active
        headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        if not headers:
            return {}

        idx_dni = None
        idx_nota = None
        for i, header in enumerate(headers):
            norm = _normalizar_header_excel(header)
            if norm == "dni" and idx_dni is None:
                idx_dni = i
            if norm in {"nota", "puntaje", "calificacion"} and idx_nota is None:
                idx_nota = i

        if idx_dni is None:
            return {}

        if idx_nota is None and len(headers) >= 2:
            idx_nota = 1 if idx_dni != 1 else 0

        resultado: dict[str, Any] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or idx_dni >= len(row):
                continue
            dni = _normalizar_dni(row[idx_dni])
            if not dni:
                continue

            nota = None
            if idx_nota is not None and idx_nota < len(row):
                nota = row[idx_nota]
                if isinstance(nota, str):
                    nota = nota.strip()
                    if nota == "":
                        nota = None
            resultado[dni] = nota
        return resultado
    except Exception:
        return {}


def _obtener_dnis_matriculados_por_codigo(codigo: str) -> list[str]:
    """Retorna listado de DNI con estado matriculado (estado=2) para un codigo."""
    codigo = str(codigo or "").strip()
    if not codigo:
        return []

    query = """
        SELECT DISTINCT dni
        FROM bbdd_difoca
        WHERE codigo = %s
          AND estado = 2
        ORDER BY dni ASC
    """
    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(query, (codigo,))
                rows = list(cursor.fetchall())
    except Exception:
        rows = []

    dnis: list[str] = []
    vistos: set[str] = set()
    for row in rows:
        dni = _normalizar_dni(row.get("dni"))
        if not dni or dni in vistos:
            continue
        vistos.add(dni)
        dnis.append(dni)
    return dnis


def _obtener_todos_dnis_por_codigo(codigo: str) -> list[str]:
    """Retorna todos los DNI en bbdd_difoca para un codigo (cualquier estado)."""
    codigo = str(codigo or "").strip()
    if not codigo:
        return []

    query = """
        SELECT DISTINCT dni
        FROM bbdd_difoca
        WHERE codigo = %s
        ORDER BY dni ASC
    """
    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(query, (codigo,))
                rows = list(cursor.fetchall())
    except Exception:
        rows = []

    dnis: list[str] = []
    vistos: set[str] = set()
    for row in rows:
        dni = _normalizar_dni(row.get("dni"))
        if not dni or dni in vistos:
            continue
        vistos.add(dni)
        dnis.append(dni)
    return dnis


def _obtener_dnis_matriculados_aula(codigo: str) -> list[str]:
    """Obtiene DNIs de matriculados directamente de Chamilo (course_rel_user + user).

    Replica la logica de obtener_lista_matriculados() de la app original.
    status=5 = matriculado en Chamilo, is_tutor IS NULL = excluye tutores.
    """
    curso_id = extraer_id_capacitacion(codigo)
    if not curso_id:
        return []
    try:
        curso_id_num = int(curso_id)
    except (ValueError, TypeError):
        return []

    query = """
        SELECT DISTINCT u.official_code AS dni
        FROM course_rel_user cru
        LEFT JOIN user u ON cru.user_id = u.user_id
        WHERE cru.c_id = %s
          AND cru.status = 5
          AND cru.is_tutor IS NULL
        ORDER BY u.official_code ASC
    """
    try:
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(query, (curso_id_num,))
                rows = list(cur.fetchall())
    except Exception:
        logger.exception("Error obteniendo matriculados de Chamilo para codigo=%s (c_id=%s)", codigo, curso_id)
        return []

    dnis: list[str] = []
    vistos: set[str] = set()
    for row in rows:
        dni = _normalizar_dni(row.get("dni"))
        if not dni or dni in vistos:
            continue
        vistos.add(dni)
        dnis.append(dni)
    return dnis


def _enriquecer_filas_con_sidi(filas_por_dni: dict[str, dict[str, Any]]) -> None:
    """Completa y actualiza datos demograficos desde uvw_usuarios_detalle_completo (SIDI en Railway).

    Estrategia de actualizacion:
    - Campos institucionales (region, IGED, puesto, regimen): siempre se sobreescriben con el
      valor de SIDI si este tiene datos, para reflejar cambios que el participante haya hecho
      en SIDIFOCA.
    - Campos personales (apellidos, nombres, email, telefonos, etc.): solo se rellenan si estan
      vacios en bbdd_difoca, para no sobreescribir correcciones manuales.
    """
    dnis_todos = list(filas_por_dni.keys())
    if not dnis_todos:
        return

    query = """
        SELECT
            CASE WHEN id_tipo_documento = 1 THEN 'DNI' ELSE 'OTROS' END AS tipo_documento,
            nro_documento AS dni,
            genero,
            CONCAT(COALESCE(apellido_paterno,''), ' ', COALESCE(apellido_materno,'')) AS apellidos,
            nombres,
            DATE_FORMAT(fecha_nacimiento, '%%Y-%%m-%%d') AS fecha_nacimiento,
            email,
            telefono_celular,
            telefono_fijo,
            CASE
                WHEN fec_modifica < DATE_SUB(CURDATE(), INTERVAL 6 MONTH) THEN 'No'
                ELSE 'Si'
            END AS actualizo_datos,
            SUBSTRING(descripcion_id_dre_gre, 5, 100) AS region,
            CASE
                WHEN REPLACE(descripcion_tipo_entidad, ' ', '') = 'DRE/GRE' THEN 'DRE / GRE'
                ELSE descripcion_tipo_entidad
            END AS tipo_iged,
            CASE
                WHEN REPLACE(descripcion_tipo_entidad, ' ', '') = 'DRE/GRE' THEN descripcion_id_dre_gre
                WHEN descripcion_tipo_entidad = 'UGEL' THEN descripcion_id_ugel
                ELSE NULL
            END AS nombre_iged,
            descripcion_nivel_puesto AS nivel_puesto,
            descripcion_puesto AS nombre_puesto,
            descripcion_regimen_laboral AS regimen_laboral
        FROM uvw_usuarios_detalle_completo
        WHERE nro_documento IN ({placeholders})
    """
    placeholders = ", ".join(["%s"] * len(dnis_todos))
    query = query.format(placeholders=placeholders)

    try:
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(query, dnis_todos)
                rows_sidi = list(cur.fetchall())
    except Exception:
        logger.exception("Error consultando uvw_usuarios_detalle_completo")
        return

    # Indexar por DNI normalizado.
    sidi_por_dni: dict[str, dict[str, Any]] = {}
    for row in rows_sidi:
        dni = _normalizar_dni(row.get("dni"))
        if dni:
            sidi_por_dni[dni] = row

    # Obtener codigo_iged desde iged_s3.
    codigos_iged = _obtener_mapa_codigo_iged()

    # Todos los campos se actualizan siempre desde SIDI si este tiene datos.
    campos_sobreescribir = [
        "tipo_documento", "genero", "apellidos", "nombres", "fecha_nacimiento",
        "email", "telefono_celular", "telefono_fijo",
        "actualizo_datos", "region", "tipo_iged", "nombre_iged",
        "nivel_puesto", "nombre_puesto", "regimen_laboral",
    ]
    campos_solo_vacios: list[str] = []

    enriquecidos = 0
    for dni in dnis_todos:
        sidi = sidi_por_dni.get(dni)
        if not sidi:
            continue
        fila = filas_por_dni[dni]
        enriquecidos += 1
        # Sobreescribir siempre los campos institucionales con datos de SIDI.
        for campo in campos_sobreescribir:
            valor = sidi.get(campo)
            if not _valor_vacio(valor):
                fila[campo] = valor
        # Rellenar datos personales solo si estan vacios.
        for campo in campos_solo_vacios:
            if _valor_vacio(fila.get(campo)):
                valor = sidi.get(campo)
                if not _valor_vacio(valor):
                    fila[campo] = valor
        # codigo_iged: recalcular siempre desde iged_s3 con la region/iged actualizada de SIDI.
        region = str(fila.get("region") or "").strip()
        nombre_iged = str(fila.get("nombre_iged") or "").strip()
        clave = (_normalizar_texto(region), _normalizar_texto(nombre_iged))
        codigo = codigos_iged.get(clave)
        if not _valor_vacio(codigo):
            fila["codigo_iged"] = codigo

    logger.info(
        "_enriquecer_filas_con_sidi: %d/%d DNIs actualizados con datos SIDI",
        enriquecidos, len(dnis_todos),
    )


def _obtener_mapa_codigo_iged() -> dict[tuple[str, str], str]:
    """Retorna mapa (region_norm, nombre_iged_norm) -> codigo_iged desde iged_s3."""
    try:
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT region, nombre_iged, codigo_iged FROM iged_s3")
                rows = list(cur.fetchall())
    except Exception:
        return {}

    mapa: dict[tuple[str, str], str] = {}
    for row in rows:
        region = _normalizar_texto(str(row.get("region") or "").strip())
        nombre = _normalizar_texto(str(row.get("nombre_iged") or "").strip())
        codigo = str(row.get("codigo_iged") or "").strip()
        if region and nombre and codigo:
            mapa[(region, nombre)] = codigo
    return mapa


def _obtener_filas_bbdd_por_codigo_y_dnis(codigo: str, dnis: list[str]) -> list[dict[str, Any]]:
    """Obtiene filas base de bbdd_difoca para el codigo y DNI seleccionados."""
    codigo = str(codigo or "").strip()
    if not codigo:
        return []

    campos = [
        "codigo",
        "tipo_documento",
        "dni",
        "apellidos",
        "nombres",
        "genero",
        "fecha_nacimiento",
        "telefono_celular",
        "telefono_fijo",
        "email",
        "actualizo_datos",
        "region",
        "tipo_iged",
        "codigo_iged",
        "nombre_iged",
        "nivel_puesto",
        "nombre_puesto",
        "regimen_laboral",
        "ultimo_acceso_curso",
        "estado",
        "compromiso",
        "promedio_final_general",
        "promedio_final_condicion",
        "avance_curso_certificacion",
        "situacion_participante",
        "retiros",
        "aprobados_certificados",
        "desaprobado_permanente",
        "desaprobado_abandono",
        "cuestionario_entrada",
        "cuestionario_salida",
        "encuesta",
        "ev_progreso_aprendizaje",
        "mantuvo_o_progreso",
        "progreso",
        "nivel_c_entrada",
        "nivel_c_salida",
        "obs",
    ]
    campos_sql = ", ".join(campos)

    # Carga base completa por codigo.
    # Nota: se evita filtrar DNI en SQL porque en fuentes legacy el DNI puede venir
    # con/sin ceros a la izquierda, y la comparaciÃ³n exacta rompe la coincidencia.
    query = (
        f"SELECT {campos_sql} FROM bbdd_difoca "
        "WHERE codigo = %s "
        "ORDER BY ultimo_acceso_curso DESC"
    )
    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(query, (codigo,))
                rows_all = list(cursor.fetchall())
    except Exception:
        return []

    # Si no se pasa lista de DNIs, retorna todos los registros del cÃ³digo.
    if not dnis:
        return rows_all

    # Filtrado robusto en memoria por DNI normalizado.
    dnis_objetivo = {_normalizar_dni(item) for item in dnis if _normalizar_dni(item)}
    if not dnis_objetivo:
        return rows_all
    return [
        row for row in rows_all
        if _normalizar_dni(row.get("dni")) in dnis_objetivo
    ]


# Orden base del archivo de plantilla, alineado al flujo legacy de plantillas.py.
ORDEN_COLUMNAS_PLANTILLA: list[str] = [
    "tipo_documento",
    "dni",
    "apellidos",
    "nombres",
    "genero",
    "fecha_nacimiento",
    "email",
    "telefono_celular",
    "telefono_fijo",
    "actualizo_datos",
    "region",
    "tipo_iged",
    "codigo_iged",
    "nombre_iged",
    "nivel_puesto",
    "nombre_puesto",
    "regimen_laboral",
    "ultima conexion",
    "estado",
    "compromiso",
    "%_avance_certificacion",
    "promedio_final_general",
    "promedio_final_condicion",
    "situacion_del_participante",
    "aprobados/certificados",
    "desaprobado/permanente",
    "desaprobado/abandono",
    "ev_progreso_aprendizaje",
    "mantuvo_o_progreso",
    "progreso",
    "nivel_c._entrada",
    "nivel_c._salida",
    "obs",
    "retiros",
]


def _a_float_nullable(valor: Any) -> float | None:
    """Convierte un valor a float; retorna None cuando esta vacio o no es numerico."""
    if valor is None:
        return None
    texto = str(valor).strip()
    if not texto or texto.lower() in {"nan", "none", "null"}:
        return None
    # Normaliza comas decimales para tolerar datos escritos manualmente.
    texto = texto.replace(",", ".")
    try:
        return float(texto)
    except Exception:
        return None


def _valor_vacio(valor: Any) -> bool:
    """Evalua si un valor debe tratarse como vacio para reglas de completitud."""
    if valor is None:
        return True
    if isinstance(valor, str):
        texto = valor.strip().lower()
        return texto in {"", "nan", "none", "null"}
    return False


def _normalizar_clave_formula(valor: Any) -> str:
    """Normaliza un nombre de columna/actividad para comparaciones de formula."""
    texto = str(valor or "")
    texto = texto.replace("_", " ").replace(".", " ").replace("/", " ")
    texto = re.sub(r"\s+", " ", texto).strip()
    return _normalizar_texto(texto)


def _limpiar_texto_html_simple(texto: Any) -> str:
    """Limpia artefactos HTML frecuentes en respuestas textuales de Aula."""
    valor = str(texto or "")
    if not valor:
        return ""

    # Reemplazos vistos en respuestas de track_e_attempt.
    reemplazos = {
        "&#38;lt&#38;#58;p&gt;": "",
        "&#38;lt&#38;#58;/p&gt;": "",
        "&#38;amp&#38;#58;": ":",
        "&#38;nbsp&#38;#58;": "",
        "&#58;": ":",
        "&nbsp;": " ",
        "&amp;": "&",
        "&lt;": "<",
        "&gt;": ">",
        "&quot;": "\"",
    }
    for original, nuevo in reemplazos.items():
        valor = valor.replace(original, nuevo)
    return re.sub(r"\s+", " ", valor).strip()


def _fila_base_plantilla_desde_bbdd(row: dict[str, Any]) -> dict[str, Any]:
    """Convierte fila de bbdd_difoca al esquema de columnas de plantilla."""
    # Normaliza tipo_iged para mantener forma legacy "DRE / GRE".
    tipo_iged = str(row.get("tipo_iged", "") or "").strip()
    if _normalizar_texto(tipo_iged) in {"dre/gre", "dre / gre", "dre"}:
        tipo_iged = "DRE / GRE"

    # Retiros se exporta vacio o 1 (no 0) en plantilla legacy.
    retiros = 1 if _a_int(row.get("retiros")) == 1 else None

    fila = {
        "tipo_documento": row.get("tipo_documento") or "DNI",
        "dni": _normalizar_dni(row.get("dni")),
        "apellidos": row.get("apellidos"),
        "nombres": row.get("nombres"),
        "genero": row.get("genero"),
        "fecha_nacimiento": row.get("fecha_nacimiento"),
        "email": row.get("email"),
        "telefono_celular": row.get("telefono_celular"),
        "telefono_fijo": row.get("telefono_fijo"),
        "actualizo_datos": row.get("actualizo_datos"),
        "region": row.get("region"),
        "tipo_iged": tipo_iged,
        "codigo_iged": row.get("codigo_iged"),
        "nombre_iged": row.get("nombre_iged"),
        "nivel_puesto": row.get("nivel_puesto"),
        "nombre_puesto": row.get("nombre_puesto"),
        "regimen_laboral": row.get("regimen_laboral"),
        "ultima conexion": row.get("ultimo_acceso_curso"),
        "estado": _a_int(row.get("estado")),
        "compromiso": row.get("compromiso"),
        "%_avance_certificacion": row.get("avance_curso_certificacion"),
        "promedio_final_general": row.get("promedio_final_general"),
        "promedio_final_condicion": row.get("promedio_final_condicion"),
        "situacion_del_participante": row.get("situacion_participante"),
        "aprobados/certificados": row.get("aprobados_certificados"),
        "desaprobado/permanente": row.get("desaprobado_permanente"),
        "desaprobado/abandono": row.get("desaprobado_abandono"),
        "ev_progreso_aprendizaje": row.get("ev_progreso_aprendizaje"),
        "mantuvo_o_progreso": row.get("mantuvo_o_progreso"),
        "progreso": row.get("progreso"),
        "nivel_c._entrada": row.get("nivel_c_entrada"),
        "nivel_c._salida": row.get("nivel_c_salida"),
        "obs": row.get("obs"),
        "retiros": retiros,
        # Alias heredados para formulas que usan variables con espacios.
        "cuestionario entrada": row.get("cuestionario_entrada"),
        "cuestionario salida": row.get("cuestionario_salida"),
    }
    return fila


def _fila_vacia_para_dni(dni: str) -> dict[str, Any]:
    """Crea fila base minima cuando el DNI no existe en bbdd_difoca."""
    fila = {col: None for col in ORDEN_COLUMNAS_PLANTILLA}
    fila["dni"] = _normalizar_dni(dni)
    fila["estado"] = 1
    # Alias para formulas de cuestionarios aun sin datos.
    fila["cuestionario entrada"] = None
    fila["cuestionario salida"] = None
    return fila


def _actividad_aplica_a_participante(aplica_a: str, tipo_iged: str) -> bool:
    """Determina si una actividad obligatoria aplica segun el tipo_iged del participante."""
    aplica = str(aplica_a or "Ambos").strip().lower()
    tipo = str(tipo_iged or "").strip().lower()
    if aplica in {"ambos", "todos", "general", ""}:
        return True
    if aplica in {"dre / gre", "dre/gre", "dre"}:
        return tipo == "dre / gre"
    if aplica == "ugel":
        return tipo == "ugel"
    return True


def _es_columna_binaria_en_filas(filas: list[dict[str, Any]], columna: str) -> bool:
    """Detecta si una columna se comporta como binaria estricta (0/1)."""
    valores: list[float] = []
    for fila in filas:
        val = _a_float_nullable(fila.get(columna))
        if val is None:
            continue
        valores.append(val)
    if not valores:
        return False
    return all(v in {0.0, 1.0} for v in valores)


def _actividad_completa(valor: Any, es_binaria: bool) -> bool:
    """Evalua completitud de una actividad segun su naturaleza (binaria o no)."""
    if es_binaria:
        num = _a_float_nullable(valor)
        return num is not None and num == 1.0
    return not _valor_vacio(valor)


def _codigo_curso_base(codigo: str) -> str:
    """Retorna la parte izquierda del codigo compuesto (ej: 26001I de 26001I-315)."""
    texto = str(codigo or "").strip()
    if "-" in texto:
        return texto.split("-", 1)[0].strip()
    return texto


def _obtener_mapa_actividad_plataforma(codigo: str, actividad: dict[str, Any]) -> dict[str, Any]:
    """Consulta Aula Virtual y retorna mapa DNI->valor para una actividad de plataforma."""
    curso_id = extraer_id_capacitacion(codigo)
    if not curso_id.isdigit():
        return {}

    codigo_actividad = str(actividad.get("codigo_actividad", "") or "").strip()
    if not codigo_actividad:
        return {}

    tipo = str(actividad.get("tipo", "") or "").strip().title()
    cumplimiento = _normalizar_texto(actividad.get("cumplimiento_nota", "Cumplimiento"))
    enlace = bool(actividad.get("enlace"))
    curso_base = _codigo_curso_base(codigo)
    resultado: dict[str, Any] = {}

    try:
        with _get_aula_connection() as connection:
            with connection.cursor() as cursor:
                if tipo == "Ejercicio":
                    if enlace:
                        # Para respuestas textuales (enlace=1), recupera la ultima respuesta del participante.
                        cursor.execute(
                            """
                            SELECT
                                u.official_code AS dni,
                                MAX(a.answer) AS nota
                            FROM track_e_attempt a
                            LEFT JOIN user u ON a.user_id = u.user_id
                            WHERE a.c_id = %s
                              AND a.question_id IN (
                                  SELECT question_id
                                  FROM c_quiz_rel_question
                                  WHERE exercice_id = %s
                                    AND c_id = %s
                              )
                            GROUP BY u.official_code
                            """,
                            (int(curso_id), int(codigo_actividad), int(curso_id)),
                        )
                        for row in cursor.fetchall():
                            dni = _normalizar_dni(row.get("dni"))
                            if not dni:
                                continue
                            resultado[dni] = _limpiar_texto_html_simple(row.get("nota"))
                    else:
                        # Para ejercicios de nota usa el maximo resultado registrado.
                        cursor.execute(
                            """
                            SELECT
                                u.official_code AS dni,
                                MAX(t.exe_result) AS nota
                            FROM track_e_exercises t
                            LEFT JOIN user u ON t.exe_user_id = u.user_id
                            WHERE t.c_id = %s
                              AND t.exe_exo_id = %s
                            GROUP BY t.exe_user_id, u.official_code
                            """,
                            (int(curso_id), int(codigo_actividad)),
                        )
                        for row in cursor.fetchall():
                            dni = _normalizar_dni(row.get("dni"))
                            if not dni:
                                continue
                            resultado[dni] = row.get("nota")

                elif tipo == "Tarea":
                    if enlace and cumplimiento == "cumplimiento":
                        # Cuando hay enlace, devuelve URL de descarga del ultimo archivo por usuario.
                        cursor.execute(
                            """
                            SELECT
                                u.official_code AS dni,
                                MAX(p.id) AS pub_id
                            FROM c_student_publication p
                            LEFT JOIN user u ON p.user_id = u.user_id
                            WHERE p.c_id = %s
                              AND p.active = 1
                              AND p.contains_file = 1
                              AND p.parent_id = %s
                            GROUP BY p.user_id, u.official_code
                            """,
                            (int(curso_id), int(codigo_actividad)),
                        )
                        for row in cursor.fetchall():
                            dni = _normalizar_dni(row.get("dni"))
                            pub_id = row.get("pub_id")
                            if not dni or pub_id is None:
                                continue
                            resultado[dni] = (
                                "https://aula.edutalentos.pe/main/work/download.php"
                                f"?id={int(pub_id)}&cidReq={curso_base}"
                            )
                    else:
                        # En tareas sin enlace se usa nota/cumplimiento desde qualification.
                        cursor.execute(
                            """
                            SELECT
                                u.official_code AS dni,
                                CASE
                                    WHEN MAX(p.qualification) = 0 THEN 1
                                    ELSE MAX(p.qualification)
                                END AS nota
                            FROM c_student_publication p
                            LEFT JOIN user u ON p.user_id = u.user_id
                            WHERE p.c_id = %s
                              AND p.active = 1
                              AND p.contains_file = 1
                              AND p.parent_id = %s
                            GROUP BY p.user_id, u.official_code
                            """,
                            (int(curso_id), int(codigo_actividad)),
                        )
                        for row in cursor.fetchall():
                            dni = _normalizar_dni(row.get("dni"))
                            if not dni:
                                continue
                            resultado[dni] = row.get("nota")

                elif tipo == "Encuesta":
                    # Encuestas se consideran de cumplimiento por presencia de respuesta.
                    cursor.execute(
                        """
                        SELECT DISTINCT
                            u.official_code AS dni
                        FROM c_survey_answer s
                        LEFT JOIN user u ON s.user = u.user_id
                        WHERE s.c_id = %s
                          AND s.survey_id = %s
                        """,
                        (int(curso_id), int(codigo_actividad)),
                    )
                    for row in cursor.fetchall():
                        dni = _normalizar_dni(row.get("dni"))
                        if not dni:
                            continue
                        resultado[dni] = 1
    except Exception:
        return {}

    return resultado


def _obtener_mapa_actividad_fuera(codigo: str, actividad: dict[str, Any]) -> dict[str, Any]:
    """Lee el Excel de actividad fuera desde BD y retorna mapa DNI->valor."""
    id_estructura = int(actividad.get("id_estructura") or 0)
    if not id_estructura:
        return {}
    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT contenido FROM cap_archivos_actividad_fuera WHERE id_estructura = %s LIMIT 1",
                    (id_estructura,),
                )
                row = cursor.fetchone()
                if not row or not row.get("contenido"):
                    return {}
                return _leer_excel_fuera_desde_bytes(row["contenido"])
    except Exception:
        return {}


def _aplicar_estado_matricula_y_retiros(
    codigo: str,
    filas_por_dni: dict[str, dict[str, Any]],
) -> None:
    """Actualiza estado/compromiso/retiros segun Chamilo y retiros manuales.

    Reglas:
    - En Chamilo (matriculado)           -> estado=2
    - No en Chamilo                      -> estado=1
    - En retiros manuales (siempre gana) -> estado=3, retiros=1
    """
    matriculados_aula = set(_obtener_dnis_matriculados_aula(codigo))
    retiros_manual = set(leer_retiros_manual(codigo))

    for dni, fila in filas_por_dni.items():
        # Chamilo es fuente de verdad para matricula.
        if dni in matriculados_aula:
            estado = 2
        else:
            estado = 1

        # Retiros manuales siempre ganan.
        if dni in retiros_manual:
            estado = 3
            fila["retiros"] = 1

        fila["estado"] = estado
        if estado == 2 and _valor_vacio(fila.get("compromiso")):
            fila["compromiso"] = 20
        # Normalizacion legacy: compromiso 1 o 20 se exporta como 20.
        compromiso_num = _a_float_nullable(fila.get("compromiso"))
        if compromiso_num in {1.0, 20.0}:
            fila["compromiso"] = 20
        if estado != 3 and _valor_vacio(fila.get("desaprobado/abandono")):
            # Limpia ruido de estados antiguos cuando no corresponde a retiro manual.
            fila["desaprobado/abandono"] = None


def _aplicar_actividades_a_filas(
    codigo: str,
    estructura: list[dict[str, Any]],
    filas_por_dni: dict[str, dict[str, Any]],
) -> list[str]:
    """Agrega columnas de actividades y carga sus valores (plataforma/fuera) sobre cada fila."""
    actividades_ordenadas = sorted(
        estructura,
        key=lambda x: int(x.get("id_estructura") or 0),
    )
    actividades_export: list[str] = []

    for actividad in actividades_ordenadas:
        nombre_col = str(actividad.get("actividad", "") or "").strip()
        if not nombre_col:
            continue

        origen = _normalizar_texto(actividad.get("origen", ""))
        cumplimiento = _normalizar_texto(actividad.get("cumplimiento_nota", "Cumplimiento"))
        enlace = bool(actividad.get("enlace"))
        mapa_valores: dict[str, Any] = {}

        if origen == "plataforma":
            mapa_valores = _obtener_mapa_actividad_plataforma(codigo, actividad)
        elif origen == "fuera":
            mapa_valores = _obtener_mapa_actividad_fuera(codigo, actividad)

        # Inicializa columna en todas las filas para mantener estructura estable.
        for fila in filas_por_dni.values():
            if nombre_col not in fila:
                fila[nombre_col] = None

        for dni, fila in filas_por_dni.items():
            valor = mapa_valores.get(dni)
            if cumplimiento == "nota":
                fila[nombre_col] = valor if not _valor_vacio(valor) else None
                continue

            # Cumplimiento con enlace conserva el texto/URL de evidencia.
            if enlace and not _valor_vacio(valor):
                fila[nombre_col] = valor
                continue

            # Cumplimiento clasico se registra como 1 cuando existe respuesta/evidencia.
            fila[nombre_col] = 1 if dni in mapa_valores else None

        actividades_export.append(nombre_col)

    return actividades_export


def _recalcular_avance_certificacion(
    estructura: list[dict[str, Any]],
    filas: list[dict[str, Any]],
) -> None:
    """Recalcula % de avance usando actividades obligatorias y reglas por aplica_a."""
    obligatorias = [
        row
        for row in estructura
        if _a_int(row.get("obligatoria")) == 1 and str(row.get("actividad", "")).strip()
    ]
    if not obligatorias:
        for fila in filas:
            if _valor_vacio(fila.get("%_avance_certificacion")):
                fila["%_avance_certificacion"] = None
        return

    nombres_obligatorios = [str(row.get("actividad", "")).strip() for row in obligatorias]
    mapa_binario = {
        col: _es_columna_binaria_en_filas(filas, col)
        for col in nombres_obligatorios
    }

    for fila in filas:
        tipo_iged = str(fila.get("tipo_iged", "") or "").strip()
        aplicables = [
            row
            for row in obligatorias
            if _actividad_aplica_a_participante(str(row.get("aplica_a", "Ambos")), tipo_iged)
        ]
        if not aplicables:
            fila["%_avance_certificacion"] = None
            continue

        total = len(aplicables)
        completas = 0
        for row in aplicables:
            nombre_col = str(row.get("actividad", "")).strip()
            if _actividad_completa(fila.get(nombre_col), mapa_binario.get(nombre_col, False)):
                completas += 1
        fila["%_avance_certificacion"] = (completas / total) if total > 0 else None


def _safe_eval_formula_ast(node: ast.AST, values: dict[str, float]) -> float:
    """Evalua un AST aritmetico restringido para evitar ejecucion de codigo arbitrario."""
    if isinstance(node, ast.Expression):
        return _safe_eval_formula_ast(node.body, values)
    if isinstance(node, ast.Constant):
        if isinstance(node.value, (int, float)):
            return float(node.value)
        raise ValueError("Constante no numerica en formula.")
    if isinstance(node, ast.Name):
        return float(values.get(node.id, 0.0))
    if isinstance(node, ast.UnaryOp):
        val = _safe_eval_formula_ast(node.operand, values)
        if isinstance(node.op, ast.UAdd):
            return +val
        if isinstance(node.op, ast.USub):
            return -val
        raise ValueError("Operador unario no permitido.")
    if isinstance(node, ast.BinOp):
        left = _safe_eval_formula_ast(node.left, values)
        right = _safe_eval_formula_ast(node.right, values)
        if isinstance(node.op, ast.Add):
            return left + right
        if isinstance(node.op, ast.Sub):
            return left - right
        if isinstance(node.op, ast.Mult):
            return left * right
        if isinstance(node.op, ast.Div):
            return 0.0 if right == 0 else (left / right)
        if isinstance(node.op, ast.Pow):
            return left ** right
        raise ValueError("Operador binario no permitido.")
    raise ValueError("Nodo de formula no permitido.")


def _evaluar_formula_en_fila(formula: str, fila: dict[str, Any]) -> float:
    """Evalua una formula textual usando los valores de actividades de una fila."""
    formula = str(formula or "").strip()
    if not formula:
        return 0.0

    # Normaliza comas decimales a puntos (ej: 0,3 → 0.3).
    # Solo reemplaza comas entre dígitos para no afectar otros usos.
    formula = re.sub(r"(\d),(\d)", r"\1.\2", formula)

    # Mapa normalizado de columnas disponibles en la fila.
    indice_columnas = {
        _normalizar_clave_formula(col): col
        for col in fila.keys()
    }

    valores_vars: dict[str, float] = {}
    contador = 0

    def _reemplazar(match: re.Match[str]) -> str:
        nonlocal contador
        actividad = str(match.group(1) or "").strip()
        key_norm = _normalizar_clave_formula(actividad)
        col_real = indice_columnas.get(key_norm)
        valor = _a_float(fila.get(col_real)) if col_real else 0.0
        var_name = f"v{contador}"
        valores_vars[var_name] = valor
        contador += 1
        return var_name

    # Reemplaza solo variables de actividad declaradas entre comillas simples.
    expr = re.sub(r"'([^']+)'", _reemplazar, formula)
    try:
        tree = ast.parse(expr, mode="eval")
        return float(_safe_eval_formula_ast(tree, valores_vars))
    except Exception:
        return 0.0


def _obtener_formulas_por_aplica_a(codigo: str) -> dict[str, str]:
    """Retorna la ultima formula registrada por cada alcance de aplica_a."""
    formulas = obtener_formulas_promedio(codigo)
    resultado: dict[str, str] = {}
    for row in formulas:
        aplica = str(row.get("aplica_a", "") or "").strip()
        if not aplica:
            continue
        aplica_norm = _normalizar_texto(aplica)
        if aplica_norm not in resultado:
            formula = str(row.get("formula", "") or "").strip()
            if formula:
                resultado[aplica_norm] = formula
    return resultado


def _aplicar_formula_promedio(
    codigo: str,
    filas: list[dict[str, Any]],
) -> None:
    """Calcula promedio_final_general segun formulas registradas en formula_promedio."""
    formulas = _obtener_formulas_por_aplica_a(codigo)
    formula_ambos = formulas.get("ambos")
    formula_dre = formulas.get("dre / gre") or formulas.get("dre/gre") or formulas.get("dre")
    formula_ugel = formulas.get("ugel")

    for fila in filas:
        tipo_iged = str(fila.get("tipo_iged", "") or "").strip()
        formula_sel = None

        if formula_ambos:
            formula_sel = formula_ambos
        elif tipo_iged == "DRE / GRE":
            formula_sel = formula_dre or formula_ugel
        elif tipo_iged == "UGEL":
            formula_sel = formula_ugel or formula_dre
        else:
            formula_sel = formula_dre or formula_ugel

        if formula_sel:
            fila["promedio_final_general"] = _evaluar_formula_en_fila(formula_sel, fila)
        else:
            # Sin formula registrada conserva valor existente o 0.
            fila["promedio_final_general"] = _a_float(fila.get("promedio_final_general"))


def _nivel_cuestionario(nota: float | None) -> str:
    """Convierte nota numerica de cuestionario al nivel cualitativo esperado."""
    if nota is None:
        return ""
    if nota < 11:
        return "En inicio"
    if nota <= 13:
        return "En proceso"
    if nota <= 17:
        return "Logrado"
    return "Destacado"


def _aplicar_campos_derivados(filas: list[dict[str, Any]]) -> None:
    """Calcula columnas derivadas de estado final para la plantilla."""
    for fila in filas:
        # Normalizacion legacy: compromiso 1 o 20 se exporta como 20.
        compromiso_num = _a_float_nullable(fila.get("compromiso"))
        if compromiso_num in {1.0, 20.0}:
            fila["compromiso"] = 20

        promedio = _a_float_nullable(fila.get("promedio_final_general")) or 0.0
        avance = _a_float_nullable(fila.get("%_avance_certificacion"))
        if avance is None:
            avance = 0.0

        # Regla de condicion final por nota.
        fila["promedio_final_condicion"] = "Aprobado" if promedio >= 13.5 else "Desaprobado"

        # Regla legacy: aprueba si nota final aprobada y avance completo.
        aprueba = fila["promedio_final_condicion"] == "Aprobado" and avance > 0.99
        # Retirados no pueden certificar.
        es_retirado = _a_int(fila.get("estado")) == 3
        if es_retirado:
            aprueba = False
        fila["situacion_del_participante"] = "Aprueba" if aprueba else "No aprueba"
        fila["aprobados/certificados"] = 1 if aprueba else None

        # Clasifica desaprobado en abandono vs permanente.
        if aprueba:
            fila["desaprobado/abandono"] = None
            fila["desaprobado/permanente"] = None
        else:
            fila["desaprobado/abandono"] = 1 if avance <= 0.99 else None
            fila["desaprobado/permanente"] = (
                1
                if avance > 0.99 and fila.get("promedio_final_condicion") == "Desaprobado"
                else None
            )

        # Columnas derivadas de cuestionarios de entrada/salida.
        ce = _a_float_nullable(fila.get("cuestionario entrada"))
        cs = _a_float_nullable(fila.get("cuestionario salida"))
        fila["ev_progreso_aprendizaje"] = 1 if ce is not None and cs is not None else None
        # En legacy estos campos se serializan como texto "1".
        fila["mantuvo_o_progreso"] = "1" if ce is not None and cs is not None and (cs - ce) >= 0 else ""
        fila["progreso"] = "1" if ce is not None and cs is not None and (cs - ce) > 0 else ""
        fila["nivel_c._entrada"] = _nivel_cuestionario(ce)
        fila["nivel_c._salida"] = _nivel_cuestionario(cs)

        # Observacion replicando reglas de plantillas.py (incluye comparacion exacta "sin registrar").
        tipo_iged = str(fila.get("tipo_iged", "") or "").strip()
        tipo_norm = _normalizar_texto(tipo_iged)
        es_dre_ugel = tipo_norm in {"dre/gre", "dre / gre", "ugel"}

        nombre_iged_raw = fila.get("nombre_iged")
        nombre_puesto_raw = fila.get("nombre_puesto")

        def _es_na_legacy(valor: Any) -> bool:
            if valor is None:
                return True
            texto = str(valor).strip().lower()
            return texto in {"nan", "none"}

        iged_na = _es_na_legacy(nombre_iged_raw)
        puesto_na = _es_na_legacy(nombre_puesto_raw)
        puesto_txt = "" if puesto_na else str(nombre_puesto_raw)

        if es_dre_ugel and iged_na and puesto_na:
            fila["obs"] = "Sin iged y sin puesto"
        elif es_dre_ugel and iged_na and puesto_txt == "sin registrar":
            fila["obs"] = "Sin iged y sin puesto"
        elif es_dre_ugel and iged_na and not puesto_na:
            fila["obs"] = "Sin iged"
        elif es_dre_ugel and (not iged_na) and puesto_na:
            fila["obs"] = "Sin puesto"
        elif es_dre_ugel and (not iged_na) and puesto_txt == "sin registrar":
            fila["obs"] = "Sin puesto"
        elif not es_dre_ugel:
            fila["obs"] = "otras instituciones"
        else:
            fila["obs"] = ""


def _ordenar_filas_exportacion(filas: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Ordena filas para exportacion como en legacy (region + nombre_iged)."""
    def key_fn(row: dict[str, Any]) -> tuple[str, str]:
        return (
            str(row.get("region", "") or "").strip(),
            str(row.get("nombre_iged", "") or "").strip(),
        )

    return sorted(filas, key=key_fn)


def _columnas_exportacion(actividades: list[str]) -> list[str]:
    """Construye orden final de columnas (base + actividades registradas)."""
    cols = [col for col in ORDEN_COLUMNAS_PLANTILLA]
    for actividad in actividades:
        if actividad not in cols:
            cols.append(actividad)
    return cols


def _sanitizar_nombre_archivo(nombre: str, max_len: int = 180) -> str:
    """Sanitiza nombre de archivo para Windows y recorta longitud maxima."""
    texto = str(nombre or "").strip()
    if not texto:
        return "plantilla_generada"
    texto = re.sub(r"[\\/:*?\"<>|]+", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    if len(texto) > max_len:
        texto = texto[:max_len].rstrip()
    return texto or "plantilla_generada"


def _anio_desde_codigo(codigo: str) -> int:
    """Intenta inferir anio (YYYY) desde prefijo del codigo (ej: 26xxxx -> 2026)."""
    texto = str(codigo or "").strip()
    match = re.match(r"^(\d{2})", texto)
    if match:
        return 2000 + int(match.group(1))
    return datetime.now().year


def _nombre_archivo_plantilla(codigo: str) -> str:
    """Construye nombre de archivo amigable usando metadatos de oferta formativa."""
    query = """
        SELECT
            cap_codigo AS codigo,
            creado_nombre AS especialista_cargo,
            cap_tipo AS tipo_proceso_formativo,
            cap_nombre AS denominacion_proceso_formativo
        FROM cap_capacitaciones
        WHERE cap_codigo = %s
        LIMIT 1
    """
    row: dict[str, Any] = {}
    try:
        with get_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(query, (codigo,))
                row = cursor.fetchone() or {}
    except Exception:
        row = {}

    if not row:
        return _sanitizar_nombre_archivo(f"plantilla_{codigo}")

    especialista = str(row.get("especialista_cargo", "") or "").strip()
    especialista = re.sub(r"^\([^)]*\)\s*", "", especialista)
    inicial = especialista[0].upper() if especialista else "X"
    tipo = str(row.get("tipo_proceso_formativo", "") or "").strip()
    denominacion = str(row.get("denominacion_proceso_formativo", "") or "").strip()
    nombre = f"{codigo} - {inicial} - {tipo} {denominacion}".strip()
    return _sanitizar_nombre_archivo(nombre)


def _valor_excel(valor: Any) -> Any:
    """Normaliza valor antes de escribirlo en celda Excel."""
    if _valor_vacio(valor):
        return ""
    return valor


def _nombre_columna_excel_legacy(columna: str) -> str:
    """Retorna etiqueta de cabecera Excel compatible con el formato legacy."""
    col = str(columna or "").strip()
    especiales = {
        "mantuvo_o_progreso": "MANTUVO O PROGRESÓ",
        "progreso": "PROGRESÓ",
    }
    return especiales.get(col, col.upper().replace("_", " "))


def _agregar_tabla_excel(ws: Any, display_name: str, ref: str, style_name: str) -> None:
    """Agrega tabla OpenPyXL con estilo; ignora fallas de rango/nombre."""
    from openpyxl.worksheet.table import Table, TableStyleInfo

    try:
        tabla = Table(displayName=display_name, ref=ref)
        estilo = TableStyleInfo(
            name=style_name,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tabla.tableStyleInfo = estilo
        ws.add_table(tabla)
    except Exception:
        return


def _contar_dni_unicos(filas: list[dict[str, Any]], condicion: Any) -> int:
    """Cuenta DNIs unicos que cumplan la condicion indicada."""
    vistos: set[str] = set()
    for fila in filas:
        dni = _normalizar_dni(fila.get("dni"))
        if not dni:
            continue
        try:
            ok = bool(condicion(fila))
        except Exception:
            ok = False
        if ok:
            vistos.add(dni)
    return len(vistos)


def _agregar_hoja_reportes_legacy(
    wb: Any,
    filas: list[dict[str, Any]],
    estructura: list[dict[str, Any]],
) -> None:
    """Replica la hoja Reportes del modulo legacy de plantillas."""
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(title="Reportes")
    if not filas:
        return

    actividades_estructura: list[str] = []
    seen_acts: set[str] = set()
    for row in estructura:
        act = str(row.get("actividad", "") or "").strip()
        if act and act not in seen_acts:
            seen_acts.add(act)
            actividades_estructura.append(act)

    # 1) Resumen por region.
    regiones = sorted({str(fila.get("region", "") or "").strip() for fila in filas})
    if not regiones:
        regiones = [""]

    headers_1 = ["REGION", "POSTULANTES", "MATRICULADOS", "PARTICIPANTES"]
    for act in actividades_estructura:
        headers_1.append(f"CUMPLE {str(act).upper()}")
        headers_1.append(f"% {str(act).upper()}")
    headers_1.extend(["CERTIFICADOS", "% CERTIFICADOS"])

    resumen_rows: list[list[Any]] = []
    for region in regiones:
        filas_region = [fila for fila in filas if str(fila.get("region", "") or "").strip() == region]
        postulantes = _contar_dni_unicos(filas_region, lambda _: True)
        matriculados = _contar_dni_unicos(filas_region, lambda x: _a_int(x.get("estado")) == 2)
        participantes = _contar_dni_unicos(
            filas_region,
            lambda x: _a_int(x.get("estado")) == 2 and _a_float_nullable(x.get("compromiso")) in {1.0, 20.0},
        )

        fila_out: list[Any] = [region, postulantes, matriculados, participantes]
        for act in actividades_estructura:
            cumple = _contar_dni_unicos(
                filas_region,
                lambda x, nombre=act: not _valor_vacio(x.get(nombre)),
            )
            pct = round((cumple / participantes) * 100, 1) if participantes > 0 else 0
            fila_out.extend([cumple, pct])

        certificados = _contar_dni_unicos(
            filas_region,
            lambda x: (
                _a_int(x.get("estado")) == 2
                and _a_float_nullable(x.get("compromiso")) in {1.0, 20.0}
                and (_a_float_nullable(x.get("aprobados/certificados")) or 0) == 1.0
            ),
        )
        pct_cert = round((certificados / participantes) * 100, 1) if participantes > 0 else 0
        fila_out.extend([certificados, pct_cert])
        resumen_rows.append(fila_out)

    ws.append(headers_1)
    for row in resumen_rows:
        ws.append(row)

    if resumen_rows:
        idx_part = headers_1.index("PARTICIPANTES")
        total_participantes = sum(_a_int(row[idx_part]) for row in resumen_rows)
        totals: list[Any] = ["TOTAL"]
        for idx_col, col_name in enumerate(headers_1[1:], start=1):
            if str(col_name).startswith("% "):
                base_col = str(col_name).replace("% ", "CUMPLE ")
                if base_col in headers_1:
                    idx_base = headers_1.index(base_col)
                    total_cumple = sum(_a_int(row[idx_base]) for row in resumen_rows)
                    totals.append(round((total_cumple / total_participantes) * 100, 1) if total_participantes > 0 else 0)
                else:
                    totals.append(0)
            elif str(col_name) == "% CERTIFICADOS":
                idx_cert = headers_1.index("CERTIFICADOS")
                total_cert = sum(_a_int(row[idx_cert]) for row in resumen_rows)
                totals.append(round((total_cert / total_participantes) * 100, 1) if total_participantes > 0 else 0)
            else:
                totals.append(sum(_a_float_nullable(row[idx_col]) or 0 for row in resumen_rows))
        ws.append(totals)

    end_row_1 = ws.max_row
    if end_row_1 >= 2:
        _agregar_tabla_excel(
            ws=ws,
            display_name="ResumenRegion",
            ref=f"A1:{get_column_letter(ws.max_column)}{end_row_1}",
            style_name="TableStyleLight8",
        )

    for col_idx, col_name in enumerate(headers_1, start=1):
        if not (str(col_name).startswith("% ") or str(col_name) == "% CERTIFICADOS"):
            continue
        for row_idx in range(2, end_row_1 + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = _a_float_nullable(cell.value)
            if val is None:
                continue
            cell.value = val / 100
            cell.number_format = "0.0%"

    # 2) Cobertura por IGED.
    start_row_2 = ws.max_row + 2
    ws.cell(
        row=start_row_2,
        column=1,
        value="Cobertura por IGED (conteo único nombre_iged por tipo_iged y región)",
    )

    tipos_iged = sorted({
        str(fila.get("tipo_iged", "") or "").strip()
        for fila in filas
        if str(fila.get("tipo_iged", "") or "").strip()
    })
    regiones_cov = sorted({str(fila.get("region", "") or "").strip() for fila in filas})

    if tipos_iged:
        ws.append(["REGION"] + tipos_iged)
        cobertura_rows: list[list[Any]] = []
        for region in regiones_cov:
            filas_region = [fila for fila in filas if str(fila.get("region", "") or "").strip() == region]
            fila_cov: list[Any] = [region]
            for tipo in tipos_iged:
                igeds = {
                    str(item.get("nombre_iged", "") or "").strip()
                    for item in filas_region
                    if str(item.get("tipo_iged", "") or "").strip() == tipo
                    and str(item.get("nombre_iged", "") or "").strip()
                }
                fila_cov.append(len(igeds))
            cobertura_rows.append(fila_cov)
            ws.append(fila_cov)

        totals_cov = ["TOTAL"]
        for idx in range(1, len(tipos_iged) + 1):
            totals_cov.append(sum(_a_int(row[idx]) for row in cobertura_rows))
        ws.append(totals_cov)

        tab2_start = start_row_2 + 1
        tab2_end = tab2_start + len(cobertura_rows) + 1
        _agregar_tabla_excel(
            ws=ws,
            display_name="CoberturaIGED",
            ref=f"A{tab2_start}:{get_column_letter(len(tipos_iged)+1)}{tab2_end}",
            style_name="TableStyleLight9",
        )

    # 3) Participantes por nivel de puesto.
    start_row_3 = ws.max_row + 2
    ws.cell(row=start_row_3, column=1, value="Participantes por Nivel de Puesto")

    filas_estado2 = [fila for fila in filas if _a_int(fila.get("estado")) == 2]
    niveles = sorted({
        str(fila.get("nivel_puesto", "") or "").strip()
        for fila in filas_estado2
        if str(fila.get("nivel_puesto", "") or "").strip()
    })
    regiones_nivel = sorted({str(fila.get("region", "") or "").strip() for fila in filas_estado2})
    if niveles:
        ws.append(["REGION"] + niveles)
        nivel_rows: list[list[Any]] = []
        for region in regiones_nivel:
            filas_region = [fila for fila in filas_estado2 if str(fila.get("region", "") or "").strip() == region]
            fila_nivel: list[Any] = [region]
            for nivel in niveles:
                fila_nivel.append(
                    _contar_dni_unicos(
                        filas_region,
                        lambda x, n=nivel: str(x.get("nivel_puesto", "") or "").strip() == n,
                    )
                )
            nivel_rows.append(fila_nivel)
            ws.append(fila_nivel)

        totals_nivel = ["TOTAL"]
        for idx in range(1, len(niveles) + 1):
            totals_nivel.append(sum(_a_int(row[idx]) for row in nivel_rows))
        ws.append(totals_nivel)

        tab3_start = start_row_3 + 1
        tab3_end = tab3_start + len(nivel_rows) + 1
        _agregar_tabla_excel(
            ws=ws,
            display_name="ParticipantesNivelPuesto",
            ref=f"A{tab3_start}:{get_column_letter(len(niveles)+1)}{tab3_end}",
            style_name="TableStyleLight10",
        )

    # 4) Porcentajes por nivel de entrada.
    start_row_4 = ws.max_row + 2
    ws.cell(row=start_row_4, column=1, value="Porcentaje por Nivel de Entrada")

    filas_entrada = [fila for fila in filas if not _valor_vacio(fila.get("cuestionario entrada"))]
    niveles_entrada = sorted({
        str(fila.get("nivel_c._entrada", "") or "").strip()
        for fila in filas_entrada
        if str(fila.get("nivel_c._entrada", "") or "").strip()
    })
    regiones_entrada = sorted({str(fila.get("region", "") or "").strip() for fila in filas_entrada})
    if niveles_entrada:
        ws.append(["REGION"] + niveles_entrada)
        entrada_rows: list[list[Any]] = []
        entrada_counts: list[list[int]] = []
        for region in regiones_entrada:
            filas_region = [fila for fila in filas_entrada if str(fila.get("region", "") or "").strip() == region]
            counts: list[int] = []
            for nivel in niveles_entrada:
                counts.append(
                    _contar_dni_unicos(
                        filas_region,
                        lambda x, n=nivel: str(x.get("nivel_c._entrada", "") or "").strip() == n,
                    )
                )
            total = sum(counts)
            fila_pct = [region] + [round((v / total), 4) if total > 0 else 0 for v in counts]
            entrada_rows.append(fila_pct)
            entrada_counts.append(counts)
            ws.append(fila_pct)

        total_por_col = [sum(row[idx] for row in entrada_counts) for idx in range(len(niveles_entrada))]
        total_sum = sum(total_por_col)
        ws.append(["TOTAL"] + [round((v / total_sum), 4) if total_sum > 0 else 0 for v in total_por_col])

        tab4_start = start_row_4 + 1
        tab4_end = tab4_start + len(entrada_rows) + 1
        _agregar_tabla_excel(
            ws=ws,
            display_name="PorcentajeNivelEntrada",
            ref=f"A{tab4_start}:{get_column_letter(len(niveles_entrada)+1)}{tab4_end}",
            style_name="TableStyleLight11",
        )
        for col_idx in range(2, len(niveles_entrada) + 2):
            for row_idx in range(tab4_start + 1, tab4_end + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "0.0%"

    # 5) Porcentajes por nivel de salida.
    start_row_5 = ws.max_row + 2
    ws.cell(row=start_row_5, column=1, value="Porcentaje por Nivel de Salida")

    filas_salida = [fila for fila in filas if not _valor_vacio(fila.get("cuestionario salida"))]
    niveles_salida = sorted({
        str(fila.get("nivel_c._salida", "") or "").strip()
        for fila in filas_salida
        if str(fila.get("nivel_c._salida", "") or "").strip()
    })
    regiones_salida = sorted({str(fila.get("region", "") or "").strip() for fila in filas_salida})
    if niveles_salida:
        ws.append(["REGION"] + niveles_salida)
        salida_rows: list[list[Any]] = []
        salida_counts: list[list[int]] = []
        for region in regiones_salida:
            filas_region = [fila for fila in filas_salida if str(fila.get("region", "") or "").strip() == region]
            counts: list[int] = []
            for nivel in niveles_salida:
                counts.append(
                    _contar_dni_unicos(
                        filas_region,
                        lambda x, n=nivel: str(x.get("nivel_c._salida", "") or "").strip() == n,
                    )
                )
            total = sum(counts)
            fila_pct = [region] + [round((v / total), 4) if total > 0 else 0 for v in counts]
            salida_rows.append(fila_pct)
            salida_counts.append(counts)
            ws.append(fila_pct)

        total_por_col = [sum(row[idx] for row in salida_counts) for idx in range(len(niveles_salida))]
        total_sum = sum(total_por_col)
        ws.append(["TOTAL"] + [round((v / total_sum), 4) if total_sum > 0 else 0 for v in total_por_col])

        tab5_start = start_row_5 + 1
        tab5_end = tab5_start + len(salida_rows) + 1
        _agregar_tabla_excel(
            ws=ws,
            display_name="PorcentajeNivelSalida",
            ref=f"A{tab5_start}:{get_column_letter(len(niveles_salida)+1)}{tab5_end}",
            style_name="TableStyleLight12",
        )
        for col_idx in range(2, len(niveles_salida) + 2):
            for row_idx in range(tab5_start + 1, tab5_end + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "0.0%"

    # Ajuste de anchos como en legacy.
    for idx_col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(idx_col)
        max_len = 0
        for cell in ws[col_letter]:
            texto = str(cell.value or "")
            if len(texto) > max_len:
                max_len = len(texto)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 100)


def _escribir_hoja_tabla_excel(
    ws: Any,
    codigo: str,
    filas: list[dict[str, Any]],
    columnas: list[str],
    titulo_tabla: str,
) -> None:
    """Escribe hoja BBDD TUTOR con formato de tabla estilo legacy."""
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    encabezados = ["N°", "CODIGO"] + [_nombre_columna_excel_legacy(col) for col in columnas]
    ws.append(encabezados)

    for idx, fila in enumerate(filas, 1):
        valores = [_valor_excel(fila.get(col)) for col in columnas]
        ws.append([idx, codigo] + valores)

    # Solo agrega tabla cuando existe al menos una fila de datos.
    if ws.max_row >= 2:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        _agregar_tabla_excel(
            ws=ws,
            display_name=titulo_tabla,
            ref=ref,
            style_name="TableStyleLight8",
        )

    # Encabezado en negrita y centrado.
    header_font = Font(color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # Autoajuste acotado de ancho por legibilidad.
    for idx_col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(idx_col)
        max_len = 0
        for cell in ws[col_letter]:
            texto = str(cell.value or "")
            if len(texto) > max_len:
                max_len = len(texto)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 100)


def _crear_excel_plantilla(
    ruta_salida: Path,
    codigo: str,
    filas: list[dict[str, Any]],
    columnas_plantilla: list[str],
    estructura: list[dict[str, Any]],
) -> bool:
    """Crea archivo XLSX principal de plantilla (BBDD TUTOR + Reportes)."""
    try:
        from openpyxl import Workbook
    except Exception:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "BBDD TUTOR"
    _escribir_hoja_tabla_excel(ws, codigo, filas, columnas_plantilla, "TablaBBDD")
    _agregar_hoja_reportes_legacy(wb, filas, estructura)

    try:
        ruta_salida.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(ruta_salida))
        return True
    except Exception:
        return False
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _es_actividad_cumplida(valor: Any) -> bool:
    """Evalua cumplimiento con la misma regla legacy (valor no vacio)."""
    if valor is None:
        return False
    if isinstance(valor, str):
        texto = valor.strip()
        return texto != "" and texto.lower() not in {"nan", "none", "null"}
    return True


def _crear_excel_nominal(
    ruta_salida: Path,
    filas: list[dict[str, Any]],
    columnas_plantilla: list[str],
    columnas_nominal: list[str],
    titulo_nominal: str,
    columnas_actividades: list[str],
) -> bool:
    """Genera reporte nominal con el mismo formato operativo del flujo legacy."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.formatting.rule import CellIsRule, FormulaRule
        from openpyxl.utils import get_column_letter
    except Exception:
        return False

    # Filtra solo participantes matriculados (estado=2), como en app_difoca.
    filas_estado2 = [row for row in filas if _a_int(row.get("estado")) == 2]
    filas_estado2 = _ordenar_filas_exportacion(filas_estado2)
    if not filas_estado2:
        return False

    # Usa seleccion guardada del nominal y valida contra columnas de plantilla.
    cols_validas = [col for col in columnas_nominal if col in columnas_plantilla]
    if not cols_validas:
        cols_validas = [col for col in columnas_plantilla if col in COLUMNAS_NOMINAL_BASE]

    # Orden principal + actividades al final (orden legacy).
    orden_principal = [
        "region",
        "nombre_iged",
        "dni",
        "apellidos",
        "nombres",
        "email",
        "telefono_celular",
        "telefono_fijo",
        "actualizo_datos",
        "ultima conexion",
        "estado",
        "compromiso",
    ]
    actividades_sel = [col for col in cols_validas if col not in orden_principal]
    columnas_finales = [col for col in orden_principal if col in cols_validas] + actividades_sel
    if not columnas_finales:
        columnas_finales = list(cols_validas)

    # Arma encabezados visibles en Excel.
    encabezados = ["N°"] + [str(col).upper().replace("_", " ") for col in columnas_finales]
    encabezados_upper = [str(x).upper().strip() for x in encabezados]

    # Ubica columna ACTUALIZO DATOS para regla condicional.
    idx_col_actualizo = None
    for idx, head in enumerate(encabezados_upper, start=1):
        if head == "ACTUALIZO DATOS":
            idx_col_actualizo = idx
            break

    # Detecta indices de columnas de actividades para marcar vacios en rojo.
    columnas_principales_legacy = {
        "TIPO DOCUMENTO",
        "DNI",
        "GENERO",
        "APELLIDOS",
        "NOMBRES",
        "FECHA NACIMIENTO",
        "EMAIL",
        "TELEFONO CELULAR",
        "TELEFONO FIJO",
        "ACTUALIZO DATOS",
        "REGION",
        "TIPO IGED",
        "NOMBRE IGED",
        "NIVEL DE PUESTO",
        "NOMBRE DE PUESTO",
        "REGIMEN LABORAL",
        "CODIGO DE IGED",
        "ULTIMA CONEXION",
        "ESTADO",
        "N°",
    }
    indices_actividades: list[int] = []
    actividades_set = set(columnas_actividades)
    for idx, col in enumerate(columnas_finales, start=2):
        head = str(col).upper().replace("_", " ")
        if head in columnas_principales_legacy:
            continue
        if col in actividades_set or col not in orden_principal:
            indices_actividades.append(idx)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte nominal"

    # Titulo en fila 1 con fecha en formato DD DE MES.
    meses = {
        1: "ENERO",
        2: "FEBRERO",
        3: "MARZO",
        4: "ABRIL",
        5: "MAYO",
        6: "JUNIO",
        7: "JULIO",
        8: "AGOSTO",
        9: "SEPTIEMBRE",
        10: "OCTUBRE",
        11: "NOVIEMBRE",
        12: "DICIEMBRE",
    }
    fecha_hoy = datetime.now()
    fecha_formateada = f"{fecha_hoy.day:02d} DE {meses.get(fecha_hoy.month, '')}"
    titulo_base = str(titulo_nominal or "Reporte Nominal").strip() or "Reporte Nominal"
    titulo_completo = f"{titulo_base} - {fecha_formateada}"
    total_cols = max(len(encabezados), 1)
    ws.cell(row=1, column=1, value=titulo_completo)
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    celda_titulo = ws.cell(row=1, column=1)
    celda_titulo.alignment = Alignment(horizontal="center", vertical="center")
    celda_titulo.font = Font(color="FFFFFF", bold=True)
    celda_titulo.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    ws.row_dimensions[1].height = 30

    # Encabezados de tabla (fila 2).
    ws.append(encabezados)
    for cell in ws[2]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # Datos desde fila 3.
    for idx, fila in enumerate(filas_estado2, 1):
        valores: list[Any] = []
        for col in columnas_finales:
            valor = fila.get(col)
            # En nominal las actividades se exportan como "Sí" cuando tienen dato.
            if col in columnas_actividades:
                valor = "Sí" if _es_actividad_cumplida(valor) else ""
            valores.append(_valor_excel(valor))
        ws.append([idx] + valores)

    # Tabla principal (encabezado en fila 2).
    if ws.max_row >= 3:
        _agregar_tabla_excel(
            ws=ws,
            display_name="TablaNominal",
            ref=f"A2:{get_column_letter(ws.max_column)}{ws.max_row}",
            style_name="TableStyleLight9",
        )

    # Resalta "No" en ACTUALIZO DATOS.
    if idx_col_actualizo is not None and ws.max_row >= 3:
        col_letter = get_column_letter(idx_col_actualizo)
        rango_actualizo = f"{col_letter}3:{col_letter}{ws.max_row}"
        ws.conditional_formatting.add(
            rango_actualizo,
            CellIsRule(
                operator="equal",
                formula=['"No"'],
                fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            ),
        )

    # Marca celdas vacias de actividades.
    if ws.max_row >= 3:
        for idx_col in indices_actividades:
            col_letter = get_column_letter(idx_col)
            rango = f"{col_letter}3:{col_letter}{ws.max_row}"
            ws.conditional_formatting.add(
                rango,
                FormulaRule(
                    formula=[f"LEN(TRIM({col_letter}3))=0"],
                    fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
                ),
            )

    # Ancho fijo legacy.
    for idx_col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(idx_col)].width = 20

    try:
        ruta_salida.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(ruta_salida))
        return True
    except Exception:
        return False
    finally:
        try:
            wb.close()
        except Exception:
            pass

def _crear_excel_cumplimiento_iged(
    ruta_salida: Path,
    filas: list[dict[str, Any]],
    estructura: list[dict[str, Any]],
    columnas_actividades: list[str],
    denominaciones_grupos: dict[str, str] | None = None,
) -> bool:
    """Genera cumplimiento por IGED con agrupacion/estilos equivalentes al flujo legacy."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        return False

    # Filtra solo participantes activos con compromiso valido.
    filas_base: list[dict[str, Any]] = []
    for row in filas:
        if _a_int(row.get("estado")) != 2:
            continue
        compromiso = _a_float_nullable(row.get("compromiso"))
        if compromiso not in {1.0, 20.0}:
            continue
        filas_base.append(row)
    if not filas_base:
        return False

    # Construye indice normalizado de actividad -> grupo desde estructura.
    indice_grupo: dict[str, str] = {}
    for row in estructura:
        actividad = str(row.get("actividad", "") or "").strip()
        grupo = str(row.get("grupo", "") or "").strip()
        if not actividad or not grupo:
            continue
        indice_grupo[_normalizar_texto(actividad)] = grupo

    # Conserva actividades con grupo configurado (criterio legacy).
    grupo_por_actividad: dict[str, str] = {}
    actividades: list[str] = []
    for actividad in columnas_actividades:
        act_txt = str(actividad or "").strip()
        if not act_txt:
            continue
        grupo = indice_grupo.get(_normalizar_texto(act_txt))
        if not grupo:
            continue
        grupo_por_actividad[act_txt] = grupo
        actividades.append(act_txt)

    # Fallback controlado cuando no hay grupos: todas a grupo "General".
    if not actividades:
        for actividad in columnas_actividades:
            act_txt = str(actividad or "").strip()
            if not act_txt:
                continue
            grupo_por_actividad[act_txt] = "General"
            actividades.append(act_txt)
    if not actividades:
        return False

    # Orden de grupos y actividades por grupo.
    grupos_orden: list[str] = []
    actividades_por_grupo: dict[str, list[str]] = {}
    for actividad in actividades:
        grupo = grupo_por_actividad.get(actividad, "General")
        if grupo not in actividades_por_grupo:
            actividades_por_grupo[grupo] = []
            grupos_orden.append(grupo)
        actividades_por_grupo[grupo].append(actividad)

    # Agrega por REGION + NOMBRE IGED.
    agrupado: dict[tuple[str, str], dict[str, Any]] = {}
    for row in filas_base:
        region = str(row.get("region", "") or "").strip() or "SIN REGION"
        iged = str(row.get("nombre_iged", "") or "").strip() or "SIN IGED"
        key = (region, iged)
        item = agrupado.setdefault(
            key,
            {
                "region": region,
                "nombre_iged": iged,
                "dnis": set(),
                "cumple": {},
            },
        )

        dni = _normalizar_dni(row.get("dni"))
        if dni:
            item["dnis"].add(dni)

        for act in actividades:
            if _es_actividad_cumplida(row.get(act)):
                item["cumple"][act] = int(item["cumple"].get(act, 0)) + 1

    # Filas de salida ordenadas.
    filas_reporte: list[list[Any]] = []
    for (_, _), item in sorted(agrupado.items(), key=lambda x: (x[0][0], x[0][1])):
        participantes = len(item["dnis"])
        fila = [item["region"], item["nombre_iged"], participantes]
        for act in actividades:
            cumple = int(item["cumple"].get(act, 0))
            pct = (cumple / participantes) if participantes > 0 else 0.0
            fila.extend([cumple, pct])
        filas_reporte.append(fila)

    if not filas_reporte:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Cumplimiento por IGED"

    # Columnas finales para encabezado y datos.
    columnas_base = ["region", "nombre_iged", "participantes"]
    columnas_finales: list[str] = list(columnas_base)
    grupos_encabezados: list[tuple[str, int]] = []
    for grupo in grupos_orden:
        acts_grupo = actividades_por_grupo.get(grupo, [])
        if not acts_grupo:
            continue
        grupos_encabezados.append((grupo, len(acts_grupo) * 2))
        for act in acts_grupo:
            columnas_finales.append(f"{act}__cumplimiento")
            columnas_finales.append(f"{act}__porcentaje")

    # Fila 2: grupos (con merge) y columnas base vacias.
    col_idx = len(columnas_base) + 1
    grupo_colors = ["B8CCE4", "244062"]
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    nombres_grupos = denominaciones_grupos or {}
    for idx_grupo, (grupo, ncols) in enumerate(grupos_encabezados):
        nombre_grupo = str(nombres_grupos.get(str(grupo), "") or "").strip()
        if not nombre_grupo:
            nombre_grupo = f"Grupo {grupo}" if str(grupo).isdigit() else str(grupo)

        start_col = get_column_letter(col_idx)
        end_col = get_column_letter(col_idx + ncols - 1)
        ws.merge_cells(f"{start_col}2:{end_col}2")

        color = grupo_colors[idx_grupo % len(grupo_colors)]
        font_color = "000000" if color == "B8CCE4" else "FFFFFF"
        celda_grupo = ws.cell(row=2, column=col_idx, value=nombre_grupo)
        celda_grupo.alignment = Alignment(horizontal="center", vertical="center")
        celda_grupo.font = Font(color=font_color, bold=True)
        celda_grupo.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        for c in range(col_idx, col_idx + ncols):
            ws.cell(row=2, column=c).border = thin_border
        col_idx += ncols

    for i in range(1, len(columnas_base) + 1):
        celda = ws.cell(row=2, column=i, value="")
        celda.fill = PatternFill(fill_type=None)
        celda.border = thin_border

    # Fila 3: encabezados por columna.
    columnas_excel: list[str] = []
    mapa_columna_a_actividad: dict[int, str] = {}
    mapa_columna_es_porcentaje: dict[int, bool] = {}
    idx_out = 1
    for col in columnas_finales:
        if col.endswith("__cumplimiento"):
            act = col[:-14]
            columnas_excel.append(str(act).upper().replace("_", " "))
            mapa_columna_a_actividad[idx_out] = act
            mapa_columna_es_porcentaje[idx_out] = False
        elif col.endswith("__porcentaje"):
            act = col[:-12]
            columnas_excel.append(f"{str(act).upper().replace('_', ' ')} (%)")
            mapa_columna_a_actividad[idx_out] = act
            mapa_columna_es_porcentaje[idx_out] = True
        else:
            columnas_excel.append(str(col).upper().replace("_", " "))
            mapa_columna_es_porcentaje[idx_out] = False
        idx_out += 1

    grupo_color_map: dict[str, str] = {}
    for idx_grupo, grupo in enumerate(grupos_orden):
        grupo_color_map[grupo] = grupo_colors[idx_grupo % len(grupo_colors)]

    for idx, header in enumerate(columnas_excel, start=1):
        celda = ws.cell(row=3, column=idx, value=header)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if idx <= len(columnas_base):
            celda.fill = PatternFill(start_color="244062", end_color="244062", fill_type="solid")
            celda.font = Font(color="FFFFFF", bold=True)
        else:
            actividad = mapa_columna_a_actividad.get(idx, "")
            grupo = grupo_por_actividad.get(actividad, "General")
            color = grupo_color_map.get(grupo, "B8CCE4")
            celda.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            celda.font = Font(color="000000" if color == "B8CCE4" else "FFFFFF", bold=True)
        celda.border = thin_border

    # Filas de datos desde fila 4.
    for fila_idx, fila in enumerate(filas_reporte, start=4):
        col_excel = 1
        for valor in fila:
            celda = ws.cell(row=fila_idx, column=col_excel, value=valor)
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            celda.border = thin_border

            if col_excel > len(columnas_base):
                es_porcentaje = mapa_columna_es_porcentaje.get(col_excel, False)

                # Regla solicitada: solo las columnas de porcentaje llevan sombreado.
                if es_porcentaje:
                    celda.number_format = "0%"
                    celda.fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
                    celda.font = Font(color="000000")
                else:
                    # Sin rojo ni sombreado para columnas no porcentuales.
                    celda.fill = PatternFill(fill_type=None)
                    celda.font = Font(color="000000")
            else:
                celda.fill = PatternFill(fill_type=None)
                celda.font = Font(color="000000")

            col_excel += 1

    # Ancho fijo por columna.
    for col_num in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    try:
        ruta_salida.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(ruta_salida))
        return True
    except Exception:
        return False
    finally:
        try:
            wb.close()
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Insercion en bbdd_2026 (alimenta la VIEW bbdd_difoca) - no sincronicas
# ---------------------------------------------------------------------------

# Mapeo de claves de filas_por_dni -> columnas bbdd_2026
_PLANTILLA_TO_BBDD = {
    "tipo_documento": "tipo_documento",
    "dni": "dni",
    "apellidos": "apellidos",
    "nombres": "nombres",
    "genero": "genero",
    "fecha_nacimiento": "fecha_nacimiento",
    "email": "email",
    "telefono_celular": "telefono_celular",
    "telefono_fijo": "telefono_fijo",
    "actualizo_datos": "actualizo_datos",
    "region": "region",
    "tipo_iged": "tipo_iged",
    "codigo_iged": "codigo_iged",
    "nombre_iged": "nombre_iged",
    "nivel_puesto": "nivel_puesto",
    "nombre_puesto": "nombre_puesto",
    "regimen_laboral": "regimen_laboral",
    "ultima conexion": "ultimo_acceso_curso",
    "estado": "estado",
    "compromiso": "compromiso",
    "%_avance_certificacion": "avance_curso_certificacion",
    "promedio_final_general": "promedio_final_general",
    "promedio_final_condicion": "promedio_final_condicion",
    "situacion_del_participante": "situacion_participante",
    "aprobados/certificados": "aprobados_certificados",
    "desaprobado/permanente": "desaprobado_permanente",
    "desaprobado/abandono": "desaprobado_abandono",
    "ev_progreso_aprendizaje": "ev_progreso_aprendizaje",
    "mantuvo_o_progreso": "mantuvo_o_progreso",
    "progreso": "progreso",
    "nivel_c._entrada": "nivel_c_entrada",
    "nivel_c._salida": "nivel_c_salida",
    "obs": "obs",
    "retiros": "retiros",
    "cuestionario entrada": "cuestionario_entrada",
    "cuestionario salida": "cuestionario_salida",
}


def _insertar_plantilla_en_bbdd_2026(filas: list[dict[str, Any]], codigo: str) -> None:
    """Inserta filas procesadas de plantilla no-sincronica en bbdd_2026.

    Si ya existen registros del mismo codigo, los reemplaza (DELETE + INSERT).
    """
    if not filas or not codigo:
        return

    cols_bbdd = ["codigo"] + list(_PLANTILLA_TO_BBDD.values())

    def _safe(v: Any) -> Any:
        if v is None:
            return None
        try:
            import math
            if isinstance(v, float) and math.isnan(v):
                return None
        except (ValueError, TypeError):
            pass
        s = str(v)
        if s in ("", "nan", "NaN", "None", "none", "null", "<NA>"):
            return None
        return v

    try:
        rows_to_insert = []
        for fila in filas:
            row_vals = [codigo]  # primera columna: codigo
            for key_plantilla, col_bbdd in _PLANTILLA_TO_BBDD.items():
                row_vals.append(_safe(fila.get(key_plantilla)))
            rows_to_insert.append(tuple(row_vals))

        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM bbdd_2026 WHERE codigo = %s", (codigo,))
                deleted = cur.rowcount
                if deleted:
                    logger.info(
                        "_insertar_plantilla_en_bbdd_2026: eliminados %d registros previos de %s",
                        deleted, codigo,
                    )

                placeholders = ", ".join(["%s"] * len(cols_bbdd))
                col_names = ", ".join(f"`{c}`" for c in cols_bbdd)
                sql = f"INSERT INTO bbdd_2026 ({col_names}) VALUES ({placeholders})"
                cur.executemany(sql, rows_to_insert)
                logger.info(
                    "_insertar_plantilla_en_bbdd_2026: insertados %d registros para %s",
                    len(rows_to_insert), codigo,
                )
    except Exception:
        logger.exception("Error insertando plantilla en bbdd_2026 para codigo=%s", codigo)


def generar_plantilla_seguimiento(
    codigo: str,
    anio: int | None = None,
) -> dict[str, Any]:
    """Genera plantilla XLSX real para la pestaÃ±a Seguimiento > Generacion de plantilla."""
    codigo = str(codigo or "").strip()
    if not codigo:
        return {"ok": False, "error": "Codigo vacio."}

    # Carga datasets principales requeridos por el flujo.
    estructura = obtener_estructura_por_codigo(codigo)
    nominal_config = obtener_config_nominal_reporte(codigo, estructura)
    postulantes_info = obtener_postulantes_excel_info(codigo)

    # Prioriza lista de postulantes cargada; si no existe, combina bbdd_difoca + Chamilo.
    dnis_objetivo: list[str] = []
    if postulantes_info.get("exists"):
        dnis_objetivo = _leer_excel_postulantes_dni(str(postulantes_info.get("path", "")))
    if not dnis_objetivo:
        # Combina TODOS los DNIs de bbdd_difoca (cualquier estado, incluyendo retirados)
        # con los matriculados actuales de Chamilo para capturar nuevas altas.
        dnis_bbdd = _obtener_todos_dnis_por_codigo(codigo)
        dnis_aula = _obtener_dnis_matriculados_aula(codigo)
        vistos: set[str] = set()
        for dni in dnis_bbdd + dnis_aula:
            if dni not in vistos:
                vistos.add(dni)
                dnis_objetivo.append(dni)

    filas_bbdd = _obtener_filas_bbdd_por_codigo_y_dnis(codigo, dnis_objetivo)
    if not filas_bbdd and not dnis_objetivo:
        # Fallback final cuando no hay postulantes/matriculados: usa todo el codigo.
        filas_bbdd = _obtener_filas_bbdd_por_codigo_y_dnis(codigo, [])

    if not filas_bbdd and not dnis_objetivo:
        return {
            "ok": False,
            "error": "No hay datos en bbdd_difoca ni matriculados en el aula virtual para este codigo.",
        }

    # Construye mapa base por DNI y asegura presencia de todos los DNIs objetivo.
    filas_por_dni: dict[str, dict[str, Any]] = {}
    for row in filas_bbdd:
        fila = _fila_base_plantilla_desde_bbdd(row)
        dni = _normalizar_dni(fila.get("dni"))
        if not dni:
            continue
        # Conserva la primera fila por DNI (ya ordenada por ultima conexion DESC).
        if dni in filas_por_dni:
            continue
        filas_por_dni[dni] = fila

    for dni in dnis_objetivo:
        dni_norm = _normalizar_dni(dni)
        if dni_norm and dni_norm not in filas_por_dni:
            filas_por_dni[dni_norm] = _fila_vacia_para_dni(dni_norm)

    # Excluye DNIs administrativos/heredados fuera del universo operativo.
    for dni_excluido in list(filas_por_dni.keys()):
        if dni_excluido in DNIS_EXCLUIDOS_PLANTILLA:
            filas_por_dni.pop(dni_excluido, None)

    if not filas_por_dni:
        return {"ok": False, "error": "No se pudo construir el universo de participantes."}

    # Completa datos demograficos desde SIDI para DNIs sin informacion previa.
    _enriquecer_filas_con_sidi(filas_por_dni)

    _aplicar_estado_matricula_y_retiros(codigo, filas_por_dni)
    actividades_export = _aplicar_actividades_a_filas(codigo, estructura, filas_por_dni)
    filas = list(filas_por_dni.values())
    _recalcular_avance_certificacion(estructura, filas)
    _aplicar_formula_promedio(codigo, filas)
    _aplicar_campos_derivados(filas)
    filas_ordenadas = _ordenar_filas_exportacion(filas)

    columnas_plantilla = _columnas_exportacion(actividades_export)
    columnas_nominal = list(nominal_config.get("columnas_seleccionadas", []))
    titulo_nominal = str(nominal_config.get("titulo_nominal", "Reporte Nominal")).strip() or "Reporte Nominal"

    anio_salida = int(anio) if anio else _anio_desde_codigo(codigo)
    nombre_base = _nombre_archivo_plantilla(codigo)
    dir_salida = _ruta_plantillas_generadas_dir(anio_salida, crear=True)
    ruta_main = dir_salida / f"{nombre_base}.xlsx"
    ruta_nominal = dir_salida / f"{nombre_base}_NOMINAL.xlsx"
    ruta_iged = dir_salida / f"{nombre_base}_IGED.xlsx"

    ok_main = _crear_excel_plantilla(
        ruta_salida=ruta_main,
        codigo=codigo,
        filas=filas_ordenadas,
        columnas_plantilla=columnas_plantilla,
        estructura=estructura,
    )
    if not ok_main:
        return {"ok": False, "error": "No se pudo crear el archivo Excel de plantilla."}

    ok_nominal = _crear_excel_nominal(
        ruta_salida=ruta_nominal,
        filas=filas_ordenadas,
        columnas_plantilla=columnas_plantilla,
        columnas_nominal=columnas_nominal,
        titulo_nominal=titulo_nominal,
        columnas_actividades=actividades_export,
    )
    ok_iged = _crear_excel_cumplimiento_iged(
        ruta_salida=ruta_iged,
        filas=filas_ordenadas,
        estructura=estructura,
        columnas_actividades=actividades_export,
        denominaciones_grupos=nominal_config.get("nombres_grupo", {}),
    )

    files_meta: list[dict[str, Any]] = []
    for kind, label, ruta, ok_file in [
        ("main", "Plantilla generada", ruta_main, ok_main),
        ("nominal", "Reporte nominal", ruta_nominal, ok_nominal),
        ("iged", "Cumplimiento por IGED", ruta_iged, ok_iged),
    ]:
        exists_file = bool(ok_file and ruta.exists())
        files_meta.append(
            {
                "kind": kind,
                "label": label,
                "path": str(ruta),
                "file_name": ruta.name,
                "size_bytes": int(ruta.stat().st_size) if exists_file else 0,
                "exists": exists_file,
            }
        )

    # Sincroniza tabla plantillas con la nueva ruta de salida.
    rutas_actuales = obtener_rutas_plantilla(codigo)
    guardar_rutas_plantilla(
        codigo=codigo,
        excel_path=str(ruta_main),
        py_path=str(rutas_actuales.get("py", "") or "").strip(),
    )

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    payload = {
        "path": str(ruta_main),
        "file_name": ruta_main.name,
        "size_bytes": int(ruta_main.stat().st_size) if ruta_main.exists() else 0,
        "generated_at": generated_at,
        "codigo": codigo,
        "anio": anio_salida,
        "files": files_meta,
    }
    _guardar_metadata_plantilla_generada(codigo, payload)

    # Insertar datos procesados en bbdd_2026 (alimenta la VIEW bbdd_difoca)
    _insertar_plantilla_en_bbdd_2026(filas_ordenadas, codigo)

    return {
        "ok": True,
        "path": str(ruta_main),
        "file_name": ruta_main.name,
        "size_bytes": payload["size_bytes"],
        "generated_at": generated_at,
        "total_participantes": len(filas_ordenadas),
        "files": files_meta,
    }

