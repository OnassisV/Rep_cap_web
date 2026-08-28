from io import BytesIO
from datetime import date
from types import SimpleNamespace
from unittest.mock import patch

import pandas as pd
from django.test import SimpleTestCase
from openpyxl import load_workbook

from core import estandares_calidad as estandares


class EstandaresCalidadTests(SimpleTestCase):
    def test_catalogo_tiene_cuatro_capitulos_y_preguntas_mapeadas(self):
        self.assertEqual(
            set(estandares.CAPITULOS),
            {
                "Autoría",
                "Sustento y planificación",
                "Implementación y seguimiento",
                "Evaluación",
            },
        )

        preguntas = [
            pregunta["pregunta"]
            for capitulo in estandares.CAPITULOS
            for pregunta in estandares.PREGUNTAS_CAPITULOS[capitulo]
        ]
        items = {
            item
            for pregunta in preguntas
            for item in [estandares._extraer_item(pregunta)]
            if item is not None
        }

        self.assertEqual(len(preguntas), 62)
        self.assertTrue(items.issubset(estandares.MAPEO_ESTANDARES))
        self.assertTrue(items.issubset(estandares.MAPEO_ETAPAS))

    def test_autollenado_resuelve_datos_del_proceso_y_kpis(self):
        datos = {
            "proceso_combinado": "Gestión de riesgos",
            "modalidad": "Virtual",
            "implementacion_final": "2026-08-15",
        }
        kpis = {"kpi_postulaciones": 12}

        self.assertEqual(
            estandares.resolver_valor_autollenado("proceso_sel", datos, kpis),
            "Gestión de riesgos",
        )
        self.assertEqual(
            estandares.resolver_valor_autollenado("fecha_finalizacion", datos, kpis),
            "2026-08-15",
        )
        self.assertEqual(
            estandares.resolver_valor_autollenado("modalidad", datos, kpis),
            "Virtual",
        )
        self.assertEqual(
            estandares.resolver_valor_autollenado("kpi_postulaciones", datos, kpis),
            "12",
        )

    def test_datos_proceso_devuelve_metadatos_de_capacitacion_seleccionada(self):
        capacitacion = SimpleNamespace(
            cap_tipo="Curso",
            cap_nombre="Gestión de riesgos",
            especialista_cargo="Especialista Uno",
            publico_objetivo_oferta="Equipos directivos",
            mi_objetivo_capacitacion="Fortalecer la gestión",
            pt_horas=24,
            pt_modalidad="Virtual",
            pt_implementacion_inicio=date(2026, 8, 1),
            pt_implementacion_fin=date(2026, 8, 15),
        )

        class CapacitacionQuery:
            def first(self):
                return capacitacion

        with patch.object(
            estandares.Capacitacion.objects,
            "filter",
            return_value=CapacitacionQuery(),
        ) as filtro:
            datos = estandares.obtener_datos_proceso("26001X-123")

        filtro.assert_called_once_with(cap_codigo="26001X", cap_id_curso="123")
        self.assertEqual(datos["proceso_combinado"], "Gestión de riesgos")
        self.assertEqual(datos["publico_objetivo"], "Equipos directivos")
        self.assertEqual(datos["objetivo_capacitacion"], "Fortalecer la gestión")
        self.assertEqual(datos["horas_certificacion"], 24)
        self.assertEqual(datos["modalidad"], "Virtual")
        self.assertEqual(datos["implementacion_inicio"], date(2026, 8, 1))

    def test_completitud_se_calcula_por_estandar(self):
        respuestas = pd.DataFrame(
            [
                {"codigo": "26001X", "pregunta": "A.1 Norma", "respuesta": "Ley"},
                {"codigo": "26001X", "pregunta": "A.2 Plan", "respuesta": ""},
            ]
        )

        resultado = estandares._calcular_completitud(
            respuestas,
            {"A.1": "Pertinencia", "A.2": "Pertinencia"},
        )

        self.assertEqual(resultado["Pertinencia"]["completitud"], 1)
        self.assertEqual(resultado["Pertinencia"]["total"], 2)
        self.assertEqual(resultado["Pertinencia"]["porcentaje"], 50)

    def test_reporte_individual_devuelve_excel_con_respuestas(self):
        filas = [
            {
                "capitulo": "Autoría",
                "pregunta": "Nombre de la capacitación",
                "respuesta": "Gestión de riesgos",
            }
        ]

        class Cursor:
            def __enter__(self):
                return self

            def __exit__(self, *args):
                return False

            def execute(self, query, params):
                self.rows = filas

            def fetchall(self):
                return self.rows

        class Connection:
            def cursor(self):
                return Cursor()

            def close(self):
                pass

        class CapacitacionQuery:
            def first(self):
                return None

        with patch.object(estandares, "get_connection", return_value=Connection()):
            with patch.object(
                estandares.Capacitacion.objects,
                "filter",
                return_value=CapacitacionQuery(),
            ):
                reporte = estandares.generar_reporte_individual("26001X")

        libro = load_workbook(BytesIO(reporte))
        self.assertEqual(libro.sheetnames, estandares.CAPITULOS)
        self.assertEqual(libro["Autoría"]["C4"].value, "Gestión de riesgos")