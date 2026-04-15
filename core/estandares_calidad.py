"""
Modulo temporal para gestion y reporteria de estandares de calidad DIFOCA.

Replica la funcionalidad del modulo herramientas/estandares_calidad.py de la app
Streamlit (app_difoca) para uso transitorio en la plataforma web Django.
Se eliminara una vez completados los datos del periodo anterior.
"""

from datetime import datetime
from io import BytesIO
from typing import Any
import re

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from accounts.db import get_connection
from core.indicadores_adapters import _count_distinct_non_empty, _normalize_iged_name

ANIO_VIGENTE = datetime.now().year

# ---------------------------------------------------------------------------
# Definiciones de capitulos y preguntas (replicadas del modulo Streamlit)
# ---------------------------------------------------------------------------

CAPITULOS = [
    "Autoría",
    "Sustento y planificación",
    "Implementación y seguimiento",
    "Evaluación",
]

PREGUNTAS_CAPITULOS: dict[str, list[dict[str, Any]]] = {
    "Autoría": [
        {"pregunta": "Nombre de la capacitación", "tipo": "autollenado", "source": "proceso_sel"},
        {"pregunta": "Nombre del gestor/a que completa la información", "tipo": "autollenado", "source": "nombre_especialista"},
        {"pregunta": "Fecha de finalización de la capacitación", "tipo": "autollenado", "source": "fecha_finalizacion"},
        {"pregunta": "Fecha de inicio de llenado de la información", "tipo": "autollenado", "source": "fecha_inicio_llenado"},
        {"pregunta": "Fecha de finalización de llenado de la información", "tipo": "autollenado", "source": "fecha_fin_llenado"},
    ],
    "Sustento y planificación": [
        {"pregunta": "A.1 Normatividad respecto de la cual se fortalecerá en esos procesos de gestión a los servidores de las DRE/GRE y UGEL.\nIndicar número y denominación del documento.", "tipo": "texto"},
        {"pregunta": "A.2 Colocar el link de ubicación del plan de trabajo de la capacitación.", "tipo": "texto"},
        {"pregunta": "A.3 Identificación del problema de gestión que motivó la capacitación.", "tipo": "texto"},
        {"pregunta": "A.4 Identificación de las causas del problema en términos de desempeño indicando la población objetivo.", "tipo": "texto"},
        {"pregunta": "A.5 Aplicó evaluación de situación de proceso o de la situación del del desempeño identificado para fortalecer.", "tipo": "seleccion", "opciones": ["Sí", "No"]},
        {"pregunta": "A.6 Valores de indicadores de cumplimiento o calidad respecto al proceso a intervenir", "tipo": "texto"},
        {"pregunta": "A.7 Fecha de inicio de la coordinación con el área usuaria o inicio de propuesta propia DIFOCA (momento del diagnóstico)", "tipo": "fecha"},
        {"pregunta": "A.8 Fecha de inicio de producción de recursos de aprendizaje.", "tipo": "fecha"},
        {"pregunta": "A.9 Grupo objetivo", "tipo": "autollenado", "source": "publico_objetivo"},
        {"pregunta": "A.10 Objetivo de la capacitación", "tipo": "autollenado", "source": "objetivo_capacitacion"},
        {"pregunta": "A.11 Desempeño a fortalecer", "tipo": "texto"},
        {"pregunta": "A.12 Cambio intencionado expresado en indicadores de gestión en el proceso intervenido", "tipo": "texto"},
        {"pregunta": "A.13 Duración (horas cronológicas)", "tipo": "autollenado", "source": "horas_certificacion"},
        {"pregunta": "A.14 Fecha de inicio planificada", "tipo": "autollenado", "source": "implementacion_inicio"},
        {"pregunta": "A.15 Fecha de finalización planificada", "tipo": "autollenado", "source": "implementacion_final"},
        {"pregunta": "A.16 Modalidad", "tipo": "texto"},
        {"pregunta": "A.17 Fórmula de evaluación", "tipo": "texto"},
        {"pregunta": "A.18 Indicar cuánto recursos se estimó para el desarrollo e implementación de esta capacitación (según corresponda)", "tipo": "texto"},
        {"pregunta": "A.19 Describir la manera en la que esta capacitación aseguró el cumplimiento de actividades y el logro de los aprendizajes y los desempeños fortalecidos (A través de la entrega de producto, reuniones de asesoría, tutoría y retroalimentación, otros.)", "tipo": "texto"},
        {"pregunta": "A.20 Indicar qué equipos y cómo harán el seguimiento de la capacitación en cuanto a cumplimiento, y el seguimiento en cuanto al logro de los desempeños fortalecidos", "tipo": "texto"},
        {"pregunta": "A.21 Metas propuestas de proceso, producto y resultado formuladas en el diseño o planificación de la capacitación.\nMetas de desempeño, producto, resultado e impacto (según corresponda)", "tipo": "texto"},
        {"pregunta": "A.22 Producto vinculado al desempeño o aprendizaje mejorado (parte de la evaluación)", "tipo": "texto"},
        {"pregunta": "A.23 Resultados cuanti y cualitativos programados", "tipo": "texto"},
    ],
    "Implementación y seguimiento": [
        {"pregunta": "B.1 Fecha de inicio real", "tipo": "fecha"},
        {"pregunta": "B.2 Fecha de finalización real", "tipo": "fecha"},
        {"pregunta": "B.3 Insertar link de Guía del participante, Cronograma, y otros documentos orientadores de la capacitación.", "tipo": "texto"},
        {"pregunta": "B.4 Mecanismos de convocatoria", "tipo": "texto"},
        {"pregunta": "B.5 Comunicación formal a las DRE/GRE y UGEL respecto de la convocatoria y la lista de los participantes.\nIndicar los documentos emitidos y la fecha de su remisión.", "tipo": "texto"},
        {"pregunta": "B.6 Mecanismos de seguimiento al cumplimiento de actividades así como al logro del desempeño a fortalecer.", "tipo": "texto"},
        {"pregunta": "B.7 En caso de contar con monitoreo académico o tutoría y retroalimentación, insertar link de guía, protocolo o documento orientador para su aplicación.\nIndicar cómo se socializó para su aplicación adecuada y homologada.", "tipo": "texto"},
        {"pregunta": "B.8 Dificultades identificadas en la implementación (plataforma, seguimiento, actividades, reporte, otras prestaciones, etc.)", "tipo": "texto"},
        {"pregunta": "B.9 Soluciones implementadas (plataforma, seguimiento, actividades, reporte, otras prestaciones, etc.)", "tipo": "texto"},
        {"pregunta": "B.10 Describir cómo se ha socializado el aprendizaje de estas recomendaciones con el equipo, y/o cómo se ha integrado al procedimiento de la gestión de la capacitación.", "tipo": "texto"},
        {"pregunta": "B.11 N° Postulaciones", "tipo": "autollenado", "source": "kpi_postulaciones"},
        {"pregunta": "B.12 Matriculaciones (N° y % de matriculados respecto a postulantes)", "tipo": "autollenado", "source": "kpi_matriculaciones_pct"},
        {"pregunta": "B.13 Participaciones (N° y % de participantes respecto a matriculados)", "tipo": "autollenado", "source": "kpi_participaciones_pct"},
        {"pregunta": "B.14 Finalizaciones (N° y % de finalizados respecto a participantes)", "tipo": "autollenado", "source": "kpi_finalizaciones_pct"},
    ],
    "Evaluación": [
        {"pregunta": "C.1 Cobertura de DRE y UGEL en participantes (# DRE coberturadas y # UGEL coberturadas)", "tipo": "autollenado", "source": "kpi_cobertura_participantes_dual"},
        {"pregunta": "C.2 Cobertura de DRE y UGEL en aprobados (# DRE fortalecidas y # UGEL fortalecidas)", "tipo": "autollenado", "source": "kpi_cobertura_aprobados_dual"},
        {"pregunta": "C.3 N° y % de servidores que aprueban respecto a finalizados", "tipo": "autollenado", "source": "kpi_certificaciones_pct"},
        {"pregunta": "C.4 % Servidores que muestran progreso / certifican", "tipo": "autollenado", "source": "kpi_tasa_progreso_fmt"},
        {"pregunta": "C.5 Índice de satisfacción expresado en % general y por aspectos", "tipo": "autollenado", "source": "kpi_satisfaccion_aspectos"},
        {"pregunta": "C.6 Dato de transferencia o puesta en práctica del aprendizaje o desempeño en % de DRE o UGEL.", "tipo": "texto"},
        {"pregunta": "C.7 Indicador de hito o producto del bien o servicio que se hará seguimiento para la evaluación de impacto de la capacitación.", "tipo": "texto"},
        {"pregunta": "C.8 Indicar cómo los recursos de aprendizaje, casuística u otras actividades que formaron parte de la capacitación contribuyeron al logro de los desempeños fortalecidos.", "tipo": "texto"},
        {"pregunta": "C.9 Indicar si hubo cambio de calendarización.", "tipo": "seleccion", "opciones": ["Sí", "No"]},
        {"pregunta": "C.10 Indicar en qué etapa se aplicaron los cambios: módulo o sesión inicial (compromiso), en la implementación general, en las actividades de cierre de hitos, módulos, o en el módulo o sesión de cierre. Indicar si fue en todas y qué lo motivó.", "tipo": "texto"},
        {"pregunta": "C.11 Indicar si hubo cambio en la cantidad de horas inicialmente planificada.", "tipo": "seleccion", "opciones": ["Sí", "No"]},
        {"pregunta": "C.12 Indicar si hubo cambio en la fórmula de evaluación", "tipo": "seleccion", "opciones": ["Sí", "No"]},
        {"pregunta": "C.13 Indicar la inversión fuera del costo de personal en la implementación de esta capacitación.", "tipo": "texto"},
        {"pregunta": "C.14 Mencionar el índice del Alfa de Crombach aplicado a las herramientas de evaluación inicial/final.", "tipo": "texto"},
        {"pregunta": "C.15 Metas propuestas de dsempeño, producto, resultado e impacto.", "tipo": "texto"},
        {"pregunta": "C.16 Fecha de publicación de los certificados de aprobación de la capacitación", "tipo": "fecha"},
        {"pregunta": "C.17 Documento formal informativo de la situación final de los participantes dirigido a las DRE/GRE y UGEL. Indicar la fecha de su remisión.", "tipo": "fecha"},
        {"pregunta": "C.18 Indicar fecha y conclusiones de la reunión de balance con la UO o IGED contraparte de la capacitación.", "tipo": "fecha"},
        {"pregunta": "C.19 Resultados cuanti y cualitativos alcanzados. Indicar si se hizo alguna evaluación a un grupo de control.", "tipo": "texto"},
        {"pregunta": "C.20 Balance y recomendaciones aplicados a la gestión de la capacitación", "tipo": "texto"},
    ],
}

# Mapeo de preguntas a estandares de calidad (usado en el reporte de analisis).
MAPEO_ESTANDARES = {
    "A.1": "Pertinencia", "A.2": "Pertinencia", "A.3": "Pertinencia", "A.4": "Pertinencia",
    "A.5": "Pertinencia", "A.6": "Pertinencia", "A.7": "Pertinencia", "A.8": "Pertinencia",
    "A.9": "Pertinencia", "A.10": "Pertinencia", "A.11": "Pertinencia", "A.12": "Pertinencia",
    "A.13": "Eficacia", "A.14": "Eficacia", "A.15": "Eficacia", "A.16": "Eficacia",
    "A.17": "Eficacia", "A.18": "Eficacia", "A.19": "Eficacia", "A.20": "Eficacia",
    "A.21": "Mejora continua", "A.22": "Mejora continua", "A.23": "Mejora continua",
    "B.1": "Eficacia", "B.2": "Eficacia", "B.3": "Eficacia",
    "B.4": "Eficacia", "B.5": "Eficacia", "B.6": "Eficacia", "B.7": "Eficacia",
    "B.8": "Mejora continua", "B.9": "Mejora continua", "B.10": "Mejora continua",
    "B.11": "Eficacia", "B.12": "Eficacia", "B.13": "Eficacia", "B.14": "Eficacia",
    "C.1": "Eficacia", "C.2": "Eficacia", "C.3": "Eficacia", "C.4": "Eficacia",
    "C.5": "Eficacia", "C.6": "Eficacia", "C.7": "Eficacia", "C.8": "Eficacia",
    "C.9": "Eficacia", "C.10": "Eficacia", "C.11": "Eficacia", "C.12": "Eficacia",
    "C.13": "Eficacia", "C.14": "Eficacia", "C.15": "Eficacia", "C.16": "Eficacia",
    "C.17": "Eficacia", "C.18": "Eficacia", "C.19": "Mejora continua", "C.20": "Mejora continua",
}

MAPEO_ETAPAS = {
    "A.1": "Diagnóstico y sustento", "A.2": "Diagnóstico y sustento",
    "A.3": "Diagnóstico y sustento", "A.4": "Diagnóstico y sustento",
    "A.5": "Diagnóstico y sustento", "A.6": "Diagnóstico y sustento",
    "A.7": "Diagnóstico y sustento",
    "A.8": "Diseño/planificación", "A.9": "Diseño/planificación",
    "A.10": "Diseño/planificación", "A.11": "Diseño/planificación",
    "A.12": "Diseño/planificación", "A.13": "Diseño/planificación",
    "A.14": "Diseño/planificación", "A.15": "Diseño/planificación",
    "A.16": "Diseño/planificación", "A.17": "Diseño/planificación",
    "A.18": "Producción y generación de condiciones",
    "A.19": "Producción y generación de condiciones",
    "A.20": "Producción y generación de condiciones",
    "A.21": "Producción y generación de condiciones",
    "A.22": "Producción y generación de condiciones",
    "A.23": "Producción y generación de condiciones",
    "B.1": "Implementación", "B.2": "Implementación", "B.3": "Implementación",
    "B.4": "Implementación", "B.5": "Implementación", "B.6": "Implementación",
    "B.7": "Implementación", "B.8": "Implementación", "B.9": "Implementación",
    "B.10": "Implementación", "B.11": "Implementación", "B.12": "Implementación",
    "B.13": "Implementación", "B.14": "Implementación",
    "C.1": "Evaluación y documentación", "C.2": "Evaluación y documentación",
    "C.3": "Evaluación y documentación", "C.4": "Evaluación y documentación",
    "C.5": "Evaluación y documentación", "C.6": "Evaluación y documentación",
    "C.7": "Evaluación y documentación", "C.8": "Evaluación y documentación",
    "C.9": "Evaluación y documentación", "C.10": "Evaluación y documentación",
    "C.11": "Evaluación y documentación", "C.12": "Evaluación y documentación",
    "C.13": "Evaluación y documentación", "C.14": "Evaluación y documentación",
    "C.15": "Evaluación y documentación", "C.16": "Evaluación y documentación",
    "C.17": "Evaluación y documentación", "C.18": "Evaluación y documentación",
    "C.19": "Evaluación y documentación", "C.20": "Evaluación y documentación",
}

_EXTRAER_ITEM_RE = re.compile(r"^([ABC]\.\d+)")


# ---------------------------------------------------------------------------
# CRUD
# ---------------------------------------------------------------------------

def obtener_anios_disponibles() -> list[int]:
    """Retorna lista de anios con capacitaciones cerradas o en implementacion."""
    try:
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT DISTINCT anio
                    FROM oferta_formativa_difoca
                    WHERE anio IS NOT NULL AND condicion IN ('Cerrado', 'En implementacion')
                    ORDER BY anio DESC
                    """
                )
                return [int(f["anio"]) for f in cur.fetchall() if f.get("anio")]
    except Exception:
        return [ANIO_VIGENTE]


def obtener_procesos_formativos(
    nombre_especialista: str | None,
    anio: int | None = None,
) -> list[dict[str, Any]]:
    """Retorna procesos formativos visibles para el especialista."""
    anio_filtro = anio or ANIO_VIGENTE
    try:
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT codigo, tipo_proceso_formativo, denominacion_proceso_formativo,
                           especialista_cargo, anio, condicion,
                           publico_objetivo, objetivo_capacitacion,
                           horas_certificacion, implementacion_inicio, implementacion_final
                    FROM oferta_formativa_difoca
                    WHERE anio = %s AND condicion IN ('Cerrado', 'En implementacion')
                    ORDER BY codigo
                    """,
                    (anio_filtro,),
                )
                filas = list(cur.fetchall())
    except Exception:
        return []

    for f in filas:
        f["proceso_combinado"] = (
            f"{f.get('tipo_proceso_formativo', '')} {f.get('denominacion_proceso_formativo', '')}".strip()
        )

    if nombre_especialista == "(OS) Chrystian Varillas":
        return filas

    if nombre_especialista:
        return [f for f in filas if f.get("especialista_cargo") == nombre_especialista]

    return []


def obtener_datos_proceso(codigo: str) -> dict[str, Any]:
    """Retorna metadatos de un proceso formativo para autollenado."""
    try:
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT tipo_proceso_formativo, denominacion_proceso_formativo,
                           especialista_cargo, publico_objetivo, objetivo_capacitacion,
                           horas_certificacion, implementacion_inicio, implementacion_final
                    FROM oferta_formativa_difoca
                    WHERE codigo = %s
                    LIMIT 1
                    """,
                    (codigo,),
                )
                fila = cur.fetchone()
        if fila:
            fila["proceso_combinado"] = (
                f"{fila.get('tipo_proceso_formativo', '')} {fila.get('denominacion_proceso_formativo', '')}".strip()
            )
            return fila
    except Exception:
        pass
    return {}


def cargar_respuestas(codigo: str, capitulo: str) -> dict[str, str]:
    """Carga las ultimas respuestas registradas por pregunta (cualquier usuario)."""
    try:
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT t1.pregunta, t1.respuesta
                    FROM estandares_calidad t1
                    INNER JOIN (
                        SELECT codigo, capitulo, pregunta, MAX(fecha_guardado) AS max_fecha
                        FROM estandares_calidad
                        WHERE codigo = %s AND capitulo = %s
                        GROUP BY codigo, capitulo, pregunta
                    ) t2
                    ON t1.codigo = t2.codigo AND t1.capitulo = t2.capitulo
                       AND t1.pregunta = t2.pregunta AND t1.fecha_guardado = t2.max_fecha
                    WHERE t1.codigo = %s AND t1.capitulo = %s
                    """,
                    (codigo, capitulo, codigo, capitulo),
                )
                filas = cur.fetchall()
        return {f["pregunta"]: f["respuesta"] for f in filas}
    except Exception:
        return {}


def guardar_respuestas(
    codigo: str,
    usuario: str,
    capitulo: str,
    respuestas: dict[str, str],
) -> None:
    """Guarda o actualiza respuestas para un codigo/capitulo."""
    ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_connection() as conn:
        with conn.cursor() as cur:
            for pregunta, respuesta in respuestas.items():
                cur.execute(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM estandares_calidad
                    WHERE codigo = %s AND usuario = %s AND capitulo = %s AND pregunta = %s
                    """,
                    (codigo, usuario, capitulo, pregunta),
                )
                existe = cur.fetchone()["cnt"] > 0
                if existe:
                    cur.execute(
                        """
                        UPDATE estandares_calidad
                        SET respuesta = %s, fecha_guardado = %s
                        WHERE codigo = %s AND usuario = %s AND capitulo = %s AND pregunta = %s
                        """,
                        (str(respuesta), ahora, codigo, usuario, capitulo, pregunta),
                    )
                else:
                    cur.execute(
                        """
                        INSERT INTO estandares_calidad
                            (codigo, usuario, capitulo, pregunta, respuesta, fecha_guardado)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        """,
                        (codigo, usuario, capitulo, pregunta, str(respuesta), ahora),
                    )
        conn.commit()


def eliminar_respuestas(
    codigo: str,
    capitulo: str,
    usuario: str | None = None,
) -> int:
    """Elimina respuestas. Si usuario es None elimina todas las del codigo/capitulo."""
    with get_connection() as conn:
        with conn.cursor() as cur:
            if usuario:
                cur.execute(
                    "DELETE FROM estandares_calidad WHERE codigo = %s AND capitulo = %s AND usuario = %s",
                    (codigo, capitulo, usuario),
                )
            else:
                cur.execute(
                    "DELETE FROM estandares_calidad WHERE codigo = %s AND capitulo = %s",
                    (codigo, capitulo),
                )
            borrados = cur.rowcount
        conn.commit()
    return borrados


def obtener_capitulos_existentes(codigo: str) -> set[str]:
    """Retorna set con los capitulos que ya tienen respuestas para un codigo."""
    try:
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT DISTINCT capitulo FROM estandares_calidad WHERE codigo = %s",
                    (codigo,),
                )
                return {f["capitulo"] for f in cur.fetchall()}
    except Exception:
        return set()


# ---------------------------------------------------------------------------
# KPI para campos autollenado
# ---------------------------------------------------------------------------

def calcular_kpis(codigo: str) -> dict[str, Any]:
    """Calcula KPIs de una capacitacion para rellenar campos automaticos."""
    codigo = str(codigo).strip()
    if not codigo:
        return {}

    try:
        conn = get_connection()
        try:
            df_bbdd = pd.read_sql(
                "SELECT * FROM bbdd_difoca WHERE codigo = %s", conn, params=(codigo,)
            )
            df_oferta = pd.read_sql(
                "SELECT * FROM oferta_formativa_difoca WHERE codigo = %s", conn, params=(codigo,)
            )
            try:
                df_satisf = pd.read_sql(
                    "SELECT * FROM satisfaccion WHERE codigo = %s", conn, params=(codigo,)
                )
            except Exception:
                df_satisf = pd.DataFrame()
        finally:
            conn.close()

        if df_bbdd.empty or df_oferta.empty:
            return {}

        merged = pd.merge(df_bbdd, df_oferta, on="codigo", how="right", suffixes=("_b", "_o"))

        kpi: dict[str, Any] = {}

        postulaciones = int(merged["dni"].nunique())
        matriculaciones = int((merged["estado"] == 2).sum())
        kpi["kpi_postulaciones"] = postulaciones

        if postulaciones > 0:
            pct = round(matriculaciones / postulaciones * 100, 1)
            kpi["kpi_matriculaciones_pct"] = f"{matriculaciones} ({pct}%)"
        else:
            kpi["kpi_matriculaciones_pct"] = str(matriculaciones)

        estado_num = pd.to_numeric(merged.get("estado"), errors="coerce")
        compromiso_num = pd.to_numeric(merged.get("compromiso"), errors="coerce")
        aprobados_flag = pd.to_numeric(merged.get("aprobados_certificados"), errors="coerce").fillna(0) >= 1
        desaprobado_perm_flag = pd.to_numeric(merged.get("desaprobado_permanente"), errors="coerce").fillna(0) >= 1
        base_participa = (estado_num == 2) & (compromiso_num.isin([20, 1]))

        participaciones = int(base_participa.sum())

        if matriculaciones > 0:
            pct_part = round(participaciones / matriculaciones * 100, 1)
            kpi["kpi_participaciones_pct"] = f"{participaciones} ({pct_part}%)"
        else:
            kpi["kpi_participaciones_pct"] = str(participaciones)

        merged["Finalizados"] = (
            base_participa & (aprobados_flag | desaprobado_perm_flag)
        ).astype(int)
        finalizaciones = int((merged["Finalizados"] == 1).sum())

        if participaciones > 0:
            pct_fin = round(finalizaciones / participaciones * 100, 1)
            kpi["kpi_finalizaciones_pct"] = f"{finalizaciones} ({pct_fin}%)"
        else:
            kpi["kpi_finalizaciones_pct"] = str(finalizaciones)

        certificaciones = int((base_participa & aprobados_flag).sum())

        if finalizaciones > 0:
            pct_cert = round(certificaciones / finalizaciones * 100, 1)
            kpi["kpi_certificaciones_pct"] = f"{certificaciones} ({pct_cert}%)"
        else:
            kpi["kpi_certificaciones_pct"] = str(certificaciones)

        # Cobertura DRE/UGEL participantes
        if "nombre_iged" in merged.columns and "tipo_iged" in merged.columns:
            nombres_iged = _normalize_iged_name(merged["nombre_iged"])
            tipo_norm = merged["tipo_iged"].astype(str).str.upper().str.replace(" ", "", regex=False)
            dre_cob = _count_distinct_non_empty(nombres_iged[base_participa & (tipo_norm == "DRE/GRE")])
            ugel_cob = _count_distinct_non_empty(nombres_iged[base_participa & (tipo_norm == "UGEL")])
            kpi["kpi_cobertura_participantes_dual"] = f"{dre_cob} DRE/GRE, {ugel_cob} UGEL"
        else:
            kpi["kpi_cobertura_participantes_dual"] = ""

        # Cobertura DRE/UGEL aprobados
        if "nombre_iged" in merged.columns and "tipo_iged" in merged.columns:
            nombres_iged = _normalize_iged_name(merged["nombre_iged"])
            tipo_norm = merged["tipo_iged"].astype(str).str.upper().str.replace(" ", "", regex=False)
            dre_fort = _count_distinct_non_empty(nombres_iged[base_participa & (tipo_norm == "DRE/GRE") & aprobados_flag])
            ugel_fort = _count_distinct_non_empty(nombres_iged[base_participa & (tipo_norm == "UGEL") & aprobados_flag])
            kpi["kpi_cobertura_aprobados_dual"] = f"{dre_fort} DRE/GRE, {ugel_fort} UGEL"
        else:
            kpi["kpi_cobertura_aprobados_dual"] = ""

        # Tasa de progreso
        ev_prog_cert = (
            (merged.get("ev_progreso_aprendizaje", pd.Series(dtype="int")) == 1)
            & (merged.get("aprobados_certificados", pd.Series(dtype="int")) == 1)
        ).sum()
        progreso_cert = (
            (merged.get("progreso", pd.Series(dtype="int")) == 1)
            & (merged.get("aprobados_certificados", pd.Series(dtype="int")) == 1)
        ).sum()
        if ev_prog_cert > 0:
            tasa = progreso_cert / ev_prog_cert
            kpi["kpi_tasa_progreso_fmt"] = f"{round(tasa * 100)}%"
        else:
            kpi["kpi_tasa_progreso_fmt"] = ""

        # Satisfaccion
        if not df_satisf.empty and "satisfaccion" in df_satisf.columns:
            total_sat = len(df_satisf)
            total_satisfecho = int(
                df_satisf["satisfaccion"]
                .str.strip()
                .str.lower()
                .isin(["satisfecho", "satisfecha"])
                .sum()
            )
            val = total_satisfecho / total_sat if total_sat > 0 else 0

            if "aspecto" in df_satisf.columns:
                aspecto_stats = (
                    df_satisf.groupby("aspecto")
                    .agg(
                        total=("satisfaccion", "count"),
                        satisfecho=(
                            "satisfaccion",
                            lambda x: x.str.strip().str.lower().isin(["satisfecho", "satisfecha"]).sum(),
                        ),
                    )
                    .reset_index()
                )
                aspecto_stats["pct"] = (aspecto_stats["satisfecho"] / aspecto_stats["total"] * 100).round(2)
                partes = [f"{row['aspecto']}: {row['pct']:.2f}%" for _, row in aspecto_stats.iterrows()]
                kpi["kpi_satisfaccion_aspectos"] = f"Global: {val * 100:.2f}% | " + " | ".join(partes)
            else:
                kpi["kpi_satisfaccion_aspectos"] = f"Global: {val * 100:.2f}%"
        else:
            kpi["kpi_satisfaccion_aspectos"] = ""

        return kpi
    except Exception:
        return {}


def resolver_valor_autollenado(
    source: str,
    datos_proceso: dict[str, Any],
    kpis: dict[str, Any],
    nombre_especialista: str | None = None,
) -> str:
    """Resuelve el valor de un campo autollenado segun su source."""
    if source == "proceso_sel":
        return str(datos_proceso.get("proceso_combinado", ""))
    if source == "nombre_especialista":
        return nombre_especialista or ""
    if source == "fecha_finalizacion":
        return str(datos_proceso.get("implementacion_final", ""))
    if source == "fecha_inicio_llenado":
        return datetime.now().strftime("%Y-%m-%d")
    if source == "fecha_fin_llenado":
        return datetime.now().strftime("%Y-%m-%d")
    if source.startswith("kpi_"):
        v = kpis.get(source, "")
        return str(v) if v is not None else ""
    val = datos_proceso.get(source, "")
    return str(val) if val is not None else ""


# ---------------------------------------------------------------------------
# Generacion de reporte Excel de analisis (5 hojas) - replica tab3 Streamlit
# ---------------------------------------------------------------------------

def _extraer_item(pregunta: str) -> str | None:
    m = _EXTRAER_ITEM_RE.match(str(pregunta).strip())
    return m.group(1) if m else None


def _calcular_completitud(df_resp: pd.DataFrame, mapeo: dict[str, str]) -> dict[str, dict[str, Any]]:
    """Calcula completitud agrupada por las categorias del mapeo dado."""
    categorias = sorted(set(mapeo.values()))
    resultados: dict[str, dict[str, Any]] = {}

    for cat in categorias:
        preguntas_cat = [k for k, v in mapeo.items() if v == cat]
        total_esperado = 0
        total_completado = 0

        for codigo in df_resp["codigo"].unique():
            df_cod = df_resp[df_resp["codigo"] == codigo]
            for item in preguntas_cat:
                mask = df_cod["pregunta"].str.startswith(item + " ")
                if mask.any():
                    total_esperado += 1
                    resp = df_cod.loc[mask, "respuesta"].iloc[0]
                    if resp and str(resp).strip() and str(resp).strip().lower() != "nan":
                        total_completado += 1

        porcentaje = (total_completado / total_esperado * 100) if total_esperado > 0 else 0
        resultados[cat] = {
            "completitud": total_completado,
            "total": total_esperado,
            "porcentaje": porcentaje,
        }
    return resultados


def generar_reporte_analisis(anio: int | None = None) -> bytes | None:
    """Genera reporte Excel profesional con 5 hojas (replica generar_reporte_excel_estandares)."""
    anio_filtro = anio or ANIO_VIGENTE

    try:
        conn = get_connection()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT ec.codigo, ec.pregunta, ec.respuesta, ec.fecha_guardado,
                           of2.denominacion_proceso_formativo, of2.especialista_cargo,
                           of2.capacitacion_presencialidad
                    FROM estandares_calidad ec
                    LEFT JOIN oferta_formativa_difoca of2 ON ec.codigo = of2.codigo
                    WHERE of2.anio = %s AND of2.condicion = 'Cerrado'
                    ORDER BY ec.codigo, ec.pregunta
                    """,
                    (anio_filtro,),
                )
                rows = cur.fetchall()
        finally:
            conn.close()

        if not rows:
            return None

        df_resp = pd.DataFrame(rows)

        buf = BytesIO()
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        hdr_font = Font(bold=True, color="FFFFFF", size=12)
        hdr_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        data_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        pct_align = Alignment(horizontal="center", vertical="center")
        borde = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        green_fill = PatternFill(start_color="D4E6B7", end_color="D4E6B7", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        red_fill = PatternFill(start_color="F8CECC", end_color="F8CECC", fill_type="solid")

        def pct_fill(pct: float) -> PatternFill:
            if pct >= 90:
                return green_fill
            if pct >= 70:
                return yellow_fill
            return red_fill

        # -- Hoja 1: Resumen General -----------------------------------------
        ws1 = wb.create_sheet("📊 Resumen General")
        ws1["A1"] = "REPORTE DE ANÁLISIS DE ESTÁNDARES DE CALIDAD"
        ws1["A1"].font = Font(bold=True, size=16, color="2E86AB")
        ws1.merge_cells("A1:F1")
        ws1["A2"] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws1["A2"].font = Font(italic=True, size=10)
        ws1.merge_cells("A2:F2")

        total_caps = int(df_resp["codigo"].nunique())
        total_resp = int(df_resp["respuesta"].notna().sum() & (df_resp["respuesta"] != "").sum())
        total_resp = int(
            len(df_resp[df_resp["respuesta"].notna() & (df_resp["respuesta"] != "")])
        )
        total_preg = int(df_resp["pregunta"].nunique())

        ws1["A4"] = "MÉTRICAS GENERALES"
        ws1["A4"].font = hdr_font
        ws1["A4"].fill = hdr_fill
        ws1.merge_cells("A4:B4")

        metricas = [
            ("Total de Capacitaciones Cerradas", total_caps),
            ("Total de Preguntas del Formulario", total_preg),
            ("Total de Respuestas Completadas", total_resp),
            (
                "Porcentaje de Completitud General",
                f"{(total_resp / (total_caps * total_preg) * 100):.1f}%"
                if total_caps * total_preg > 0
                else "0.0%",
            ),
        ]
        for i, (met, val) in enumerate(metricas, start=5):
            ws1[f"A{i}"] = met
            ws1[f"B{i}"] = val
            ws1[f"A{i}"].font = Font(bold=True)

        # -- Hoja 2: Por Estandares -------------------------------------------
        resultado_est = _calcular_completitud(df_resp, MAPEO_ESTANDARES)
        ws2 = wb.create_sheet("🎯 Por Estándares")
        ws2["A1"] = "ANÁLISIS POR ESTÁNDARES DE CALIDAD"
        ws2["A1"].font = Font(bold=True, size=14, color="2E86AB")
        ws2.merge_cells("A1:D1")

        headers_est = ["Estándar", "Completadas", "Total", "Porcentaje"]
        for ci, h in enumerate(headers_est, 1):
            c = ws2.cell(row=3, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = hdr_align
            c.border = borde

        for ri, (est, datos) in enumerate(resultado_est.items(), 4):
            ws2.cell(row=ri, column=1, value=est).border = borde
            ws2.cell(row=ri, column=2, value=datos["completitud"]).border = borde
            ws2.cell(row=ri, column=3, value=datos["total"]).border = borde
            pc = ws2.cell(row=ri, column=4, value=f"{datos['porcentaje']:.1f}%")
            pc.border = borde
            pc.alignment = pct_align
            pc.fill = pct_fill(datos["porcentaje"])

        # -- Hoja 3: Por Etapas -----------------------------------------------
        resultado_eta = _calcular_completitud(df_resp, MAPEO_ETAPAS)
        ws3 = wb.create_sheet("📈 Por Etapas")
        ws3["A1"] = "ANÁLISIS POR ETAPAS DEL PROCESO"
        ws3["A1"].font = Font(bold=True, size=14, color="2E86AB")
        ws3.merge_cells("A1:D1")

        headers_eta = ["Etapa", "Completadas", "Total", "Porcentaje"]
        for ci, h in enumerate(headers_eta, 1):
            c = ws3.cell(row=3, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = hdr_align
            c.border = borde

        etapas_orden = [
            "Diagnóstico y sustento",
            "Diseño/planificación",
            "Producción y generación de condiciones",
            "Implementación",
            "Evaluación y documentación",
        ]
        for ri, etapa in enumerate(etapas_orden, 4):
            datos = resultado_eta.get(etapa, {"completitud": 0, "total": 0, "porcentaje": 0})
            ws3.cell(row=ri, column=1, value=etapa).border = borde
            ws3.cell(row=ri, column=2, value=datos["completitud"]).border = borde
            ws3.cell(row=ri, column=3, value=datos["total"]).border = borde
            pc = ws3.cell(row=ri, column=4, value=f"{datos['porcentaje']:.1f}%")
            pc.border = borde
            pc.alignment = pct_align
            pc.fill = pct_fill(datos["porcentaje"])

        # -- Hoja 4: Por Especialista ------------------------------------------
        ws4 = wb.create_sheet("👥 Por Especialista")
        ws4["A1"] = "ANÁLISIS POR ESPECIALISTA"
        ws4["A1"].font = Font(bold=True, size=14, color="2E86AB")
        ws4.merge_cells("A1:E1")

        headers_esp = ["Especialista", "Capacitaciones", "Con Estándares", "Completitud", "Porcentaje"]
        for ci, h in enumerate(headers_esp, 1):
            c = ws4.cell(row=3, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = hdr_align
            c.border = borde

        esp_data: list[dict[str, Any]] = []
        for especialista in df_resp["especialista_cargo"].dropna().unique():
            caps_esp = df_resp[df_resp["especialista_cargo"] == especialista]["codigo"].unique()
            total_c = len(caps_esp)
            caps_con = 0
            t_resp_e = 0
            t_preg_e = 0
            for cod in caps_esp:
                df_c = df_resp[df_resp["codigo"] == cod]
                resp_ok = int(len(df_c[df_c["respuesta"].notna() & (df_c["respuesta"] != "")]))
                t_preg_c = len(df_c)
                if resp_ok > 0:
                    caps_con += 1
                t_resp_e += resp_ok
                t_preg_e += t_preg_c
            pct_e = (t_resp_e / t_preg_e * 100) if t_preg_e > 0 else 0
            esp_data.append(
                {
                    "especialista": especialista,
                    "total": total_c,
                    "con": caps_con,
                    "completitud": f"{t_resp_e}/{t_preg_e}",
                    "porcentaje": pct_e,
                }
            )

        esp_data.sort(key=lambda x: x["porcentaje"], reverse=True)

        for ri, d in enumerate(esp_data, 4):
            ws4.cell(row=ri, column=1, value=d["especialista"]).border = borde
            ws4.cell(row=ri, column=2, value=d["total"]).border = borde
            ws4.cell(row=ri, column=3, value=d["con"]).border = borde
            ws4.cell(row=ri, column=4, value=d["completitud"]).border = borde
            pc = ws4.cell(row=ri, column=5, value=f"{d['porcentaje']:.1f}%")
            pc.border = borde
            pc.alignment = pct_align
            pc.fill = pct_fill(d["porcentaje"])

        # -- Hoja 5: Datos Detallados ------------------------------------------
        ws5 = wb.create_sheet("📋 Datos Detallados")
        df_det = df_resp.copy()
        df_det["modalidad"] = df_det["capacitacion_presencialidad"].apply(
            lambda x: "Presencial" if x == 1 else "Virtual"
        )
        df_det = df_det[
            ["codigo", "denominacion_proceso_formativo", "especialista_cargo", "modalidad", "pregunta", "respuesta", "fecha_guardado"]
        ]
        df_det.columns = ["Código", "Denominación", "Especialista", "Modalidad", "Pregunta", "Respuesta", "Fecha"]

        ws5["A1"] = "DATOS DETALLADOS - TODAS LAS RESPUESTAS"
        ws5["A1"].font = Font(bold=True, size=14, color="2E86AB")
        ws5.merge_cells("A1:G1")

        for r_idx, row in enumerate(dataframe_to_rows(df_det, index=False, header=True), 3):
            for c_idx, value in enumerate(row, 1):
                cell = ws5.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:
                    cell.font = hdr_font
                    cell.fill = hdr_fill
                    cell.alignment = hdr_align
                cell.border = borde
                if c_idx == 6:
                    cell.alignment = data_align

        col_widths = {"A": 15, "B": 40, "C": 20, "D": 15, "E": 15, "F": 25, "G": 15}
        for ws in wb.worksheets:
            for letter, w in col_widths.items():
                try:
                    ws.column_dimensions[letter].width = w
                except Exception:
                    pass

        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    except Exception:
        return None


# ---------------------------------------------------------------------------
# Generacion de reporte Excel individual por capacitacion
# ---------------------------------------------------------------------------

def generar_reporte_individual(codigo: str, anio: int | None = None) -> bytes | None:
    """Genera reporte individual de una capacitacion (sin usar plantilla)."""
    try:
        conn = get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT t1.capitulo, t1.pregunta, t1.respuesta
                FROM estandares_calidad t1
                INNER JOIN (
                    SELECT codigo, capitulo, pregunta, MAX(fecha_guardado) AS max_fecha
                    FROM estandares_calidad
                    WHERE codigo = %s
                    GROUP BY codigo, capitulo, pregunta
                ) t2
                ON t1.codigo = t2.codigo AND t1.capitulo = t2.capitulo
                   AND t1.pregunta = t2.pregunta AND t1.fecha_guardado = t2.max_fecha
                WHERE t1.codigo = %s
                ORDER BY t1.capitulo, t1.pregunta
                """,
                (codigo, codigo),
            )
            filas_resp = list(cursor.fetchall())

            cursor.execute(
                """
                SELECT tipo_proceso_formativo, denominacion_proceso_formativo
                FROM oferta_formativa_difoca WHERE codigo = %s LIMIT 1
                """,
                (codigo,),
            )
            info_proc = cursor.fetchone()
        finally:
            conn.close()

        if not filas_resp:
            return None

        nombre_proc = ""
        if info_proc:
            nombre_proc = f"{info_proc.get('tipo_proceso_formativo', '')} {info_proc.get('denominacion_proceso_formativo', '')}".strip()

        resp_por_cap: dict[str, dict[str, str]] = {}
        for f in filas_resp:
            cap = f["capitulo"]
            resp_por_cap.setdefault(cap, {})[f["pregunta"]] = f["respuesta"] or ""

        buf = BytesIO()
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        borde = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        wrap = Alignment(vertical="center", wrap_text=True)

        for cap in CAPITULOS:
            ws = wb.create_sheet(cap)
            ws["A1"] = cap.upper()
            ws["A1"].font = Font(bold=True, size=14, color="2E86AB")
            ws.merge_cells("A1:C1")
            ws["A2"] = f"Código: {codigo} — {nombre_proc}"
            ws["A2"].font = Font(italic=True, size=10)
            ws.merge_cells("A2:C2")

            ws.cell(row=3, column=1, value="N°").font = hdr_font
            ws.cell(row=3, column=1).fill = hdr_fill
            ws.cell(row=3, column=1).border = borde
            ws.cell(row=3, column=2, value="Pregunta").font = hdr_font
            ws.cell(row=3, column=2).fill = hdr_fill
            ws.cell(row=3, column=2).border = borde
            ws.cell(row=3, column=3, value="Respuesta").font = hdr_font
            ws.cell(row=3, column=3).fill = hdr_fill
            ws.cell(row=3, column=3).border = borde

            preguntas_def = PREGUNTAS_CAPITULOS.get(cap, [])
            respuestas_cap = resp_por_cap.get(cap, {})

            for idx, pdef in enumerate(preguntas_def, 1):
                preg_txt = pdef["pregunta"]
                resp_txt = respuestas_cap.get(preg_txt, "(Sin respuesta)")
                row = idx + 3
                ws.cell(row=row, column=1, value=idx).border = borde
                c_preg = ws.cell(row=row, column=2, value=preg_txt)
                c_preg.border = borde
                c_preg.alignment = wrap
                c_resp = ws.cell(row=row, column=3, value=resp_txt)
                c_resp.border = borde
                c_resp.alignment = wrap

            ws.column_dimensions["A"].width = 6
            ws.column_dimensions["B"].width = 60
            ws.column_dimensions["C"].width = 50

        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    except Exception:
        return None
