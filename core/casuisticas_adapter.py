"""Logica de negocio de las Casuisticas.

Una casuistica es un caso reportado por un Visor sobre un participante (DNI)
con problemas de inscripcion/matricula. Es una conversacion por turnos entre
el Visor y el administrador (`core.models.Casuistica`, `CasuisticaMensaje`,
`CasuisticaEvidencia`). Este modulo concentra las reglas del ciclo de vida
para que las vistas (portal del Visor y bandeja interna) solo orquesten HTTP.

Estados: ABIERTO -> EN_PLATAFORMA (solo admin, define accion) -> CERRADO
(solo admin). Un mensaje nuevo sobre un caso cerrado lo reabre; sobre un caso
"En plataforma" lo regresa a Abierto (ya no es solo "pendiente en plataforma").
"""

from __future__ import annotations

import io
from typing import Any, Iterable

from django.db import transaction
from django.utils import timezone

from core import drive_storage
from core.models import (
    AccionPlataforma,
    Casuistica,
    CasuisticaEvidencia,
    CasuisticaMensaje,
)

MAX_EVIDENCIA_BYTES = 8 * 1024 * 1024  # 8 MB por archivo adjunto.
TIPOS_EVIDENCIA_PERMITIDOS = {
    "image/jpeg", "image/png", "image/webp", "image/heic", "image/heif",
    "application/pdf",
}


class CasuisticaError(ValueError):
    """Error de validacion o de regla de negocio, mostrable al usuario."""


class CasuisticaDuplicada(CasuisticaError):
    """Ya existe un caso de este Visor para el mismo DNI en el mismo curso.

    Se evita abrir un hilo nuevo por error: `caso_existente` permite a la
    vista redirigir directo a la conversacion ya registrada.
    """

    def __init__(self, caso_existente: Casuistica):
        self.caso_existente = caso_existente
        super().__init__("Ya tienes un caso registrado para este DNI en este curso.")


def _validar_archivo(archivo) -> None:
    if archivo.size > MAX_EVIDENCIA_BYTES:
        raise CasuisticaError(f"'{archivo.name}' supera el tamaño máximo permitido (8 MB).")
    content_type = str(getattr(archivo, "content_type", "") or "")
    if content_type not in TIPOS_EVIDENCIA_PERMITIDOS:
        raise CasuisticaError(f"'{archivo.name}' no es un formato permitido (foto o PDF).")


def _adjuntar_evidencias(mensaje: CasuisticaMensaje, dni: str, archivos: Iterable[Any]) -> None:
    for archivo in archivos:
        contenido = archivo.read()
        file_id = drive_storage.subir_evidencia(
            nombre=f"{dni}_{archivo.name}",
            contenido=contenido,
            mime_type=str(getattr(archivo, "content_type", "")),
        )
        CasuisticaEvidencia.objects.create(
            mensaje=mensaje,
            drive_file_id=file_id,
            nombre_original=archivo.name,
            mime_type=str(getattr(archivo, "content_type", "")),
            tamano_bytes=len(contenido),
        )


def _voltear_turno(autor_tipo: str) -> str:
    """El turno pasa siempre al otro interlocutor."""
    return Casuistica.Turno.ADMIN if autor_tipo == CasuisticaMensaje.AutorTipo.VISOR else Casuistica.Turno.VISOR


def crear_caso(
    *, capacitacion, email_visor: str, nombre_visor: str, dni: str,
    nombre_participante: str, asunto: str, texto: str, archivos: Iterable[Any] = (),
) -> Casuistica:
    """El Visor reporta un caso nuevo: queda Abierto y en turno del administrador."""
    dni = str(dni or "").strip()
    if not dni.isdigit() or not (8 <= len(dni) <= 12):
        raise CasuisticaError("El documento debe tener solo números (8 a 12 dígitos).")

    caso_existente = (
        Casuistica.objects.filter(
            capacitacion=capacitacion, dni_participante=dni, email_visor__iexact=email_visor,
        )
        .order_by("-creado_en")
        .first()
    )
    if caso_existente is not None:
        raise CasuisticaDuplicada(caso_existente)

    asunto = str(asunto or "").strip()
    if not asunto:
        raise CasuisticaError("Escribe un asunto breve para el caso.")
    texto = str(texto or "").strip()
    if not texto:
        raise CasuisticaError("Describe el problema encontrado.")
    archivos = list(archivos)
    for archivo in archivos:
        _validar_archivo(archivo)

    with transaction.atomic():
        caso = Casuistica.objects.create(
            capacitacion=capacitacion,
            email_visor=email_visor,
            nombre_visor=nombre_visor,
            dni_participante=dni,
            nombre_participante=str(nombre_participante or "").strip(),
            asunto=asunto,
            estado=Casuistica.Estado.ABIERTO,
            turno=Casuistica.Turno.ADMIN,
        )
        mensaje = CasuisticaMensaje.objects.create(
            casuistica=caso,
            tipo=CasuisticaMensaje.Tipo.MENSAJE,
            autor_tipo=CasuisticaMensaje.AutorTipo.VISOR,
            autor_nombre=nombre_visor,
            autor_email=email_visor,
            texto=texto,
        )
        _adjuntar_evidencias(mensaje, dni, archivos)
    return caso


def responder_caso(
    caso: Casuistica, *, autor_tipo: str, autor_nombre: str, autor_email: str,
    texto: str, archivos: Iterable[Any] = (),
) -> CasuisticaMensaje:
    """Agrega una respuesta al hilo y voltea el turno.

    Si el caso estaba Cerrado, este mensaje lo reabre (el otro interlocutor
    también puede reabrir, no solo el administrador). Si estaba En plataforma,
    vuelve a Abierto: ya llegó algo nuevo que revisar.
    """
    texto = str(texto or "").strip()
    archivos = list(archivos)
    if not texto and not archivos:
        raise CasuisticaError("Escribe un mensaje o adjunta una evidencia.")
    for archivo in archivos:
        _validar_archivo(archivo)

    with transaction.atomic():
        if caso.estado == Casuistica.Estado.CERRADO or caso.estado == Casuistica.Estado.ANULADO:
            caso.estado = Casuistica.Estado.ABIERTO
            caso.veces_reabierto += 1
            CasuisticaMensaje.objects.create(
                casuistica=caso, tipo=CasuisticaMensaje.Tipo.REAPERTURA,
                autor_tipo=autor_tipo, autor_nombre=autor_nombre, autor_email=autor_email,
                texto="Caso reabierto automáticamente por una nueva observación.",
            )
        elif caso.estado == Casuistica.Estado.EN_PLATAFORMA:
            caso.estado = Casuistica.Estado.ABIERTO
            CasuisticaMensaje.objects.create(
                casuistica=caso, tipo=CasuisticaMensaje.Tipo.CAMBIO_ESTADO,
                autor_tipo=CasuisticaMensaje.AutorTipo.SISTEMA,
                texto="El caso volvió a Abierto por una nueva observación.",
            )

        mensaje = CasuisticaMensaje.objects.create(
            casuistica=caso, tipo=CasuisticaMensaje.Tipo.MENSAJE, autor_tipo=autor_tipo,
            autor_nombre=autor_nombre, autor_email=autor_email, texto=texto,
        )
        _adjuntar_evidencias(mensaje, caso.dni_participante, archivos)

        caso.turno = _voltear_turno(autor_tipo)
        caso.save(update_fields=["estado", "turno", "veces_reabierto", "actualizado_en"])
    return mensaje


def reabrir_caso(caso: Casuistica, *, autor_tipo: str, autor_nombre: str, autor_email: str, motivo: str = "") -> None:
    """Reapertura explícita (botón "Reabrir"), sin necesidad de escribir un mensaje."""
    if caso.estado != Casuistica.Estado.CERRADO:
        raise CasuisticaError("Solo se pueden reabrir casos cerrados.")
    with transaction.atomic():
        caso.estado = Casuistica.Estado.ABIERTO
        caso.veces_reabierto += 1
        caso.turno = _voltear_turno(autor_tipo)
        caso.save(update_fields=["estado", "turno", "veces_reabierto", "actualizado_en"])
        texto = f"Caso reabierto por {autor_nombre or autor_email}."
        if motivo.strip():
            texto += f" Motivo: {motivo.strip()}"
        CasuisticaMensaje.objects.create(
            casuistica=caso, tipo=CasuisticaMensaje.Tipo.REAPERTURA, autor_tipo=autor_tipo,
            autor_nombre=autor_nombre, autor_email=autor_email, texto=texto,
        )


def pasar_a_plataforma(
    caso: Casuistica, *, accion: AccionPlataforma, admin_nombre: str, admin_email: str, detalle: str = "",
) -> None:
    """Solo el administrador: pausa el caso y deja la acción a ejecutar en Plataforma.

    `detalle` es el dato puntual que esa acción necesita (ej. el correo nuevo
    si la acción es "Cambiar correo"): se guarda aparte para poder exportarlo
    como columna propia en el Excel, no solo mezclado en el texto del hilo.
    """
    if caso.estado == Casuistica.Estado.CERRADO or caso.estado == Casuistica.Estado.ANULADO:
        raise CasuisticaError("El caso está cerrado o anulado; reábrelo antes de definir una acción.")
    with transaction.atomic():
        caso.estado = Casuistica.Estado.EN_PLATAFORMA
        caso.turno = ""
        caso.accion_a_realizar = accion
        caso.accion_detalle = detalle.strip()
        caso.accion_definida_en = timezone.now()
        caso.accion_definida_por = admin_email or admin_nombre
        caso.save(update_fields=[
            "estado", "turno", "accion_a_realizar", "accion_detalle",
            "accion_definida_en", "accion_definida_por", "actualizado_en",
        ])
        texto = f"Estado: En plataforma. Acción a realizar: {accion.nombre}."
        if detalle.strip():
            texto += f" Detalle: {detalle.strip()}"
        CasuisticaMensaje.objects.create(
            casuistica=caso, tipo=CasuisticaMensaje.Tipo.CAMBIO_ESTADO,
            autor_tipo=CasuisticaMensaje.AutorTipo.ADMIN, autor_nombre=admin_nombre,
            autor_email=admin_email, texto=texto,
        )


def cerrar_caso(caso: Casuistica, *, admin_nombre: str, admin_email: str, nota: str = "") -> None:
    """Solo el administrador puede cerrar un caso, desde cualquier estado previo."""
    if caso.estado == Casuistica.Estado.CERRADO:
        raise CasuisticaError("El caso ya está cerrado.")
    with transaction.atomic():
        caso.estado = Casuistica.Estado.CERRADO
        caso.turno = ""
        caso.cerrado_en = timezone.now()
        caso.cerrado_por = admin_email or admin_nombre
        caso.save(update_fields=["estado", "turno", "cerrado_en", "cerrado_por", "actualizado_en"])
        texto = f"Caso cerrado por {admin_nombre or admin_email}."
        if nota.strip():
            texto += f" Nota: {nota.strip()}"
        CasuisticaMensaje.objects.create(
            casuistica=caso, tipo=CasuisticaMensaje.Tipo.CAMBIO_ESTADO,
            autor_tipo=CasuisticaMensaje.AutorTipo.ADMIN, autor_nombre=admin_nombre,
            autor_email=admin_email, texto=texto,
        )


def anular_caso(caso: Casuistica, *, admin_nombre: str, admin_email: str, motivo: str = "") -> None:
    """Solo el administrador puede anular un caso (desestimado sin atención), desde cualquier estado previo."""
    if caso.estado == Casuistica.Estado.ANULADO:
        raise CasuisticaError("El caso ya está anulado.")
    with transaction.atomic():
        caso.estado = Casuistica.Estado.ANULADO
        caso.turno = ""
        caso.cerrado_en = timezone.now()
        caso.cerrado_por = admin_email or admin_nombre
        caso.save(update_fields=["estado", "turno", "cerrado_en", "cerrado_por", "actualizado_en"])
        texto = f"Caso anulado por {admin_nombre or admin_email}."
        if motivo.strip():
            texto += f" Motivo: {motivo.strip()}"
        CasuisticaMensaje.objects.create(
            casuistica=caso, tipo=CasuisticaMensaje.Tipo.CAMBIO_ESTADO,
            autor_tipo=CasuisticaMensaje.AutorTipo.ADMIN, autor_nombre=admin_nombre,
            autor_email=admin_email, texto=texto,
        )


def crear_accion_plataforma(nombre: str, *, creado_por: str) -> AccionPlataforma:
    """Agrega una acción nueva al catálogo (popup 'agregar nueva' del desplegable)."""
    nombre = str(nombre or "").strip()
    if not nombre:
        raise CasuisticaError("Escribe el nombre de la acción.")
    accion, creada = AccionPlataforma.objects.get_or_create(
        nombre__iexact=nombre,
        defaults={"nombre": nombre, "creado_por": creado_por, "activo": True},
    )
    if not creada and not accion.activo:
        accion.activo = True
        accion.save(update_fields=["activo"])
    return accion


def exportar_excel_plataforma(casos: Iterable[Casuistica], *, marcar_exportado: bool = True, admin_email: str = "") -> bytes:
    """Genera el Excel de casos 'En plataforma' con la acción a realizar por DNI."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    columnas = [
        "DNI", "Nombre participante", "Curso", "Código curso", "Acción a realizar",
        "Detalle de la acción", "Definida por", "Definida el", "Asunto del caso", "Visor",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Para plataforma"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, nombre in enumerate(columnas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=nombre)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    casos_exportados: list[Casuistica] = []
    for row_idx, caso in enumerate(casos, start=2):
        cap = caso.capacitacion
        codigo = str(cap.cap_codigo or "").strip()
        curso_codigo = f"{codigo}-{cap.cap_id_curso}" if codigo and cap.cap_id_curso else codigo
        fila = [
            str(caso.dni_participante).zfill(8) if caso.dni_participante.isdigit() else caso.dni_participante,
            caso.nombre_participante,
            cap.cap_nombre,
            curso_codigo,
            caso.accion_a_realizar.nombre if caso.accion_a_realizar else "",
            caso.accion_detalle,
            caso.accion_definida_por,
            timezone.localtime(caso.accion_definida_en).strftime("%d/%m/%Y %H:%M") if caso.accion_definida_en else "",
            caso.asunto,
            caso.nombre_visor or caso.email_visor,
        ]
        for col_idx, valor in enumerate(fila, start=1):
            ws.cell(row=row_idx, column=col_idx, value=valor)
        casos_exportados.append(caso)

    anchos = {1: 12, 2: 28, 3: 36, 4: 16, 5: 30, 6: 34, 7: 22, 8: 18, 9: 34, 10: 24}
    for col_idx, ancho in anchos.items():
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = ancho
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    if marcar_exportado and casos_exportados:
        ahora = timezone.now()
        Casuistica.objects.filter(id__in=[c.id for c in casos_exportados]).update(
            exportado_en=ahora, exportado_por=admin_email,
        )

    return buf.getvalue()
