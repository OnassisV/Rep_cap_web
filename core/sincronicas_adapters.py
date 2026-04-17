"""Adaptadores para el modulo de Sincronicas y Evidencias.

Replica la logica de plantillas_sincronicas_streamlit.py en contexto Django:
  - Gestion de archivos de entrada/salida por capacitacion
  - Procesamiento de DataFrames (limpiar, combinar, calcular KPIs)
  - Generacion de Excel con pestanas BBDD TUTOR, REPORTES, ASISTENCIA
"""

from __future__ import annotations

import hashlib
import json
import logging
import os
import re
import unicodedata
from copy import deepcopy
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

from accounts.db import get_connection

logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parent.parent
SINCRONICAS_DIR = BASE_DIR / "Actividades_fuera" / "sincronicas"

NOTA_APROBACION = 13.5


# ---------------------------------------------------------------------------
# Utilidades de archivos
# ---------------------------------------------------------------------------

def _storage_dir(codigo: str, crear: bool = True) -> Path:
    safe = str(codigo).replace("/", "_").replace("\\", "_").replace(":", "_")[:50]
    path = SINCRONICAS_DIR / safe
    if crear:
        path.mkdir(parents=True, exist_ok=True)
    return path


def _listar_archivos(storage: Path) -> tuple[list[str], list[str]]:
    if not storage.is_dir():
        return [], []
    entrada, salida = [], []
    for f in sorted(storage.iterdir()):
        if f.suffix.lower() in (".xlsx", ".xls") and f.is_file():
            if f.name.startswith("entrada_"):
                entrada.append(f.name)
            elif f.name.startswith("salida_"):
                salida.append(f.name)
    return entrada, salida


def _sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _calcular_hash_insumos(storage: Path, entradas: list[str], salidas: list[str]) -> str:
    hashes = []
    for nombre in sorted(entradas):
        hashes.append(_sha256_file(storage / nombre))
    for nombre in sorted(salidas):
        hashes.append(_sha256_file(storage / nombre))
    return hashlib.sha256("|".join(hashes).encode()).hexdigest()


# ---------------------------------------------------------------------------
# Manifest de procesamiento (idempotencia)
# ---------------------------------------------------------------------------

def _manifest_path(storage: Path) -> Path:
    return storage / "manifest_procesamiento.json"


def _cargar_manifest(storage: Path) -> dict:
    mp = _manifest_path(storage)
    if mp.exists():
        try:
            return json.loads(mp.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {"version": 1, "runs": {}}


def _guardar_manifest(storage: Path, manifest: dict) -> None:
    mp = _manifest_path(storage)
    tmp = mp.with_suffix(".tmp")
    tmp.write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")
    os.replace(str(tmp), str(mp))


def _obtener_run_completado(manifest: dict, input_hash: str) -> dict | None:
    runs = manifest.get("runs", {})
    run = runs.get(input_hash)
    if run and run.get("status") == "completed":
        output_path = run.get("output_path", "")
        if output_path and os.path.exists(output_path):
            return run
    return None


def _bytes_desde_blob(blob: Any) -> bytes:
    if blob is None:
        return b""
    if isinstance(blob, memoryview):
        return blob.tobytes()
    if isinstance(blob, (bytes, bytearray)):
        return bytes(blob)
    try:
        return bytes(blob)
    except Exception:
        return b""


def _obtener_run_db(codigo: str, input_hash: str | None = None) -> dict | None:
    """Lee el último procesamiento persistido en MySQL."""
    try:
        from core.models import CapSincronicaProcesamiento

        qs = CapSincronicaProcesamiento.objects.filter(
            codigo=str(codigo).strip(),
            status="completed",
        )
        if input_hash:
            qs = qs.filter(input_hash=input_hash)
        run = qs.order_by("-procesado_en").first()
        if not run:
            return None
        return {
            "status": run.status,
            "timestamp": run.procesado_en.isoformat() if run.procesado_en else "",
            "input_hash": run.input_hash,
            "output_path": "",
            "output_file": run.output_file,
            "output_content": _bytes_desde_blob(run.output_blob),
            "stats": json.loads(run.stats_json or "{}"),
        }
    except Exception:
        logger.exception("Error leyendo procesamiento persistido de %s", codigo)
        return None


def _guardar_run_db(
    codigo: str,
    input_hash: str,
    output_file: str,
    output_bytes: bytes,
    stats: dict[str, Any],
) -> None:
    """Guarda el resultado procesado en MySQL para reutilización y descarga."""
    try:
        from core.models import Capacitacion, CapSincronicaProcesamiento

        cap = Capacitacion.objects.filter(cap_codigo=str(codigo).strip()).order_by("-cap_anio", "-actualizado_en").first()
        CapSincronicaProcesamiento.objects.update_or_create(
            codigo=str(codigo).strip(),
            input_hash=input_hash,
            defaults={
                "capacitacion": cap,
                "status": "completed",
                "output_file": str(output_file or "").strip(),
                "output_blob": output_bytes,
                "stats_json": json.dumps(stats or {}, ensure_ascii=False),
            },
        )
    except Exception:
        logger.exception("Error guardando procesamiento persistido de %s", codigo)


# ---------------------------------------------------------------------------
# Guardar archivos subidos (evitar duplicados por contenido)
# ---------------------------------------------------------------------------

def _indexar_hashes(storage: Path, prefijo: str) -> dict[str, str]:
    idx: dict[str, str] = {}
    for f in storage.iterdir():
        if f.is_file() and f.name.startswith(prefijo) and f.suffix.lower() in (".xlsx", ".xls"):
            idx[_sha256_file(f)] = f.name
    return idx


def guardar_archivos_subidos(
    archivos: list[Any],
    tipo: str,
    storage: Path,
    timestamp: str,
) -> list[str]:
    """Guarda archivos subidos evitando duplicados por SHA-256. Retorna nombres guardados."""
    prefijo = f"{tipo}_"
    existentes = _indexar_hashes(storage, prefijo)
    guardados = []
    for i, archivo in enumerate(archivos, 1):
        contenido = archivo.read()
        archivo.seek(0)
        h = _sha256_bytes(contenido)
        if h in existentes:
            continue
        nombre = f"{prefijo}{timestamp}_{i}_{archivo.name}"
        dest = storage / nombre
        dest.write_bytes(contenido)
        existentes[h] = nombre
        guardados.append(nombre)
    return guardados


def eliminar_archivo(storage: Path, nombre: str) -> bool:
    """Elimina un archivo de insumo del storage."""
    target = storage / nombre
    if target.is_file() and target.parent == storage:
        target.unlink()
        return True
    return False


# ---------------------------------------------------------------------------
# Listar capacitaciones sincronicas desde oferta
# ---------------------------------------------------------------------------

def obtener_capacitaciones_sincronicas(
    role_effective: str,
    display_name: str,
    username: str,
) -> list[dict[str, Any]]:
    """Retorna filas de oferta_formativa_difoca donde tipo_proceso contiene 'sincron'."""
    sql = """
        SELECT codigo, anio, condicion,
               tipo_proceso_formativo, denominacion_proceso_formativo,
               especialista_cargo
        FROM oferta_formativa_difoca
        ORDER BY anio DESC, condicion, codigo
    """
    try:
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(sql)
                rows = cur.fetchall()
    except Exception:
        logger.exception("Error leyendo oferta para sincronicas")
        return []

    result = []
    norm_role = str(role_effective or "").strip().lower()
    norm_display = str(display_name or "").strip().lower()
    norm_user = str(username or "").strip().lower()

    for row in rows:
        tipo = str(row.get("tipo_proceso_formativo", "")).strip()
        if "sincr" not in tipo.lower():
            continue
        cond = str(row.get("condicion", "")).strip().lower()
        if cond not in ("cerrado", "implementacion", "en implementacion"):
            continue
        if norm_role == "usuario estandar":
            esp = str(row.get("especialista_cargo", "")).strip().lower()
            if esp != norm_display and esp != norm_user:
                continue
        result.append({
            "codigo": str(row.get("codigo", "")).strip(),
            "anio": str(row.get("anio", "")).strip(),
            "condicion": str(row.get("condicion", "")).strip(),
            "tipo_proceso_formativo": tipo,
            "denominacion_proceso_formativo": str(row.get("denominacion_proceso_formativo", "")).strip(),
            "especialista_cargo": str(row.get("especialista_cargo", "")).strip(),
        })
    return result


def aplicar_filtro_anio_sync(
    filas: list[dict[str, Any]], anio_param: str,
) -> tuple[list[str], str, list[dict[str, Any]]]:
    """Aplica selector de anio a la lista de sincronicas."""
    anios = sorted({str(f.get("anio", "")).strip() for f in filas if str(f.get("anio", "")).strip()}, reverse=True)
    if not anios:
        return [], "", []
    anio_sel = anio_param if anio_param in anios else anios[0]
    filtradas = [f for f in filas if str(f.get("anio", "")).strip() == anio_sel]
    return anios, anio_sel, filtradas


def obtener_info_archivos(codigo: str) -> dict[str, Any]:
    """Retorna info de archivos guardados para un codigo."""
    storage = _storage_dir(codigo, crear=False)
    entradas, salidas = _listar_archivos(storage)
    manifest = _cargar_manifest(storage) if storage.is_dir() else {}
    procesado = False
    output_file = ""
    run_db = _obtener_run_db(codigo)
    if entradas and salidas:
        try:
            ih = _calcular_hash_insumos(storage, entradas, salidas)
            run = _obtener_run_completado(manifest, ih) or _obtener_run_db(codigo, ih)
            if run:
                procesado = True
                output_file = run.get("output_file", "")
        except Exception:
            logger.exception("Error calculando estado de procesamiento para %s", codigo)
    elif run_db:
        procesado = True
        output_file = str(run_db.get("output_file", ""))
    return {
        "entradas": entradas,
        "salidas": salidas,
        "total_entradas": len(entradas),
        "total_salidas": len(salidas),
        "procesado": procesado,
        "output_file": output_file,
    }


# ---------------------------------------------------------------------------
# Procesamiento de archivos individuales
# ---------------------------------------------------------------------------

def _limpiar_respuestas_preguntas(df: pd.DataFrame) -> pd.DataFrame:
    """Elimina prefijos tipo 'Pregunta: ' y puntos finales de columnas de texto."""
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].astype(str).str.replace(
                r"^(?:Pregunta|Respuesta|Opci[oó]n)\s*:\s*", "", regex=True
            )
            df[col] = df[col].str.rstrip(".")
    return df


def procesar_archivo_individual(df: pd.DataFrame, nombre_archivo: str | None = None) -> pd.DataFrame:
    """Aplica todas las transformaciones a un DataFrame individual de entrada o salida."""
    df_p = deepcopy(df)

    # 0. Limpiar nombres de columnas: quitar \n, \xa0, espacios extra
    df_p.columns = [str(c).replace('\n', '').replace('\xa0', ' ').strip() for c in df_p.columns]

    # 1. Columnas a eliminar
    cols_eliminar = [c for c in df_p.columns if str(c).startswith("Puntos:") or str(c).startswith("Comentarios:")]

    # 2. Columnas IGED para combinar
    cols_iged = [c for c in df_p.columns if str(c).startswith("Indique la IGED")]

    # 3. Limpiar respuestas
    df_p = _limpiar_respuestas_preguntas(df_p)

    # 4. Combinar IGED
    if cols_iged:
        df_p["IGED"] = ""
        for idx, row in df_p.iterrows():
            vals = []
            for ci in cols_iged:
                v = str(row[ci]) if pd.notna(row[ci]) else ""
                if v and v != "nan" and v.strip():
                    vals.append(v.strip())
            df_p.at[idx, "IGED"] = "; ".join(set(vals)) if vals else ""
        cols_eliminar.extend(cols_iged)

    # 5. Eliminar columnas marcadas
    if cols_eliminar:
        df_p = df_p.drop(columns=cols_eliminar, errors="ignore")

    # 6. Eliminar columnas vacias
    cols_vacias = []
    for col in df_p.columns:
        s = df_p[col].astype(str).str.strip()
        if s.isin(["", "nan", "None"]).all() or df_p[col].isna().all():
            cols_vacias.append(col)
    if cols_vacias:
        df_p = df_p.drop(columns=cols_vacias, errors="ignore")

    # 7. Renombrar
    renames = {}
    if "Indique la Región a la que pertenece" in df_p.columns:
        renames["Indique la Región a la que pertenece"] = "REGION"
    if "Total de puntos" in df_p.columns:
        renames["Total de puntos"] = "NOTA"
    if renames:
        df_p = df_p.rename(columns=renames)

    # 8. Reordenar IGED despues de REGION
    if "REGION" in df_p.columns and "IGED" in df_p.columns:
        cols = list(df_p.columns)
        cols.remove("IGED")
        pos = cols.index("REGION")
        cols.insert(pos + 1, "IGED")
        df_p = df_p[cols]

    # 9. Eliminar columna Nombre
    df_p = df_p.drop(columns=["Nombre"], errors="ignore")

    # 10. Formatear DNI
    col_dni = None
    if "Documento de Identidad" in df_p.columns:
        col_dni = "Documento de Identidad"
    elif "DNI" in df_p.columns:
        col_dni = "DNI"
    if col_dni:
        if col_dni != "DNI":
            df_p = df_p.rename(columns={col_dni: "DNI"})
        df_p["DNI"] = df_p["DNI"].astype(str).str.replace("O", "0").str.replace("o", "0")

        def fmt_dni(d):
            if pd.isna(d):
                return ""
            s = str(d).strip()
            if "." in s:
                s = s.split(".")[0]
            return s.zfill(8) if s.isdigit() else s

        df_p["DNI"] = df_p["DNI"].apply(fmt_dni)

    # 10.1 Depurar DNIs invalidos
    if "DNI" in df_p.columns:
        dn = df_p["DNI"].fillna("").astype(str).str.strip()
        con_dato = dn != ""
        no_num = con_dato & ~dn.str.fullmatch(r"\d+")
        repetido = con_dato & dn.str.fullmatch(r"(\d)\1{7}")
        descartar = no_num | repetido
        if descartar.any():
            df_p = df_p.loc[~descartar].copy()

    # 11. Eliminar duplicados por DNI
    if "DNI" in df_p.columns and "Hora de inicio" in df_p.columns:
        try:
            df_p["Hora de inicio"] = pd.to_datetime(df_p["Hora de inicio"], errors="coerce")
            con_dni = df_p[(df_p["DNI"] != "") & df_p["DNI"].notna()]
            sin_dni = df_p[~((df_p["DNI"] != "") & df_p["DNI"].notna())]
            con_dni = con_dni.sort_values("Hora de inicio").drop_duplicates(subset=["DNI"], keep="first")
            df_p = pd.concat([con_dni, sin_dni], ignore_index=True)
            df_p = df_p.sort_values("Hora de inicio").reset_index(drop=True)
        except Exception:
            pass

    if nombre_archivo is not None:
        df_p["ARCHIVO_ORIGEN"] = nombre_archivo

    return df_p


# ---------------------------------------------------------------------------
# Procesamiento completo
# ---------------------------------------------------------------------------

def _normalizar_texto_ascii_upper(s: pd.Series) -> pd.Series:
    """Convierte a MAYUSCULAS quitando tildes vocalicas (conserva Ñ)."""
    result = s.fillna("").astype(str).str.strip().str.upper()
    for orig, repl in [("Á", "A"), ("É", "E"), ("Í", "I"), ("Ó", "O"), ("Ú", "U"), ("Ü", "U")]:
        result = result.str.replace(orig, repl, regex=False)
    return result


def _obtener_catalogo_iged() -> pd.DataFrame:
    """Lee catalogo iged_s3 desde MySQL y normaliza para match."""
    try:
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute('SELECT region, `NOMBRE IGED` AS nombre_iged, `tipo IGED` AS tipo_iged, CODIGO_R AS codigo_iged FROM iged_s3')
                rows = cur.fetchall()
        df = pd.DataFrame(list(rows or []))
        if not df.empty:
            # Normalizar region y nombre_iged igual que df_usuarios para match
            df["region"] = _normalizar_texto_ascii_upper(df["region"])
            df["nombre_iged"] = _normalizar_texto_ascii_upper(df["nombre_iged"])
        return df
    except Exception:
        logger.exception("Error leyendo iged_s3")
        return pd.DataFrame(columns=["region", "nombre_iged", "tipo_iged", "codigo_iged"])


def _obtener_nombre_capacitacion(codigo: str) -> dict[str, str]:
    """Obtiene anio/tipo/denominacion de cap_capacitaciones (Django ORM)."""
    try:
        from core.models import Capacitacion
        cap = Capacitacion.objects.filter(cap_codigo=str(codigo).strip()).first()
        if cap:
            return {
                "anio": str(cap.cap_anio or "").strip(),
                "tipo": str(cap.cap_tipo or "").strip(),
                "denominacion": str(cap.cap_nombre or "").strip(),
            }
    except Exception:
        logger.exception("Error leyendo nombre capacitacion")
    return {"anio": "", "tipo": "", "denominacion": ""}


def procesar_sincronicas(codigo: str) -> dict[str, Any]:
    """Procesa archivos de entrada y salida para una capacitacion sincronica.

    Retorna dict con: ok, error, df_tutor, df_usuarios, df_entrada, df_salida,
    output_path, output_file, stats.
    """
    logger.info("procesar_sincronicas: inicio para codigo=%s", codigo)
    storage = _storage_dir(codigo)
    entradas, salidas = _listar_archivos(storage)
    logger.info("procesar_sincronicas: entradas=%d, salidas=%d", len(entradas), len(salidas))

    if not entradas:
        return {"ok": False, "error": "No hay archivos de entrada."}
    if not salidas:
        return {"ok": False, "error": "No hay archivos de salida."}

    # Idempotencia
    input_hash = _calcular_hash_insumos(storage, entradas, salidas)
    manifest = _cargar_manifest(storage)
    run_prev = _obtener_run_db(codigo, input_hash) or _obtener_run_completado(manifest, input_hash)
    if run_prev:
        logger.info("procesar_sincronicas: reutilizando run previo")
        return {
            "ok": True,
            "reutilizado": True,
            "output_path": run_prev.get("output_path", ""),
            "output_file": run_prev.get("output_file", ""),
            "output_content": run_prev.get("output_content", b""),
            "stats": run_prev.get("stats", {}),
        }

    # Procesar archivos
    logger.info("procesar_sincronicas: leyendo archivos Excel...")
    dfs_entrada, dfs_salida = [], []
    for nombre in entradas:
        df = pd.read_excel(storage / nombre)
        dfs_entrada.append(procesar_archivo_individual(df, nombre))
    for nombre in salidas:
        df = pd.read_excel(storage / nombre)
        dfs_salida.append(procesar_archivo_individual(df, nombre))

    df_entrada = pd.concat(dfs_entrada, ignore_index=True)
    df_salida = pd.concat(dfs_salida, ignore_index=True)
    logger.info("procesar_sincronicas: df_entrada=%d filas, df_salida=%d filas", len(df_entrada), len(df_salida))

    # Crear df_usuarios
    dnis_e = set(df_entrada["DNI"].dropna().astype(str)) if "DNI" in df_entrada.columns else set()
    dnis_s = set(df_salida["DNI"].dropna().astype(str)) if "DNI" in df_salida.columns else set()
    todos_dnis = dnis_e | dnis_s

    df_usuarios = pd.DataFrame({"DNI": list(todos_dnis)})

    # Mapear notas
    if "NOTA" in df_entrada.columns:
        df_usuarios["cuestionario entrada"] = df_usuarios["DNI"].map(df_entrada.set_index("DNI")["NOTA"].to_dict())
    if "NOTA" in df_salida.columns:
        df_usuarios["cuestionario salida"] = df_usuarios["DNI"].map(df_salida.set_index("DNI")["NOTA"].to_dict())

    # Mapeo inteligente: prioriza salida sobre entrada
    columnas_mapear = {
        "Nombres": "nombres",
        "Apellidos": "apellidos",
        "Correo electrónico": "email",
        "Número de celular": "celular",
        "EMAIL": "email",
        "CELULAR": "celular",
        "Entidad Pública a la que pertenece:": "Entidad Pública a la que pertenece:",
        "Entidad Pública": "Entidad Pública",
        "REGION": "REGION",
        "IGED": "IGED",
        "Niveles de puesto": "Niveles de puesto",
        "Nombre del Puesto (Escriba puntualmente el cargo o puesto al que pertenece en la entidad)": "Nombre del Puesto",
    }

    for col_origen, col_destino in columnas_mapear.items():
        df_usuarios[col_destino] = ""
        if col_origen in df_entrada.columns:
            df_usuarios[col_destino] = df_usuarios["DNI"].map(df_entrada.set_index("DNI")[col_origen].to_dict())
        if col_origen in df_salida.columns:
            mapa_s = df_salida.set_index("DNI")[col_origen].to_dict()
            for dni in df_usuarios["DNI"]:
                if dni in mapa_s:
                    vs = mapa_s[dni]
                    if pd.notna(vs) and str(vs).strip() not in ("", "nan"):
                        df_usuarios.loc[df_usuarios["DNI"] == dni, col_destino] = vs
        df_usuarios[col_destino] = df_usuarios[col_destino].fillna("")

    # Normalizar nombres
    def _norm_nombre(texto):
        if pd.isnull(texto) or texto == "":
            return ""
        t = str(texto).upper().strip()
        t = re.sub(r"[^A-ZÁÉÍÓÚÜÑ\s\-'.]", "", t)
        return re.sub(r"\s+", " ", t).strip()

    for col in ("nombres", "apellidos"):
        if col in df_usuarios.columns:
            df_usuarios[col] = df_usuarios[col].apply(_norm_nombre)

    # Agregar codigo
    df_usuarios["codigo"] = codigo

    # Renombrar
    df_usuarios = df_usuarios.rename(columns={
        "DNI": "dni",
        "REGION": "region",
        "IGED": "nombre_iged",
        "Niveles de puesto": "nivel_puesto",
        "Nombre del Puesto": "nombre_puesto",
        "celular": "telefono_celular",
    })

    # Normalizar region/iged para match con BD
    if "region" in df_usuarios.columns:
        df_usuarios["region"] = _normalizar_texto_ascii_upper(df_usuarios["region"])
        df_usuarios["region"] = df_usuarios["region"].str.replace(r"^LIMA PROVINCIA$", "LIMA PROVINCIAS", regex=True)
    if "nombre_iged" in df_usuarios.columns:
        df_usuarios["nombre_iged"] = _normalizar_texto_ascii_upper(df_usuarios["nombre_iged"])
        df_usuarios["nombre_iged"] = df_usuarios["nombre_iged"].str.replace(r"^DRE LIMA PROVINCIA$", "DRE LIMA PROVINCIAS", regex=True)
        df_usuarios["nombre_iged"] = df_usuarios["nombre_iged"].replace(["", "NAN"], pd.NA)

    # Merge con iged_s3
    logger.info("procesar_sincronicas: merge con iged_s3...")
    iged_cat = _obtener_catalogo_iged()
    if not iged_cat.empty:
        df_usuarios = df_usuarios.merge(iged_cat, on=["region", "nombre_iged"], how="left")

        # Fallback: si no matcheó por region+nombre_iged, corregir region con la BD
        sin_codigo = df_usuarios["codigo_iged"].isna() & df_usuarios["nombre_iged"].notna()
        if sin_codigo.any():
            cat_by_nombre = iged_cat.drop_duplicates(subset=["nombre_iged"])
            for idx in df_usuarios.index[sin_codigo]:
                nombre = df_usuarios.at[idx, "nombre_iged"]
                match = cat_by_nombre[cat_by_nombre["nombre_iged"] == nombre]
                if len(match) == 1:
                    df_usuarios.at[idx, "region"] = match.iloc[0]["region"]
                    df_usuarios.at[idx, "tipo_iged"] = match.iloc[0]["tipo_iged"]
                    df_usuarios.at[idx, "codigo_iged"] = match.iloc[0]["codigo_iged"]

        if "codigo_iged" in df_usuarios.columns:
            def _fmt_cod(x):
                if pd.isnull(x):
                    return ""
                s = str(x)
                if "." in s:
                    s = s.split(".")[0]
                return s.zfill(6) if s.isdigit() else ""
            df_usuarios["codigo_iged"] = df_usuarios["codigo_iged"].apply(_fmt_cod)
    else:
        for c in ("tipo_iged", "codigo_iged"):
            if c not in df_usuarios.columns:
                df_usuarios[c] = ""

    # Crear df_tutor: filtrar DRE/GRE y UGEL
    mask_tutor = (
        df_usuarios.get("Entidad Pública a la que pertenece:", pd.Series(dtype=str))
        .fillna("")
        .astype(str)
        .str.strip()
        == "DRE/GRE y UGEL"
    )
    df_tutor = df_usuarios[mask_tutor].copy()
    df_tutor = df_tutor.drop(columns=["Entidad Pública a la que pertenece:", "Entidad Pública"], errors="ignore")

    # Tipo documento
    df_tutor["tipo_documento"] = df_tutor["dni"].apply(
        lambda d: "DNI" if str(d).strip().isdigit() and len(str(d).strip()) == 8 else "OTROS"
    )

    # Estado, compromiso, avance
    df_tutor["estado"] = df_tutor.apply(
        lambda r: 2 if pd.notna(r.get("cuestionario salida")) else 1, axis=1,
    )
    df_tutor["compromiso"] = df_tutor["estado"].apply(lambda x: 20 if x == 2 else 0)
    df_tutor["%_avance_certificacion"] = df_tutor["estado"].apply(lambda x: 1 if x == 2 else 0.5)

    # Situacion
    df_tutor["situacion_del_participante"] = df_tutor["cuestionario salida"].apply(
        lambda x: "Aprueba" if pd.notna(x) and x >= NOTA_APROBACION else "No aprueba"
    )
    df_tutor["promedio_final_general"] = df_tutor["cuestionario salida"]
    df_tutor["promedio_final_condicion"] = df_tutor["promedio_final_general"].apply(
        lambda x: "Aprobado" if pd.notna(x) and x >= NOTA_APROBACION else "Desaprobado"
    )

    # Resultados
    df_tutor["Aprobados/Certificados"] = df_tutor["cuestionario salida"].apply(
        lambda x: 1 if pd.notna(x) and x >= NOTA_APROBACION else ""
    )
    df_tutor["Desaprobado/Permanente"] = df_tutor["cuestionario salida"].apply(
        lambda x: 1 if pd.notna(x) and x < NOTA_APROBACION else ""
    )
    df_tutor["Desaprobado/Abandono"] = df_tutor["cuestionario salida"].apply(
        lambda x: "" if pd.notna(x) else 1
    )

    # Progreso
    df_tutor["ev_progreso_aprendizaje"] = df_tutor.apply(
        lambda r: 1 if pd.notna(r.get("cuestionario entrada")) and pd.notna(r.get("cuestionario salida")) else "",
        axis=1,
    )
    ce = pd.to_numeric(df_tutor.get("cuestionario entrada"), errors="coerce")
    cs = pd.to_numeric(df_tutor.get("cuestionario salida"), errors="coerce")
    diff = cs - ce
    df_tutor["mantuvo_o_progres\u00f3"] = np.where(diff >= 0, "1", "")
    df_tutor["progres\u00f3"] = np.where(diff > 0, "1", "")

    # Niveles
    def _nivel(s: pd.Series) -> pd.Series:
        n = pd.to_numeric(s, errors="coerce")
        conds = [n < 11, (n >= 11) & (n <= 13), (n >= 14) & (n <= 17), n > 17]
        vals = ["En inicio", "En proceso", "Logrado", "Destacado"]
        return pd.Series(np.select(conds, vals, default=""), index=s.index)

    df_tutor["Nivel C. Entrada"] = _nivel(df_tutor.get("cuestionario entrada", pd.Series(dtype=float)))
    df_tutor["Nivel C. Salida"] = _nivel(df_tutor.get("cuestionario salida", pd.Series(dtype=float)))

    # Limpiar NaN
    df_tutor["cuestionario entrada"] = df_tutor["cuestionario entrada"].fillna("")
    df_tutor["cuestionario salida"] = df_tutor["cuestionario salida"].fillna("")

    # Numerar
    df_tutor["N°"] = range(1, len(df_tutor) + 1)

    # Ordenar columnas
    col_order = [
        "N°", "codigo", "tipo_documento", "dni", "nombres", "apellidos",
        "email", "telefono_celular", "region", "tipo_iged", "codigo_iged",
        "nombre_iged", "nivel_puesto", "nombre_puesto",
        "estado", "compromiso", "%_avance_certificacion",
        "promedio_final_general", "promedio_final_condicion",
        "situacion_del_participante", "Aprobados/Certificados",
        "Desaprobado/Permanente", "Desaprobado/Abandono",
        "cuestionario entrada", "cuestionario salida",
        "ev_progreso_aprendizaje", "mantuvo_o_progres\u00f3", "progres\u00f3",
        "Nivel C. Entrada", "Nivel C. Salida",
    ]
    existing = [c for c in col_order if c in df_tutor.columns]
    df_tutor = df_tutor[existing]

    # Generar Excel
    logger.info("procesar_sincronicas: generando Excel... df_tutor=%d filas", len(df_tutor))
    output_bytes = _generar_excel(df_tutor, df_usuarios, codigo)
    cap_info = _obtener_nombre_capacitacion(codigo)
    anio_cap = cap_info.get("anio") or str(datetime.now().year)

    if cap_info.get("tipo") and cap_info.get("denominacion"):
        nombre_safe = re.sub(r'[\\/:*?"<>|]+', " ", f"{cap_info['tipo']} {cap_info['denominacion']}")
        output_file = f"{codigo} - {nombre_safe.strip()}.xlsx"
    else:
        output_file = f"PLANTILLA_{codigo}.xlsx"

    # Guardar resultado dentro del directorio de la capacitacion.
    output_path = storage / output_file
    output_path.write_bytes(output_bytes)
    logger.info("procesar_sincronicas: archivo guardado en %s (%d bytes)", output_path, len(output_bytes))

    # Insertar en bbdd_2026 para que la VIEW bbdd_difoca refleje los datos
    _insertar_en_bbdd_2026(df_tutor, codigo)

    # Actualizar manifest
    stats = {
        "filas_entrada": len(df_entrada),
        "filas_salida": len(df_salida),
        "usuarios_unicos": len(df_usuarios),
        "tutores": len(df_tutor),
    }
    manifest["capacitacion_id"] = codigo
    manifest.setdefault("runs", {})[input_hash] = {
        "status": "completed",
        "timestamp": datetime.now().isoformat(),
        "input_hash": input_hash,
        "output_path": str(output_path),
        "output_file": output_file,
        "stats": stats,
        "rows_df_usuarios": len(df_usuarios),
        "rows_df_tutor": len(df_tutor),
    }
    _guardar_manifest(storage, manifest)
    _guardar_run_db(codigo, input_hash, output_file, output_bytes, stats)

    return {
        "ok": True,
        "reutilizado": False,
        "output_path": str(output_path),
        "output_file": output_file,
        "stats": stats,
    }


# ---------------------------------------------------------------------------
# Insercion en bbdd_2026 (alimenta la VIEW bbdd_difoca)
# ---------------------------------------------------------------------------

# Mapeo de columnas BBDD TUTOR -> columnas bbdd_2026
_TUTOR_TO_BBDD = {
    "N\u00b0": "numero",
    "codigo": "codigo",
    "tipo_documento": "tipo_documento",
    "dni": "dni",
    "nombres": "nombres",
    "apellidos": "apellidos",
    "email": "email",
    "telefono_celular": "telefono_celular",
    "region": "region",
    "tipo_iged": "tipo_iged",
    "codigo_iged": "codigo_iged",
    "nombre_iged": "nombre_iged",
    "nivel_puesto": "nivel_puesto",
    "nombre_puesto": "nombre_puesto",
    "estado": "estado",
    "compromiso": "compromiso",
    "%_avance_certificacion": "avance_curso_certificacion",
    "promedio_final_general": "promedio_final_general",
    "promedio_final_condicion": "promedio_final_condicion",
    "situacion_del_participante": "situacion_participante",
    "Aprobados/Certificados": "aprobados_certificados",
    "Desaprobado/Permanente": "desaprobado_permanente",
    "Desaprobado/Abandono": "desaprobado_abandono",
    "cuestionario entrada": "cuestionario_entrada",
    "cuestionario salida": "cuestionario_salida",
    "ev_progreso_aprendizaje": "ev_progreso_aprendizaje",
    "mantuvo_o_progres\u00f3": "mantuvo_o_progreso",
    "progres\u00f3": "progreso",
    "Nivel C. Entrada": "nivel_c_entrada",
    "Nivel C. Salida": "nivel_c_salida",
}


def _insertar_en_bbdd_2026(df_tutor: pd.DataFrame, codigo: str) -> None:
    """Inserta df_tutor en bbdd_2026. Si ya existen registros del mismo codigo, los reemplaza."""
    try:
        df = df_tutor.copy()
        # Renombrar columnas al esquema de bbdd_2026
        df = df.rename(columns=_TUTOR_TO_BBDD)
        # Solo conservar columnas que existen en bbdd_2026
        cols_bbdd = list(_TUTOR_TO_BBDD.values())
        for c in cols_bbdd:
            if c not in df.columns:
                df[c] = None
        df = df[cols_bbdd]
        # Reemplazar strings vacios y "nan" por None para MySQL
        df = df.where(df.notnull(), None)
        df = df.replace({"": None, "nan": None, "NaN": None})

        with get_connection() as conn:
            with conn.cursor() as cur:
                # Borrar registros previos del mismo codigo
                cur.execute("DELETE FROM bbdd_2026 WHERE codigo = %s", (codigo,))
                deleted = cur.rowcount
                if deleted:
                    logger.info("_insertar_en_bbdd_2026: eliminados %d registros previos de %s", deleted, codigo)

                # Insertar nuevos registros
                placeholders = ", ".join(["%s"] * len(cols_bbdd))
                col_names = ", ".join(f"`{c}`" for c in cols_bbdd)
                sql = f"INSERT INTO bbdd_2026 ({col_names}) VALUES ({placeholders})"

                def _safe(v):
                    """Convierte NaN/NA/pd.NA a None para MySQL."""
                    if v is None:
                        return None
                    try:
                        if pd.isna(v):
                            return None
                    except (ValueError, TypeError):
                        pass
                    if isinstance(v, float) and (v != v):  # NaN check
                        return None
                    s = str(v)
                    if s in ("", "nan", "NaN", "<NA>"):
                        return None
                    return v

                rows = [tuple(_safe(row[c]) for c in cols_bbdd) for _, row in df.iterrows()]
                cur.executemany(sql, rows)
                logger.info("_insertar_en_bbdd_2026: insertados %d registros para %s", len(rows), codigo)
    except Exception:
        logger.exception("Error insertando en bbdd_2026 para codigo=%s", codigo)


# ---------------------------------------------------------------------------
# Generacion de Excel con formato
# ---------------------------------------------------------------------------

def _crear_reportes_dinamicos(df_tutor: pd.DataFrame) -> list[tuple[str, pd.DataFrame]]:
    """Genera tablas pivot de reportes."""
    reportes = []

    if df_tutor.empty or "region" not in df_tutor.columns:
        return reportes

    # 1. Cumplimiento
    estado_num = pd.to_numeric(df_tutor.get("estado"), errors="coerce")
    ce_num = pd.to_numeric(df_tutor.get("cuestionario entrada"), errors="coerce")
    cs_num = pd.to_numeric(df_tutor.get("cuestionario salida"), errors="coerce")
    cert = pd.to_numeric(df_tutor.get("Aprobados/Certificados"), errors="coerce")

    cumpl = pd.DataFrame({
        "region": df_tutor["region"],
        "Postulantes": 1,
        "Matriculados": (estado_num >= 1).astype(int),
        "Participantes": (estado_num == 2).astype(int),
        "Cuest. Entrada": ce_num.notna().astype(int),
        "Cuest. Salida": cs_num.notna().astype(int),
        "Certificados": cert.fillna(0).astype(int),
    })
    cumpl_pivot = cumpl.groupby("region").sum().reset_index()
    if not cumpl_pivot.empty and cumpl_pivot["Matriculados"].sum() > 0:
        cumpl_pivot["% Cuest. Entrada"] = cumpl_pivot["Cuest. Entrada"] / cumpl_pivot["Matriculados"]
        cumpl_pivot["% Certificados"] = cumpl_pivot["Certificados"] / cumpl_pivot["Matriculados"]
    reportes.append(("Cumplimiento", cumpl_pivot))

    # 2. Cobertura IGED
    tipo_iged = df_tutor.get("tipo_iged", pd.Series(dtype=str)).fillna("").astype(str).str.strip().str.upper()
    nombre_iged = df_tutor.get("nombre_iged", pd.Series(dtype=str)).fillna("").astype(str).str.strip()
    cob = pd.DataFrame({"region": df_tutor["region"], "tipo_iged": tipo_iged, "nombre_iged": nombre_iged})
    dre_cob = cob[cob["tipo_iged"].str.contains("DRE", na=False)].groupby("region")["nombre_iged"].nunique().reset_index()
    dre_cob.columns = ["region", "DRE/GRE"]
    ugel_cob = cob[cob["tipo_iged"] == "UGEL"].groupby("region")["nombre_iged"].nunique().reset_index()
    ugel_cob.columns = ["region", "UGEL"]
    cob_pivot = dre_cob.merge(ugel_cob, on="region", how="outer").fillna(0)
    reportes.append(("Cobertura IGED", cob_pivot))

    # 3. Nivel de puesto
    if "nivel_puesto" in df_tutor.columns:
        np_df = df_tutor[["region", "nivel_puesto"]].copy()
        np_df["nivel_puesto"] = np_df["nivel_puesto"].fillna("Sin dato").astype(str).str.strip()
        np_pivot = pd.crosstab(np_df["region"], np_df["nivel_puesto"])
        np_pivot = np_pivot.reset_index()
        reportes.append(("Nivel Puesto", np_pivot))

    # 4-5. Niveles entrada/salida
    for col, titulo in [("Nivel C. Entrada", "Nivel C. Entrada"), ("Nivel C. Salida", "Nivel C. Salida")]:
        if col in df_tutor.columns:
            niv = df_tutor[["region", col]].copy()
            niv[col] = niv[col].fillna("Sin dato").astype(str).str.strip()
            niv_pivot = pd.crosstab(niv["region"], niv[col])
            niv_pivot = niv_pivot.reset_index()
            reportes.append((titulo, niv_pivot))

    return reportes


def _sanitizar_df_para_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Reemplaza todo tipo de NA/NaN/pd.NA por cadena vacia para openpyxl."""
    out = df.copy()
    for col in out.columns:
        out[col] = out[col].fillna("").infer_objects()
        out[col] = out[col].apply(lambda v: "" if pd.isna(v) is True else v)
    return out


def _generar_excel(df_tutor: pd.DataFrame, df_usuarios: pd.DataFrame, codigo: str) -> bytes:
    """Genera bytes de un Excel con pestanas BBDD TUTOR, REPORTES, ASISTENCIA."""
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    # Sanitizar DataFrames para evitar pd.NA que openpyxl no soporta
    df_tutor = _sanitizar_df_para_excel(df_tutor)
    df_usuarios = _sanitizar_df_para_excel(df_usuarios)

    buf = BytesIO()
    wb = Workbook()

    # --- Pestaña BBDD TUTOR ---
    ws_tutor = wb.active
    ws_tutor.title = "BBDD TUTOR"
    # Encabezados en mayúsculas con _ → espacio (como el original)
    df_tutor_exp = df_tutor.copy()
    if "region" in df_tutor_exp.columns and "nombre_iged" in df_tutor_exp.columns:
        df_tutor_exp = df_tutor_exp.sort_values(["region", "nombre_iged"])
    df_tutor_exp.columns = [str(c).upper().replace("_", " ") for c in df_tutor_exp.columns]
    _escribir_df_a_sheet(ws_tutor, df_tutor_exp, "TablaTutor", "TableStyleLight8")

    # --- Pestaña REPORTES ---
    ws_rep = wb.create_sheet("REPORTES")
    reportes = _crear_reportes_dinamicos(df_tutor)
    fila_actual = 1
    estilos = ["TableStyleLight8", "TableStyleLight9", "TableStyleLight10", "TableStyleLight11", "TableStyleLight12"]
    for i, (titulo, df_rep) in enumerate(reportes):
        ws_rep.cell(row=fila_actual, column=1, value=titulo)
        fila_actual += 1
        df_rep = _sanitizar_df_para_excel(df_rep)
        _escribir_df_a_sheet(ws_rep, df_rep, f"Reporte{i + 1}", estilos[i % len(estilos)], start_row=fila_actual)
        fila_actual += len(df_rep) + 3

    # --- Pestaña ASISTENCIA ---
    ws_asis = wb.create_sheet("ASISTENCIA")
    # Agregar tipo_documento a df_usuarios si no existe
    if "tipo_documento" not in df_usuarios.columns:
        df_usuarios["tipo_documento"] = df_usuarios.get("dni", pd.Series(dtype=str)).apply(
            lambda d: "DNI" if str(d).strip().isdigit() and len(str(d).strip()) == 8 else "OTROS"
        )
    cols_asis = [
        "codigo", "tipo_documento", "codigo_iged", "region", "tipo_iged",
        "nombre_iged", "Entidad P\u00fablica a la que pertenece:", "Entidad P\u00fablica",
        "dni", "nombres", "apellidos", "email", "telefono_celular",
        "nivel_puesto", "nombre_puesto",
        "cuestionario entrada", "cuestionario salida",
    ]
    for col_faltante in [c for c in cols_asis if c not in df_usuarios.columns]:
        df_usuarios[col_faltante] = ""
    df_asis = df_usuarios[cols_asis].copy()
    sort_cols = [c for c in ["Entidad Pública a la que pertenece:", "region", "nombre_iged"] if c in df_asis.columns]
    if sort_cols:
        df_asis = df_asis.sort_values(sort_cols, kind="mergesort", na_position="last").reset_index(drop=True)
    # N° numeración
    df_asis.insert(0, "N°", range(1, len(df_asis) + 1))
    # Encabezados UPPER sin _ ni : finales
    encabezados_asis = {
        "codigo": "CODIGO", "tipo_documento": "TIPO DOCUMENTO", "codigo_iged": "CODIGO IGED",
        "region": "REGION", "tipo_iged": "TIPO IGED", "nombre_iged": "NOMBRE IGED",
        "Entidad Pública a la que pertenece:": "ENTIDAD PÚBLICA A LA QUE PERTENECE",
        "Entidad Pública": "ENTIDAD PÚBLICA",
        "dni": "DNI", "nombres": "NOMBRES", "apellidos": "APELLIDOS",
        "email": "EMAIL", "telefono_celular": "TELEFONO CELULAR",
        "nivel_puesto": "NIVEL PUESTO", "nombre_puesto": "NOMBRE PUESTO",
        "cuestionario entrada": "CUESTIONARIO ENTRADA", "cuestionario salida": "CUESTIONARIO SALIDA",
    }
    df_asis = df_asis.rename(columns=encabezados_asis)
    _escribir_df_a_sheet(ws_asis, df_asis, "TablaAsistencia", "TableStyleLight9")

    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _escribir_df_a_sheet(
    ws, df: pd.DataFrame, table_name: str, style: str, start_row: int = 1,
) -> None:
    """Escribe un DataFrame en una hoja openpyxl con formato de tabla."""
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    if df.empty:
        ws.cell(row=start_row, column=1, value="Sin datos")
        return

    # Headers
    for c_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=c_idx, value=str(col_name))

    # Data
    for r_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for c_idx, val in enumerate(row, 1):
            try:
                is_na = pd.isna(val)
            except (ValueError, TypeError):
                is_na = False
            if is_na:
                cell_val = ""
            else:
                cell_val = val
            ws.cell(row=r_idx, column=c_idx, value=cell_val)

    # Table
    end_row = start_row + len(df)
    end_col = get_column_letter(len(df.columns))
    ref = f"A{start_row}:{end_col}{end_row}"
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name=style, showFirstColumn=False, showLastColumn=False, showRowStripes=True)
    ws.add_table(tbl)

    # Ajustar ancho de columnas automaticamente
    for i, col_cells in enumerate(ws.iter_cols(min_row=start_row, max_row=end_row, min_col=1, max_col=len(df.columns)), 1):
        max_length = 0
        col_letter = get_column_letter(i)
        for cell in col_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 100)
        ws.column_dimensions[col_letter].width = adjusted_width


# ---------------------------------------------------------------------------
# Descargar resultado existente
# ---------------------------------------------------------------------------

def obtener_resultado_procesamiento(codigo: str) -> dict[str, Any]:
    """Retorna info del ultimo procesamiento exitoso."""
    storage = _storage_dir(codigo, crear=False)
    run = None

    if storage.is_dir():
        entradas, salidas = _listar_archivos(storage)
        if entradas and salidas:
            try:
                ih = _calcular_hash_insumos(storage, entradas, salidas)
                run = _obtener_run_completado(_cargar_manifest(storage), ih) or _obtener_run_db(codigo, ih)
            except Exception:
                logger.exception("Error leyendo resultado local de %s", codigo)

    if not run:
        run = _obtener_run_db(codigo)

    if not run:
        return {"exists": False}

    return {
        "exists": True,
        "output_path": run.get("output_path", ""),
        "output_file": run.get("output_file", ""),
        "output_content": run.get("output_content", b""),
        "stats": run.get("stats", {}),
        "timestamp": run.get("timestamp", ""),
    }
