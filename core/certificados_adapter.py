"""Adaptador Django para generacion de certificados PDF en memoria.

Replica la logica de Certificados26.py sin dependencias de escritorio (tkinter).
Recibe bytes en memoria y devuelve un ZIP con los PDFs generados.
"""

from __future__ import annotations

import io
import re
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT, TA_JUSTIFY

try:
    from PIL import Image
    PIL_DISPONIBLE = True
except ImportError:
    PIL_DISPONIBLE = False

try:
    import qrcode as _qrcode_mod
    QRCODE_DISPONIBLE = True
except ImportError:
    QRCODE_DISPONIBLE = False


# ---------------------------------------------------------------------------
# Rutas de recursos del servidor
# ---------------------------------------------------------------------------
_RESOURCES_DIR = Path(__file__).resolve().parent / "resources" / "certificados"

FONDO_PATH = _RESOURCES_DIR / "fondo.png"
LOGO_PATH = _RESOURCES_DIR / "logoedutalentos-sinfondo.png"

# Layout del anverso (coincide con Certificados26.py)
Y_BASE_FIRMAS = 0.10 * inch
BOX_H_FIRMAS = 2.15 * inch
PADDING_FECHA = 0.35 * inch


# ---------------------------------------------------------------------------
# Utilidades de texto
# ---------------------------------------------------------------------------

def _obtener_lineas_ajustadas(
    c: canvas.Canvas,
    texto: str,
    max_ancho: float,
    font_name: str = "Helvetica-Bold",
    font_size: float = 18,
) -> list[str]:
    """Envuelve texto por palabras respetando el ancho maximo."""
    palabras = texto.split()
    lineas: list[str] = []
    linea_actual = ""
    for palabra in palabras:
        if c.stringWidth(linea_actual + palabra, font_name, font_size) < max_ancho:
            linea_actual += palabra + " "
        else:
            lineas.append(linea_actual.strip())
            linea_actual = palabra + " "
    if linea_actual:
        lineas.append(linea_actual.strip())
    return lineas


def _obtener_lineas_ajustadas_max(
    c: canvas.Canvas,
    texto: str,
    max_ancho: float,
    font_name: str = "Helvetica",
    font_size: float = 12,
    max_lines: int | None = 2,
) -> list[str]:
    """Envuelve texto respetando saltos de linea, limita a max_lines con truncado."""
    texto = str(texto or "")
    texto = texto.replace("\r\n", "\n").replace("\r", "\n")
    partes = texto.split("\n")
    lineas: list[str] = []
    for parte in partes:
        if parte.strip() == "":
            lineas.append("")
            continue
        lineas.extend(_obtener_lineas_ajustadas(c, parte, max_ancho, font_name, font_size))
    if max_lines is not None and len(lineas) > max_lines:
        lineas = lineas[:max_lines]
        last = lineas[-1].rstrip()
        ell = "…"
        while last and c.stringWidth(last + ell, font_name, font_size) > max_ancho:
            last = last[:-1].rstrip()
        lineas[-1] = (last + ell) if last else ell
    return lineas


def _ajustar_texto_sin_limite(
    c: canvas.Canvas,
    texto: str,
    max_ancho: float,
    font_name: str,
    font_size_max: float,
    font_size_min: float = 9,
) -> tuple[list[str], float]:
    """Reduce fuente hasta que todas las palabras quepan sin truncar (sin límite de líneas)."""
    # Intentar primero con el tamaño exacto (puede ser float como 13.5)
    if font_size_max != int(font_size_max):
        lineas = _obtener_lineas_ajustadas_max(c, texto, max_ancho, font_name, float(font_size_max), max_lines=None)
        if lineas:
            return lineas, float(font_size_max)
    for fs in range(int(font_size_max), int(font_size_min) - 1, -1):
        lineas = _obtener_lineas_ajustadas_max(c, texto, max_ancho, font_name, float(fs), max_lines=None)
        if lineas:
            return lineas, float(fs)
    lineas = _obtener_lineas_ajustadas_max(c, texto, max_ancho, font_name, font_size_min, max_lines=None)
    return lineas, font_size_min


def _ajustar_texto_a_2_lineas(
    c: canvas.Canvas,
    texto: str,
    max_ancho: float,
    font_name: str,
    font_size_max: float,
    font_size_min: float = 9,
) -> tuple[list[str], float]:
    """Reduce fuente hasta que el texto quepa en max 2 lineas."""
    for fs in range(int(font_size_max), int(font_size_min) - 1, -1):
        lineas = _obtener_lineas_ajustadas_max(c, texto, max_ancho, font_name, float(fs), max_lines=None)
        if len(lineas) <= 2:
            return lineas, float(fs)
    lineas = _obtener_lineas_ajustadas_max(c, texto, max_ancho, font_name, font_size_min, max_lines=2)
    return lineas, font_size_min


# ---------------------------------------------------------------------------
# Utilidades de imagen
# ---------------------------------------------------------------------------

def _trim_whitespace(im: Any, white_threshold: int = 245, alpha_threshold: int = 10) -> Any:
    """Recorta margenes blancos/translucidos de una imagen PIL."""
    if not PIL_DISPONIBLE:
        return im
    try:
        im_rgba = im.convert("RGBA")
        pix = im_rgba.load()
        w, h = im_rgba.size
        min_x, min_y, max_x, max_y = w, h, -1, -1
        for y in range(h):
            for x in range(w):
                r, g, b, a = pix[x, y]
                if a > alpha_threshold and (r < white_threshold or g < white_threshold or b < white_threshold):
                    if x < min_x:
                        min_x = x
                    if y < min_y:
                        min_y = y
                    if x > max_x:
                        max_x = x
                    if y > max_y:
                        max_y = y
        if max_x >= 0 and max_y >= 0:
            pad = 6
            min_x = max(0, min_x - pad)
            min_y = max(0, min_y - pad)
            max_x = min(w - 1, max_x + pad)
            max_y = min(h - 1, max_y + pad)
            return im_rgba.crop((min_x, min_y, max_x + 1, max_y + 1))
    except Exception:
        pass
    return im


def _draw_image_fit(
    c: canvas.Canvas,
    img_reader: Any,
    x: float,
    y: float,
    box_w: float,
    box_h: float,
) -> None:
    """Dibuja una imagen manteniendo proporciones dentro de un recuadro."""
    try:
        c.drawImage(img_reader, x, y, box_w, box_h, mask="auto", preserveAspectRatio=True)
    except Exception:
        pass


def _dibujar_bloque_firmas(
    c: canvas.Canvas,
    firmas_readers: list[Any],
    ancho_pagina: float,
    n_firmas_esperadas: int = 1,
    y_base: float = Y_BASE_FIRMAS,
    box_h: float = BOX_H_FIRMAS,
) -> None:
    """Dibuja 1 o 2 firmas centradas en el anverso."""
    firmas_readers = [r for r in (firmas_readers or []) if r is not None]
    if not firmas_readers:
        return

    side_margin = 0.60 * inch
    gap = 0.25 * inch
    firma_scale_2 = 0.85
    firma_scale_1 = 0.68

    n = len(firmas_readers)

    if n == 1:
        reader = firmas_readers[0]
        box_w = min(6.40 * inch, ancho_pagina - 2 * side_margin)
        x = (ancho_pagina - box_w) / 2
        box_w2 = box_w * firma_scale_1
        box_h2 = box_h * firma_scale_1
        x2 = x + (box_w - box_w2) / 2
        y2 = y_base + (box_h - box_h2) / 2
        try:
            c.drawImage(reader, x2, y2, box_w2, box_h2, mask="auto", preserveAspectRatio=True)
        except Exception:
            pass
        return

    # 2 firmas lado a lado
    firmas_readers = firmas_readers[:2]
    box_w = (ancho_pagina - 2 * side_margin - gap) / 2
    box_w2 = box_w * firma_scale_2
    box_h2 = box_h * firma_scale_2
    for i, reader in enumerate(firmas_readers):
        x_cell = side_margin + i * (box_w + gap)
        x_inner = x_cell + (box_w - box_w2) / 2
        y_inner = y_base + (box_h - box_h2) / 2
        try:
            c.drawImage(reader, x_inner, y_inner, box_w2, box_h2, mask="auto", preserveAspectRatio=True)
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Lectura de tabla Excel
# ---------------------------------------------------------------------------

def _leer_tabla_excel(
    wb: openpyxl.Workbook,
    hoja_nombre: str = "Hoja2",
    celda_inicio: str = "A2",
    n_rows: int | None = None,
    n_cols: int | None = 4,
    stop_when_blank_row: bool = True,
    max_scan_rows: int = 200,
    max_scan_cols: int = 20,
) -> dict[str, Any]:
    """Lee tabla desde Hoja2 del workbook ya abierto. Preserva merges, anchos y alineaciones."""
    vacio: dict[str, Any] = {
        "data": None, "excel_spans": [], "has_excel_merges": False,
        "excel_col_widths": [], "excel_halign": [], "excel_valign": [],
    }
    if hoja_nombre not in wb.sheetnames:
        return vacio
    ws = wb[hoja_nombre]

    # Pre-leer rangos combinados
    merged_rects: list[tuple[int, int, int, int, Any]] = []
    try:
        for m in ws.merged_cells.ranges:
            min_c, min_r, max_c, max_r = m.bounds
            val = ws.cell(row=min_r, column=min_c).value
            merged_rects.append((min_r, min_c, max_r, max_c, val))
    except Exception:
        merged_rects = []

    def _eff_val(r: int, c: int) -> Any:
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip():
            return v
        for r1, c1, r2, c2, val in merged_rects:
            if r1 <= r <= r2 and c1 <= c <= c2:
                if val is not None and str(val).strip():
                    return val
        return v

    try:
        col_letter, row_number = coordinate_from_string(celda_inicio)
    except Exception:
        col_letter, row_number = "A", 2
    min_row = int(row_number)
    min_col = int(column_index_from_string(col_letter))

    # Determinar max_col
    if n_cols is not None:
        max_col = min_col + int(n_cols) - 1
    else:
        # Escanea todas las filas para encontrar la columna más a la derecha con contenido real
        max_col = min_col
        scan_rows_for_col = min(max_scan_rows, 60)  # no necesitamos escanear 200 filas para detectar columnas
        for i in range(min_row, min_row + scan_rows_for_col):
            for j in range(min_col, min_col + max_scan_cols):
                v = _eff_val(i, j)
                if v is not None and str(v).strip():
                    if j > max_col:
                        max_col = j
                else:
                    # si hay un hueco en esta fila, no seguir buscando más columnas en ella
                    break

    # Determinar max_row
    last_nonempty = min_row - 1
    scan_to = (min_row + int(n_rows) - 1) if n_rows is not None else (min_row + max_scan_rows - 1)
    for i in range(min_row, scan_to + 1):
        row_vals = [_eff_val(i, j) for j in range(min_col, max_col + 1)]
        is_blank = all(v is None or (isinstance(v, str) and not v.strip()) for v in row_vals)
        if is_blank and stop_when_blank_row:
            break
        if not is_blank:
            last_nonempty = i
    max_row = last_nonempty if last_nonempty >= min_row else min_row

    # Construir spans y mapa
    excel_spans: list[tuple[int, int, int, int]] = []
    merged_cell_map: dict[tuple[int, int], tuple[int, int, Any]] = {}
    for r1, c1, r2, c2, val in merged_rects:
        or0 = max(r1, min_row)
        oc0 = max(c1, min_col)
        or1 = min(r2, max_row)
        oc1 = min(c2, max_col)
        if or0 <= or1 and oc0 <= oc1:
            excel_spans.append((oc0 - min_col, or0 - min_row, oc1 - min_col, or1 - min_row))
            for rr in range(or0, or1 + 1):
                for cc in range(oc0, oc1 + 1):
                    merged_cell_map[(rr, cc)] = (or0, oc0, val)

    # Anchos de columna
    try:
        default_w = ws.sheet_format.defaultColWidth or 8.43
    except Exception:
        default_w = 8.43
    excel_col_widths = []
    for cc in range(min_col, max_col + 1):
        try:
            w = ws.column_dimensions[get_column_letter(cc)].width
            if w is None or float(w) <= 0:
                w = default_w
            excel_col_widths.append(float(w))
        except Exception:
            excel_col_widths.append(float(default_w))

    # Alineaciones
    excel_halign: list[list[str]] = []
    excel_valign: list[list[str]] = []
    for rr in range(min_row, max_row + 1):
        row_h, row_v = [], []
        for cc in range(min_col, max_col + 1):
            rr_use, cc_use = rr, cc
            if (rr, cc) in merged_cell_map:
                or0, oc0, _ = merged_cell_map[(rr, cc)]
                rr_use, cc_use = or0, oc0
            try:
                al = ws.cell(row=rr_use, column=cc_use).alignment
                h = (al.horizontal or "") if al else ""
                v = (al.vertical or "") if al else ""
            except Exception:
                h, v = "", ""
            row_h.append(str(h).lower() if h else "")
            row_v.append(str(v).lower() if v else "")
        excel_halign.append(row_h)
        excel_valign.append(row_v)

    # Leer datos
    reemplazos = {
        "■": "• ", "▪": "• ", "◾": "• ", "▫": "• ", "◽": "• ",
        "—": "- ", "–": "- ", "•": "• ", "‣": "• ", "◦": "• ", "⁃": "• ",
    }
    datos: list[list[str]] = []
    for r in range(min_row, max_row + 1):
        fila = []
        for c in range(min_col, max_col + 1):
            if (r, c) in merged_cell_map:
                or0, oc0, val = merged_cell_map[(r, c)]
                valor = val if r == or0 and c == oc0 else ""
            else:
                valor = ws.cell(row=r, column=c).value
            valor_str = "" if valor is None else str(valor)
            for simbolo, reemplazo in reemplazos.items():
                valor_str = valor_str.replace(simbolo, reemplazo)
            fila.append(valor_str)
        datos.append(fila)

    return {
        "data": datos,
        "excel_spans": excel_spans,
        "has_excel_merges": bool(excel_spans),
        "excel_col_widths": excel_col_widths,
        "excel_halign": excel_halign,
        "excel_valign": excel_valign,
    }


# ---------------------------------------------------------------------------
# Tabla PDF (reverso)
# ---------------------------------------------------------------------------

def _crear_tabla_pdf(
    c: canvas.Canvas,
    datos_tabla: list[list[str]],
    ancho_pagina: float,
    alto_pagina: float,
    excel_spans: list[tuple[int, int, int, int]] | None = None,
    excel_col_widths: list[float] | None = None,
    excel_halign: list[list[str]] | None = None,
    tabla_width_pct: float = 0.92,
) -> float | None:
    """Crea tabla de actividades en el PDF del reverso. Retorna y debajo de la tabla."""
    if not datos_tabla:
        return None

    excel_spans = excel_spans or []

    # Normalizar filas
    filas: list[list[str]] = []
    max_len = 0
    for row in datos_tabla:
        if row is None:
            continue
        r = ["" if v is None else str(v).strip() for v in row]
        if not excel_spans and all(v == "" for v in r):
            continue
        filas.append(r)
        max_len = max(max_len, len(r))

    if not filas or max_len == 0:
        return None

    for i in range(len(filas)):
        if len(filas[i]) < max_len:
            filas[i] = filas[i] + [""] * (max_len - len(filas[i]))

    # Aplicar spans
    spans_ok = []
    for c0, r0, c1, r1 in excel_spans:
        if 0 <= r0 < len(filas) and 0 <= r1 < len(filas) and 0 <= c0 < max_len and 0 <= c1 < max_len:
            spans_ok.append((c0, r0, c1, r1))
            for rr in range(r0, r1 + 1):
                for cc in range(c0, c1 + 1):
                    if not (rr == r0 and cc == c0):
                        filas[rr][cc] = ""
    excel_spans = spans_ok

    # Detectar encabezado
    def _norm(v: str) -> str:
        return re.sub(r"\s+", " ", (v or "").strip())

    def _looks_like_header(row: list[str]) -> bool:
        joined = " ".join(_norm(x).lower() for x in row if _norm(x))
        if not joined:
            return False
        if _norm(row[0]).lower().isdigit():
            return False
        keys = ["actividad", "actividades", "modalidad", "hora", "horas", "contenido",
                "curso", "sesion", "tema", "modulo", "fecha"]
        if any(k in joined for k in keys):
            return True
        nonempty = [_norm(x) for x in row if _norm(x)]
        if not nonempty:
            return False
        short = sum(1 for x in nonempty if len(x) <= 20)
        alpha = sum(1 for x in nonempty if not x.isdigit())
        return (short / len(nonempty) >= 0.7) and (alpha / len(nonempty) >= 0.8)

    has_header = _looks_like_header(filas[0])

    base_fs = 8.0
    head_fs = 8.8

    style_cell_left = ParagraphStyle(
        "cell_left", fontName="Helvetica", fontSize=base_fs,
        leading=base_fs + 1.0, alignment=TA_LEFT, spaceBefore=0, spaceAfter=0,
    )
    style_cell_center = ParagraphStyle("cell_center", parent=style_cell_left, alignment=TA_CENTER)
    style_cell_right = ParagraphStyle("cell_right", parent=style_cell_left, alignment=TA_RIGHT)
    style_cell_justify = ParagraphStyle("cell_justify", parent=style_cell_left, alignment=TA_JUSTIFY)
    style_head_center = ParagraphStyle(
        "head_center", fontName="Helvetica-Bold", fontSize=head_fs,
        leading=head_fs + 1.2, alignment=TA_CENTER, spaceBefore=0, spaceAfter=0,
        textColor=colors.white,
    )
    style_head_left = ParagraphStyle("head_left", parent=style_head_center, alignment=TA_LEFT)

    from xml.sax.saxutils import escape as _xml_escape

    def _map_excel_h(h: str) -> str | None:
        h = (h or "").strip().lower()
        if h in ("center", "centercontinuous", "distributed"):
            return "center"
        if h == "right":
            return "right"
        if h == "justify":
            return "justify"
        if h == "left":
            return "left"
        return None

    def _pick_style(is_header: bool, halign: str | None, raw: str) -> ParagraphStyle:
        if is_header:
            return style_head_left if halign == "left" else style_head_center
        if halign == "center":
            return style_cell_center
        if halign == "right":
            return style_cell_right
        if halign == "justify":
            return style_cell_justify
        if halign == "left":
            return style_cell_left
        if re.fullmatch(r"\d+(\.\d+)?", _norm(str(raw or ""))):
            return style_cell_center
        return style_cell_left

    def _as_para(v: str, is_header: bool = False, halign: str | None = None) -> Paragraph:
        raw = "" if v is None else str(v)
        raw = raw.replace("\r\n", "\n").replace("\r", "\n").strip()
        if is_header:
            raw = _norm(raw)
        s = _xml_escape(raw).replace("\n", "<br/>")
        return Paragraph(s, _pick_style(is_header, halign, raw))

    table_data = []
    for i, row in enumerate(filas):
        is_h = has_header and i == 0
        out_row = []
        for j, v in enumerate(row):
            h = None
            if excel_halign and i < len(excel_halign) and j < len(excel_halign[i]):
                h = _map_excel_h(excel_halign[i][j])
            out_row.append(_as_para(v, is_header=is_h, halign=h))
        table_data.append(out_row)

    # Calcular anchos
    page_max_w = ancho_pagina - 1.6 * inch
    pct = max(0.50, min(1.00, float(tabla_width_pct or 0.92)))
    target_w = page_max_w * pct
    min_col_w = 0.55 * inch

    def _fit_to_width(widths: list[float], desired: float, hard_max: float) -> list[float]:
        if not widths:
            return widths
        total = sum(widths) or 1.0
        desired = min(hard_max, max(min_col_w * len(widths), desired))
        factor = desired / total
        scaled = [max(min_col_w, w * factor) for w in widths]
        total2 = sum(scaled)
        if total2 < desired and total2 > 0:
            extra = desired - total2
            wsum = sum(scaled) or 1.0
            scaled = [w + extra * (w / wsum) for w in scaled]
        return scaled

    if excel_col_widths and len(excel_col_widths) == max_len:
        raw_w = [max(float(w or 8.43), 0.01) for w in excel_col_widths]
        tot = sum(raw_w) or 1.0
        natural = _fit_to_width([(w / tot) * target_w for w in raw_w], target_w, page_max_w)
    else:
        pad = 10
        natural = []
        for col in range(max_len):
            fs = head_fs if has_header else base_fs
            m = 0.0
            for rr in filas:
                t = rr[col] if col < len(rr) else ""
                for part in str(t or "").replace("\r\n", "\n").split("\n"):
                    part = re.sub(r"\s+", " ", part).strip()
                    if part:
                        m = max(m, c.stringWidth(part[:220], "Helvetica", fs))
            w = max(m + pad, min_col_w)
            w = min(w, page_max_w * (0.75 if max_len <= 2 else 0.60))
            natural.append(w)
        natural = _fit_to_width(natural, target_w, page_max_w)

    table_w = sum(natural)
    x_left = (ancho_pagina - table_w) / 2

    styles_list: list[tuple] = [
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]
    if has_header:
        styles_list += [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ]
    data_start = 1 if has_header else 0
    for col in range(max_len):
        vals = [_norm(filas[r][col]) for r in range(data_start, len(filas)) if col < len(filas[r]) and _norm(filas[r][col])]
        if vals and sum(1 for v in vals if re.fullmatch(r"\d+(\.\d+)?", v)) / len(vals) >= 0.7:
            styles_list.append(("ALIGN", (col, data_start), (col, -1), "CENTER"))
    for c0, r0, c1, r1 in excel_spans:
        styles_list.append(("SPAN", (c0, r0), (c1, r1)))
        styles_list.append(("VALIGN", (c0, r0), (c1, r1), "MIDDLE"))

    tabla = Table(table_data, colWidths=natural)
    tabla.setStyle(TableStyle(styles_list))

    y_top = alto_pagina - 2.05 * inch
    y_bottom = 1.65 * inch
    max_height = max(1.0, y_top - y_bottom)

    cur_fs = base_fs
    for _ in range(8):
        _w, h = tabla.wrapOn(c, table_w, max_height)
        if h <= max_height or cur_fs < 6.2:
            break
        cur_fs -= 0.35
        for st in (style_cell_left, style_cell_center, style_cell_right, style_cell_justify):
            st.fontSize = cur_fs
            st.leading = cur_fs + 0.95

    y_draw = y_top - h
    tabla.drawOn(c, x_left, y_draw)
    return max(y_draw - 0.35 * inch, y_bottom)


# ---------------------------------------------------------------------------
# Generar QR en BytesIO
# ---------------------------------------------------------------------------

def _generar_qr_reader(url: str) -> Any | None:
    """Genera un QR y lo retorna como ImageReader de ReportLab."""
    if not QRCODE_DISPONIBLE:
        return None
    try:
        img = _qrcode_mod.make(url)
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        return ImageReader(buf)
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Funcion principal de generacion
# ---------------------------------------------------------------------------

def _clasificar_equipo(nombre_puesto: str) -> str:
    """Mapea el texto libre de `nombre_puesto` a una de las 5 categorias de equipo.

    Categorias (devueltas en Title Case para que encajen en la frase
    "Integrante del ... de la {IGED}" del certificado):
    - Equipo de Directivos y Jefaturas
    - Equipo de Gestión Pedagógica
    - Equipo de Gestión Institucional
    - Equipo Administrativo
    - Equipo Técnico y de Apoyo Especializado  (default y fallback)
    """
    import re
    import unicodedata

    raw = str(nombre_puesto or "").strip()
    if not raw:
        return "Equipo Técnico y de Apoyo Especializado"

    # Normaliza: sin tildes, mayusculas, separadores como espacios.
    norm = "".join(
        ch for ch in unicodedata.normalize("NFD", raw)
        if unicodedata.category(ch) != "Mn"
    ).upper()
    # Convierte separadores comunes en espacios para que las siglas queden tokenizadas.
    norm_tokens = " " + re.sub(r"[^A-Z0-9]+", " ", norm).strip() + " "

    def has_kw(kw: str) -> bool:
        return kw in norm

    def has_token(tok: str) -> bool:
        return f" {tok} " in norm_tokens

    # 1) Directivos y jefaturas
    directivos_kw = (
        "DIRECTOR", "DIRECTORA", "SUBDIRECTOR", "GERENTE", "GERENCIA",
        "JEFE", "JEFA", "JEFATURA", "JEDE", "JEGE", "JEFEDE", "JEJFE",
    )
    if any(has_kw(kw) for kw in directivos_kw):
        return "Equipo de Directivos y Jefaturas"
    if has_token("DIRECCION"):
        return "Equipo de Directivos y Jefaturas"

    # 2) Gestion pedagogica
    pedagogica_kw = (
        "PEDAGOG", "AGEBRE", "AGEBATP", "AGEBA", "AGEBE", "AGEBTH",
        "DOCENTE", "PROFESOR", "PROFESORA", "FORMADOR",
        "ACOMPANANTE", "ACOMPANAMIENTO",
        "TUTORIA", "TUTOR", "ORIENTACION EDUCATIVA", "ORIENTADOR",
        "PSICOLOG", "EDUCUNA", "PRONOEI",
        "CONVIVENCIA",
        "EICE",
        "EDUCACION BASICA", "EDUCACION INICIAL", "EDUCACION PRIMARIA",
        "EDUCACION SECUNDARIA", "EDUCACION FISICA", "EDUCACION ARTISTICA",
        "ESPECIALISTA EN EDUCACION", "ESPECIALISTA DE EDUCACION",
        "ESPECIALISTA PEDAGOG",
        "ASGESE",
        "SUPERVISION Y GESTION DEL SERVICIO EDUCATIVO",
    )
    if any(has_kw(kw) for kw in pedagogica_kw):
        return "Equipo de Gestión Pedagógica"
    if any(has_token(t) for t in ("AGP", "DGP", "UGP", "EBR", "EBA", "EBE", "EBTH", "ICE")):
        return "Equipo de Gestión Pedagógica"

    # 3) Gestion institucional
    institucional_kw = (
        "INSTITUCIONAL",
        "PLANIFICAC", "PLANEAMIENTO", "PLANIFICADOR",
        "PRESUPUESTO", "RACIONALIZA", "MODERNIZACION",
        "ESTADIST", "INVESTIGAC",
        "CALIDAD DE LA INFORMAC", "CALIDAD DE INFORMAC",
        "PREVAED", "COMPROMISOS DE DESEMPENO", "CONVENIO FED",
    )
    if any(has_kw(kw) for kw in institucional_kw):
        return "Equipo de Gestión Institucional"
    if any(has_token(t) for t in ("AGI", "DGI", "UGI", "POI", "UPDI")):
        return "Equipo de Gestión Institucional"

    # 4) Administrativo
    administrativo_kw = (
        "ADMINISTRATIV", "ADMINISTRACION", "ADMINISTRADOR",
        "ABASTECIMIENTO", "LOGISTIC",
        "TESORER", "CONTABIL", "PATRIMONI", "ALMACEN",
        "SECRETARI", "OFICINISTA",
        "PERSONAL", "RECURSOS HUMANOS",
        "ESCALAFON", "REMUNERACION", "PENSION", "PLANILLA",
        "TRAMITE", "MESA DE PARTES", "ARCHIVO", "NOTIFIC",
        "ASESORIA JURIDIC", "JURIDIC", "ABOGAD", "LEGAL",
        "COMUNICAC", "IMAGEN", "RELACIONES PUB", "RELACIONISTA",
        "INTEGRIDAD", "CONTROL INTERNO", "PROCESOS ADMIN",
        "BIENESTAR",
        "ATENCION AL USUARIO", "QUEJAS", "RECLAMOS", "TRANSPARENC",
        "ACTAS Y CERTIFICADOS", "CONSTANCIA DE PAGOS",
        "FINANZAS", "FINANCIST", "CONTADOR", "AUDITOR",
        "INSPECTORIA", "INSPECCION",
        "ADQUISICIONES", "CONTRATACION",
    )
    if any(has_kw(kw) for kw in administrativo_kw):
        return "Equipo Administrativo"
    if any(has_token(t) for t in (
        "AGA", "DGA", "OGA", "RRHH", "ORRHH", "ARH", "EAP", "EARH",
        "AAJ", "OEAJ", "OCI", "PAD", "CPPADD", "COPROA",
    )):
        return "Equipo Administrativo"

    # 5) Default / fallback: tecnico y apoyo especializado
    return "Equipo Técnico y de Apoyo Especializado"


def generar_certificados_zip(
    params: dict[str, Any],
    excel_bytes: bytes,
    firma_bytes_list: list[bytes],
    progress_callback: Callable[[int, int, int, int, int, str], None] | None = None,
    participantes_rows: list[dict[str, Any]] | None = None,
) -> tuple[io.BytesIO, int, list[str]]:
    """Genera un ZIP en memoria con los PDFs de certificados.

    Args:
        params: dict con 'curso_nombre', 'curso_descripcion', 'curso_codigo',
                'n_firmas' (int, 1 o 2), 'tabla_width_pct' (float, 0.5-1.0).
        excel_bytes: contenido del archivo .xlsx de tabla de actividades (Hoja2).
        firma_bytes_list: lista de bytes de las imagenes de firma (1 o 2).
        participantes_rows: lista opcional de participantes ya filtrados desde BD.

    Returns:
        (zip_bytesio, n_certificados, lista_errores)
    """
    curso_nombre = str(params.get("curso_nombre", "")).strip()
    curso_descripcion = str(params.get("curso_descripcion", "")).strip()
    curso_codigo = str(params.get("curso_codigo", "")).strip()
    n_firmas = int(params.get("n_firmas", 1) or 1)
    tabla_width_pct = float(params.get("tabla_width_pct", 0.92) or 0.92)

    errores: list[str] = []

    # Construir universo de participantes.
    columnas_requeridas = {"DNI", "NOMBRES", "APELLIDOS", "NOTAS"}
    if participantes_rows is None:
        excel_buf_participantes = io.BytesIO(excel_bytes)
        try:
            df_alumnos = pd.read_excel(excel_buf_participantes, dtype={"DNI": str}, engine="openpyxl")
        except Exception as exc:
            return io.BytesIO(), 0, [f"No se pudo leer el archivo Excel: {exc}"]
        faltantes_cols = columnas_requeridas - set(df_alumnos.columns)
        if faltantes_cols:
            return io.BytesIO(), 0, [f"El Excel no tiene las columnas requeridas: {', '.join(sorted(faltantes_cols))}"]
    else:
        df_alumnos = pd.DataFrame(participantes_rows)
        faltantes_cols = columnas_requeridas - set(df_alumnos.columns)
        if faltantes_cols:
            return io.BytesIO(), 0, [f"Faltan columnas en participantes automáticos: {', '.join(sorted(faltantes_cols))}"]
        if df_alumnos.empty:
            return io.BytesIO(), 0, ["No se encontraron participantes con los filtros requeridos."]
        # Normaliza DNI a texto para conservar ceros a la izquierda en casos especiales.
        df_alumnos["DNI"] = df_alumnos["DNI"].astype(str)

    # Leer tabla de actividades con openpyxl (opcional: si excel_bytes está vacío, no hay tabla)
    datos_tabla = None
    tabla_spans: list = []
    tabla_col_widths: list = []
    tabla_halign: list = []
    if excel_bytes:
        excel_buf = io.BytesIO(excel_bytes)
        try:
            wb = openpyxl.load_workbook(excel_buf, data_only=True)
            res_tabla = _leer_tabla_excel(wb, "Hoja2", "A2", n_cols=None, n_rows=200)
            if not res_tabla.get("data"):
                res_tabla = _leer_tabla_excel(wb, "Hoja2", "B2", n_cols=None, n_rows=200)
            wb.close()
            datos_tabla = res_tabla.get("data")
            tabla_spans = res_tabla.get("excel_spans", [])
            tabla_col_widths = res_tabla.get("excel_col_widths", [])
            tabla_halign = res_tabla.get("excel_halign", [])
        except Exception as exc:
            errores.append(f"No se pudo leer la tabla del reverso: {exc}")

    # Cargar recursos del servidor
    fondo_reader = ImageReader(str(FONDO_PATH)) if FONDO_PATH.exists() else None
    logo_reader = ImageReader(str(LOGO_PATH)) if LOGO_PATH.exists() else None

    # Preparar ImageReaders de firmas
    firmas_readers: list[Any] = []
    for fb in (firma_bytes_list or [])[:2]:
        if not fb:
            continue
        try:
            firmas_readers.append(ImageReader(io.BytesIO(fb)))
        except Exception:
            pass

    # Fecha de emision
    ahora = datetime.now()
    meses = [
        "enero", "febrero", "marzo", "abril", "mayo", "junio",
        "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
    ]
    fecha_str = f"{ahora.day} de {meses[ahora.month - 1]} de {ahora.year}"

    zip_buf = io.BytesIO()
    n_ok = 0
    n_omitidos = 0
    total_filas = int(len(df_alumnos.index))
    procesadas = 0

    def _reportar_progreso(done: int, total: int, etapa: str) -> None:
        if not progress_callback:
            return
        try:
            progress_callback(done, total, n_ok, len(errores), n_omitidos, etapa)
        except Exception:
            # El callback es solo informativo para la UI.
            pass

    _reportar_progreso(0, total_filas, "iniciando")

    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for _, row in df_alumnos.iterrows():
            try:
                dni = str(row.get("DNI", "")).strip()
                nombres = str(row.get("NOMBRES", "")).strip()
                apellidos = str(row.get("APELLIDOS", "")).strip()
                notas = row.get("NOTAS", "")
                puesto = str(row.get("NOMBRE DE PUESTO", "")).strip()
                equipo = _clasificar_equipo(puesto)
                iged = str(row.get("NOMBRE IGED", "")).strip().upper()

                if not dni or not nombres:
                    n_omitidos += 1
                    continue

                pdf_buf = io.BytesIO()
                c = canvas.Canvas(pdf_buf, pagesize=landscape(A4))
                ancho, alto = landscape(A4)

                # ── PÁGINA 1: ANVERSO ──────────────────────────────────────
                c.setFillColorRGB(1, 1, 1)
                c.rect(0, 0, ancho, alto, fill=1)

                if fondo_reader:
                    try:
                        c.drawImage(fondo_reader, ancho * 0.6, 0, ancho * 0.4, alto, mask="auto")
                    except Exception:
                        pass
                if logo_reader:
                    try:
                        c.drawImage(logo_reader, 30, alto - 0.65 * inch - 30, 2.25 * inch, 0.65 * inch, mask="auto")
                    except Exception:
                        pass

                y = alto - 2 * inch
                lh = 0.6 * inch

                c.setFont("Helvetica", 12)
                c.setFillColor(colors.black)
                c.drawCentredString(ancho / 2, y,
                    "El Ministerio de Educación, a través de la Dirección de Fortalecimiento de Capacidades,")
                y -= lh
                c.drawCentredString(ancho / 2, y, "certifica que el (la) señor (a):")
                y -= lh

                c.setFont("Helvetica-Bold", 24)
                c.drawCentredString(ancho / 2, y, f"{nombres} {apellidos}")
                y -= lh

                c.setFont("Helvetica", 12)
                c.drawCentredString(ancho / 2, y, f"Integrante del {equipo} de la {iged}, culminó satisfactoriamente el")
                y -= lh

                # Ajuste dinamico solicitado:
                # - 1 linea: aumentar hasta 2x
                # - 2 lineas: aumentar hasta 1.5x
                lineas_curso_base, fs_curso_base = _ajustar_texto_sin_limite(c, curso_nombre, 600, "Helvetica-Bold", 13.5, 8)
                n_lineas_curso = len(lineas_curso_base)
                if n_lineas_curso <= 1:
                    max_fs_curso = fs_curso_base * 2.0
                elif n_lineas_curso == 2:
                    max_fs_curso = fs_curso_base * 1.5
                else:
                    max_fs_curso = fs_curso_base
                lineas_curso, fs_curso = _ajustar_texto_sin_limite(c, curso_nombre, 600, "Helvetica-Bold", max_fs_curso, 8)
                c.setFont("Helvetica-Bold", fs_curso)
                c.setFillColor(colors.red)
                gap_curso = fs_curso * 1.30
                for linea in lineas_curso:
                    c.drawCentredString(ancho / 2, y, linea)
                    y -= gap_curso
                y -= lh * 0.3  # pequeño espacio tras el bloque del nombre

                lineas_desc, fs_desc = _ajustar_texto_a_2_lineas(c, curso_descripcion, 650, "Helvetica", 12, 10)
                c.setFont("Helvetica", fs_desc)
                c.setFillColor(colors.black)
                gap_desc = fs_desc * 1.30
                for linea in lineas_desc:
                    c.drawCentredString(ancho / 2, y, linea)
                    y -= gap_desc
                y -= lh * 0.3

                c.setFont("Helvetica", 11)
                c.setFillColor(colors.black)
                c.drawRightString(ancho - 0.6 * inch, y, f"Fecha de emisión: {fecha_str}")

                _dibujar_bloque_firmas(
                    c, firmas_readers, ancho,
                    n_firmas_esperadas=n_firmas,
                    y_base=Y_BASE_FIRMAS, box_h=BOX_H_FIRMAS,
                )
                c.showPage()

                # ── PÁGINA 2: REVERSO ──────────────────────────────────────
                y_titulo = alto - 1.2 * inch
                lineas_titulo = _obtener_lineas_ajustadas(c, curso_nombre.upper(), 600, "Helvetica-Bold", 14)
                c.setFont("Helvetica-Bold", 14)
                c.setFillColor(colors.black)
                for i, linea in enumerate(lineas_titulo):
                    c.drawCentredString(ancho / 2, y_titulo, linea)
                    y_titulo -= 0.2 * inch if i < len(lineas_titulo) - 1 else 0.4 * inch

                y_promedio: float | None = None
                if datos_tabla:
                    y_promedio = _crear_tabla_pdf(
                        c, datos_tabla, ancho, alto,
                        excel_spans=tabla_spans,
                        excel_col_widths=tabla_col_widths,
                        excel_halign=tabla_halign,
                        tabla_width_pct=tabla_width_pct,
                    )

                # Muestra promedio sin decimales, redondeado al entero mas cercano.
                try:
                    promedio_texto = str(int(round(float(notas))))
                except Exception:
                    promedio_texto = str(notas).strip()

                c.setFont("Helvetica-Bold", 12)
                if y_promedio:
                    c.drawCentredString(ancho / 2, y_promedio, f"Promedio Final: {promedio_texto}")
                else:
                    c.drawString(inch, alto - 2.5 * inch, f"Promedio Final: {promedio_texto}")

                qr_url = f"https://difoca.com/consultas-cyc/qr/?d={dni}&c={curso_codigo}"
                qr_reader = _generar_qr_reader(qr_url)
                qr_size = 1.5 * inch
                qr_x = ancho - qr_size - 0.25 * inch
                qr_y = 0.25 * inch
                if qr_reader:
                    c.setFont("Helvetica", 7)
                    c.drawCentredString(qr_x + qr_size / 2, qr_y + qr_size + 1.5, "Verificar legitimidad en:")
                    c.drawImage(qr_reader, qr_x, qr_y, qr_size, qr_size, mask="auto")

                c.setFont("Helvetica", 10)
                c.drawCentredString(ancho / 2, alto - 7.5 * inch, f"Código del certificado: DIFOCA-{dni}-{curso_codigo}")

                c.save()

                pdf_bytes = pdf_buf.getvalue()
                nombre_archivo = f"DIFOCA-{dni}-{curso_codigo}.pdf"
                zf.writestr(nombre_archivo, pdf_bytes)
                n_ok += 1

            except Exception as exc:
                errores.append(f"Error en DNI {row.get('DNI', '?')}: {exc}")
            finally:
                procesadas += 1
                _reportar_progreso(procesadas, total_filas, "generando")

    _reportar_progreso(total_filas, total_filas, "empaquetando")
    zip_buf.seek(0)
    _reportar_progreso(total_filas, total_filas, "completado")
    return zip_buf, n_ok, errores


def validar_excel_certificados(excel_bytes: bytes) -> tuple[bool, str]:
    """Verifica que el Excel tenga las columnas requeridas en Hoja1."""
    try:
        buf = io.BytesIO(excel_bytes)
        df = pd.read_excel(buf, dtype={"DNI": str}, engine="openpyxl", nrows=1)
        requeridas = {"DNI", "NOMBRES", "APELLIDOS", "NOTAS"}
        faltantes = requeridas - set(df.columns)
        if faltantes:
            return False, f"Faltan columnas: {', '.join(sorted(faltantes))}"
        return True, ""
    except Exception as exc:
        return False, f"No se pudo leer el archivo: {exc}"
